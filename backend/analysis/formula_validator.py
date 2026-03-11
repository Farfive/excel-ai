"""
Formula Validation Engine
Checks formula consistency patterns across ranges — catches model errors
like inconsistent SUM ranges, broken references, mismatched patterns.
"""
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Set

from parser.xlsx_parser import WorkbookData

logger = logging.getLogger(__name__)


@dataclass
class FormulaIssue:
    severity: str  # "error", "warning", "info"
    category: str
    cell: str
    sheet: str
    message: str
    expected: Optional[str] = None
    actual: Optional[str] = None


@dataclass
class FormulaValidationReport:
    total_issues: int
    errors: int
    warnings: int
    info: int
    issues: List[FormulaIssue]
    patterns_checked: int
    formula_count: int
    consistency_score: float  # 0-100


class FormulaValidator:
    def __init__(self, workbook_data: WorkbookData) -> None:
        self.wb = workbook_data

    def run(self) -> FormulaValidationReport:
        issues: List[FormulaIssue] = []

        issues.extend(self._check_column_consistency())
        issues.extend(self._check_row_consistency())
        issues.extend(self._check_hardcoded_in_formula_columns())
        issues.extend(self._check_circular_text_refs())
        issues.extend(self._check_empty_formula_gaps())

        formula_cells = [c for c in self.wb.cells.values() if c.formula]
        formula_count = len(formula_cells)

        errors = sum(1 for i in issues if i.severity == "error")
        warnings = sum(1 for i in issues if i.severity == "warning")
        info = sum(1 for i in issues if i.severity == "info")

        consistency = 100.0
        if formula_count > 0:
            consistency = max(0.0, 100.0 - (errors * 5.0 + warnings * 2.0 + info * 0.5))

        return FormulaValidationReport(
            total_issues=len(issues),
            errors=errors,
            warnings=warnings,
            info=info,
            issues=issues,
            patterns_checked=5,
            formula_count=formula_count,
            consistency_score=round(consistency, 1),
        )

    def _check_column_consistency(self) -> List[FormulaIssue]:
        """Check that formulas in the same column follow the same pattern."""
        issues: List[FormulaIssue] = []

        by_sheet_col: Dict[str, Dict[int, List]] = defaultdict(lambda: defaultdict(list))
        for c in self.wb.cells.values():
            if c.formula:
                by_sheet_col[c.sheet_name][c.col].append(c)

        for sheet, cols in by_sheet_col.items():
            for col_idx, cells in cols.items():
                if len(cells) < 3:
                    continue

                patterns = []
                for c in sorted(cells, key=lambda x: x.row):
                    pattern = self._normalize_formula(c.formula, c.row, c.col)
                    patterns.append((c, pattern))

                pattern_counts: Dict[str, List] = defaultdict(list)
                for c, p in patterns:
                    pattern_counts[p].append(c)

                if len(pattern_counts) <= 1:
                    continue

                dominant_pattern = max(pattern_counts.items(), key=lambda x: len(x[1]))
                dominant_p, dominant_cells = dominant_pattern

                for pattern, cells_with_pattern in pattern_counts.items():
                    if pattern == dominant_p:
                        continue
                    for c in cells_with_pattern:
                        issues.append(FormulaIssue(
                            severity="warning",
                            category="column_inconsistency",
                            cell=c.cell_address,
                            sheet=sheet,
                            message=f"Formula pattern differs from {len(dominant_cells)} other cells in this column",
                            expected=dominant_p,
                            actual=pattern,
                        ))

        return issues

    def _check_row_consistency(self) -> List[FormulaIssue]:
        """Check that formulas in the same row follow the same pattern (shifted by column)."""
        issues: List[FormulaIssue] = []

        by_sheet_row: Dict[str, Dict[int, List]] = defaultdict(lambda: defaultdict(list))
        for c in self.wb.cells.values():
            if c.formula:
                by_sheet_row[c.sheet_name][c.row].append(c)

        for sheet, rows in by_sheet_row.items():
            for row_idx, cells in rows.items():
                if len(cells) < 3:
                    continue

                patterns = []
                for c in sorted(cells, key=lambda x: x.col):
                    pattern = self._normalize_formula(c.formula, c.row, c.col, by_col=True)
                    patterns.append((c, pattern))

                pattern_counts: Dict[str, List] = defaultdict(list)
                for c, p in patterns:
                    pattern_counts[p].append(c)

                if len(pattern_counts) <= 1:
                    continue

                dominant_pattern = max(pattern_counts.items(), key=lambda x: len(x[1]))
                dominant_p, dominant_cells = dominant_pattern

                for pattern, cells_with_pattern in pattern_counts.items():
                    if pattern == dominant_p:
                        continue
                    for c in cells_with_pattern:
                        issues.append(FormulaIssue(
                            severity="warning",
                            category="row_inconsistency",
                            cell=c.cell_address,
                            sheet=sheet,
                            message=f"Formula pattern differs from {len(dominant_cells)} other cells in this row",
                            expected=dominant_p,
                            actual=pattern,
                        ))

        return issues

    def _check_hardcoded_in_formula_columns(self) -> List[FormulaIssue]:
        """Detect hardcoded values in columns that are predominantly formulas."""
        issues: List[FormulaIssue] = []

        by_sheet_col: Dict[str, Dict[int, List]] = defaultdict(lambda: defaultdict(list))
        for c in self.wb.cells.values():
            if c.value is not None or c.formula:
                by_sheet_col[c.sheet_name][c.col].append(c)

        for sheet, cols in by_sheet_col.items():
            for col_idx, cells in cols.items():
                if len(cells) < 5:
                    continue
                formula_cells = [c for c in cells if c.formula]
                hardcoded_num = [c for c in cells if not c.formula and isinstance(c.value, (int, float)) and c.row > 1]

                if len(formula_cells) < len(cells) * 0.6:
                    continue

                for c in hardcoded_num:
                    issues.append(FormulaIssue(
                        severity="info",
                        category="hardcoded_in_formula_column",
                        cell=c.cell_address,
                        sheet=sheet,
                        message=f"Hardcoded numeric value ({c.value}) in a column that is {round(len(formula_cells)/len(cells)*100)}% formulas",
                    ))

        return issues

    def _check_circular_text_refs(self) -> List[FormulaIssue]:
        """Check for formulas that reference themselves."""
        issues: List[FormulaIssue] = []
        for c in self.wb.cells.values():
            if not c.formula:
                continue
            short_addr = c.cell_address.split("!")[-1] if "!" in c.cell_address else c.cell_address
            if short_addr in c.formula:
                issues.append(FormulaIssue(
                    severity="error",
                    category="self_reference",
                    cell=c.cell_address,
                    sheet=c.sheet_name,
                    message=f"Formula may reference itself: ={c.formula}",
                ))
        return issues

    def _check_empty_formula_gaps(self) -> List[FormulaIssue]:
        """Detect gaps in formula sequences (e.g. row 5 has formula, row 6 empty, row 7 has formula)."""
        issues: List[FormulaIssue] = []

        by_sheet_col: Dict[str, Dict[int, List]] = defaultdict(lambda: defaultdict(list))
        for c in self.wb.cells.values():
            if c.formula:
                by_sheet_col[c.sheet_name][c.col].append(c)

        for sheet, cols in by_sheet_col.items():
            for col_idx, cells in cols.items():
                if len(cells) < 3:
                    continue
                rows_with_formula = sorted(c.row for c in cells)
                for i in range(len(rows_with_formula) - 1):
                    gap = rows_with_formula[i + 1] - rows_with_formula[i]
                    if gap == 2:
                        missing_row = rows_with_formula[i] + 1
                        from openpyxl.utils import get_column_letter
                        missing_addr = f"{sheet}!{get_column_letter(col_idx)}{missing_row}"
                        if missing_addr not in self.wb.cells:
                            issues.append(FormulaIssue(
                                severity="info",
                                category="formula_gap",
                                cell=missing_addr,
                                sheet=sheet,
                                message=f"Empty cell in a sequence of formulas (rows {rows_with_formula[i]} to {rows_with_formula[i+1]})",
                            ))

        return issues

    @staticmethod
    def _normalize_formula(formula: str, row: int, col: int, by_col: bool = False) -> str:
        """Replace row numbers (or col letters) with placeholders to get a structural pattern."""
        if by_col:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            return re.sub(r'(?<![A-Z])' + col_letter + r'(?=\d)', 'COL', formula, flags=re.IGNORECASE)
        else:
            return re.sub(str(row), 'ROW', formula)
