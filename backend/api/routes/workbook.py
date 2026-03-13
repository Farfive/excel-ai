import json
import logging
import time
import uuid
from typing import Any, Dict

from fastapi import APIRouter, Body, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
import aiofiles
import tempfile
import os

from api.dependencies import (
    get_embedder, get_chroma, get_ollama, get_lsh,
    get_workbook_graphs, get_workbook_data_cache, get_workbook_states,
    get_audit_trails, get_cgasr_indices,
)
from api.models import (
    AskRequest, DeltaRequest, StateRequest, StateResponse,
    UploadResponse, AnomaliesResponse, AnomalyItem, ExplainResponse, DeltaResponse,
)
from parser.xlsx_parser import XLSXParser
from parser.graph_builder import DependencyGraphBuilder
from rag.chunker import ChunkBuilder
from rag.chroma_store import ChunkRecord
from rag.retrieval import RAGRetriever
from rag.cgasr_index import build_cgasr_index
from agent.excel_agent import ExcelAgent
from agent.tools import ExcelTools
from db.connection import get_db, WorkbookSession
from sqlalchemy.ext.asyncio import AsyncSession

router = APIRouter(prefix="/workbook", tags=["workbook"])
logger = logging.getLogger(__name__)


@router.post("/upload", response_model=UploadResponse)
async def upload_workbook(
    file: UploadFile = File(...),
    embedder=Depends(get_embedder),
    chroma=Depends(get_chroma),
    lsh=Depends(get_lsh),
    graphs=Depends(get_workbook_graphs),
    data_cache=Depends(get_workbook_data_cache),
    cgasr_store=Depends(get_cgasr_indices),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    start = time.time()
    workbook_uuid = str(uuid.uuid4())

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        parser = XLSXParser()
        workbook_data = parser.parse(tmp_path)

        builder = DependencyGraphBuilder()
        graph = builder.build(workbook_data)
        graph = builder.run_algorithms(graph, workbook_data)

        graphs[workbook_uuid] = graph
        data_cache[workbook_uuid] = workbook_data

        chunk_builder = ChunkBuilder()
        chunks = chunk_builder.build_chunks(workbook_data, graph, workbook_uuid)

        texts = [c.text for c in chunks]
        embeddings = embedder.embed(texts)

        chunk_records = []
        for i, chunk in enumerate(chunks):
            chunk.embedding = embeddings[i]
            lsh.add(chunk.chunk_id, embeddings[i])
            chunk_records.append(
                ChunkRecord(
                    chunk_id=chunk.chunk_id,
                    text=chunk.text,
                    embedding=embeddings[i],
                    metadata=chunk.metadata,
                )
            )

        chroma.upsert_chunks(workbook_uuid, chunk_records)

        # Build BM25 keyword index for hybrid search
        try:
            _retriever = RAGRetriever(embedder=embedder, store=chroma, ollama_client=None, graph=graph)
            _retriever.build_bm25_index(workbook_uuid)
        except Exception as e:
            logger.warning(f"BM25 index build skipped: {e}")

        # Build CGASR spectral index
        try:
            import numpy as np
            emb_matrix = np.array(embeddings, dtype=np.float32)
            cgasr_idx = build_cgasr_index(
                workbook_data, graph, existing_embeddings=emb_matrix,
            )
            cgasr_store[workbook_uuid] = cgasr_idx
            logger.info(f"CGASR index built: N={cgasr_idx.N}, K={cgasr_idx.K}, {cgasr_idx.build_time_ms}ms")
        except Exception as e:
            logger.warning(f"CGASR index build failed (retrieval will use fallback): {e}")

        anomaly_count = sum(
            1 for n in graph.nodes() if graph.nodes[n].get("is_anomaly")
        )

        cluster_ids = set(graph.nodes[n].get("cluster_id", 0) for n in graph.nodes())
        cluster_count = len(cluster_ids)
        cell_count = len(workbook_data.cells)

        session_obj = WorkbookSession(
            workbook_uuid=workbook_uuid,
            filename=file.filename,
            cell_count=cell_count,
            cluster_count=cluster_count,
            anomaly_count=anomaly_count,
        )
        db.add(session_obj)

        elapsed_ms = int((time.time() - start) * 1000)
        logger.info(f"Upload complete: {workbook_uuid} in {elapsed_ms}ms")

        return UploadResponse(
            workbook_uuid=workbook_uuid,
            cell_count=cell_count,
            cluster_count=cluster_count,
            anomaly_count=anomaly_count,
            processing_time_ms=elapsed_ms,
        )
    finally:
        os.unlink(tmp_path)


@router.post("/{uuid}/ask")
async def ask(
    uuid: str,
    request: AskRequest,
    embedder=Depends(get_embedder),
    chroma=Depends(get_chroma),
    ollama_client=Depends(get_ollama),
    graphs=Depends(get_workbook_graphs),
    data_cache=Depends(get_workbook_data_cache),
    states=Depends(get_workbook_states),
    audit_trails=Depends(get_audit_trails),
    cgasr_store=Depends(get_cgasr_indices),
):
    if uuid not in graphs:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found. Upload first.")

    graph = graphs[uuid]
    workbook_data = data_cache[uuid]
    if uuid not in states:
        states[uuid] = {}
    workbook_state = states[uuid]

    from analysis.audit_trail import AuditTrail
    if uuid not in audit_trails:
        audit_trails[uuid] = AuditTrail()
    trail = audit_trails[uuid]

    tools = ExcelTools(workbook_state=workbook_state, graph=graph, workbook_data=workbook_data)
    cgasr_idx = cgasr_store.get(uuid)
    retriever = RAGRetriever(
        embedder=embedder, store=chroma, ollama_client=ollama_client,
        graph=graph, cgasr_index=cgasr_idx,
    )
    retriever.build_bm25_index(uuid)
    agent = ExcelAgent(ollama=ollama_client, retriever=retriever, tools=tools, workbook_data=workbook_data)

    async def event_generator():
        try:
            async for event in agent.execute(
                query=request.question,
                workbook_uuid=uuid,
                approved_plan=request.approved_plan,
                mode=request.mode,
            ):
                # Record diffs into the audit trail
                if event.get("event") == "tool_result":
                    edata = event.get("data", {})
                    result_data = edata.get("data") or {}
                    diff_list = result_data.get("diff", []) if isinstance(result_data, dict) else []
                    tool_name = edata.get("tool", "")
                    reason = edata.get("args", {}).get("reason", "") if isinstance(edata.get("args"), dict) else request.question
                    for d in diff_list:
                        cell_addr = d.get("cell", "")
                        sheet_name = cell_addr.split("!")[0] if "!" in cell_addr else ""
                        trail.record_change(
                            cell=cell_addr, sheet=sheet_name,
                            old_value=d.get("before"), new_value=d.get("after"),
                            old_formula=None, new_formula=None,
                            reason=f"{tool_name}: {request.question[:60]}",
                            agent_step=edata.get("step"),
                            approved_by="agent",
                        )
                yield f"data: {json.dumps(event, default=str)}\n\n"
        except Exception as e:
            logger.error(f"Agent execution error: {e}")
            yield f"data: {json.dumps({'event': 'error', 'data': {'message': str(e)}})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/{uuid}/audit-trail")
async def get_audit_trail(
    uuid: str,
    audit_trails=Depends(get_audit_trails),
):
    """Return the full audit trail of all AI changes for this workbook."""
    trail = audit_trails.get(uuid)
    if not trail:
        return {"changes": [], "total": 0}

    return {
        "changes": [
            {
                "id": c.change_id,
                "timestamp": c.timestamp,
                "cell": c.cell,
                "sheet": c.sheet,
                "old_value": c.old_value,
                "new_value": c.new_value,
                "reason": c.reason,
                "approved_by": c.approved_by,
                "step": c.agent_step,
            }
            for c in trail.changes
        ],
        "total": len(trail.changes),
    }


@router.get("/{uuid}/summary")
async def get_summary(
    uuid: str,
    data_cache=Depends(get_workbook_data_cache),
    graphs=Depends(get_workbook_graphs),
):
    """Return a comprehensive workbook summary with per-sheet stats, key metrics, quality score, and cross-sheet deps."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    workbook_data = data_cache[uuid]
    graph = graphs.get(uuid)

    total_cells = 0
    total_formulas = 0
    total_empty = 0
    total_anomalies = 0
    sheets_summary = []

    for sheet_name in workbook_data.sheets:
        sheet_cells = [c for c in workbook_data.cells.values() if c.sheet_name == sheet_name]
        cell_count = len(sheet_cells)
        total_cells += cell_count

        formula_count = sum(1 for c in sheet_cells if c.formula)
        total_formulas += formula_count

        empty_count = sum(1 for c in sheet_cells if c.value is None and not c.formula)
        total_empty += empty_count

        number_count = sum(1 for c in sheet_cells if isinstance(c.value, (int, float)) and not c.formula)
        text_count = sum(1 for c in sheet_cells if isinstance(c.value, str))

        max_row = max((c.row for c in sheet_cells), default=0)
        max_col = max((c.col for c in sheet_cells), default=0)

        anomaly_count = 0
        if graph:
            for c in sheet_cells:
                addr = c.cell_address
                if addr in graph.nodes and graph.nodes[addr].get("is_anomaly"):
                    anomaly_count += 1
        total_anomalies += anomaly_count

        key_metrics = []
        for c in sheet_cells:
            if c.col == 1 and isinstance(c.value, str):
                label = c.value.strip()
                label_lower = label.lower()
                is_key = any(kw in label_lower for kw in [
                    'total', 'revenue', 'income', 'ebitda', 'wacc', 'npv', 'irr',
                    'net', 'gross', 'margin', 'growth', 'capex', 'debt', 'equity',
                    'cost', 'expense', 'profit', 'cash flow', 'fcf',
                ])
                if is_key:
                    from openpyxl.utils import get_column_letter
                    for vc in range(2, min(max_col + 1, 5)):
                        val_cell = workbook_data.cells.get(f"{sheet_name}!{get_column_letter(vc)}{c.row}")
                        if val_cell and val_cell.value is not None:
                            key_metrics.append({
                                "label": label,
                                "cell": f"{get_column_letter(vc)}{c.row}",
                                "value": val_cell.value,
                                "has_formula": bool(val_cell.formula),
                            })
                            break
                    if len(key_metrics) >= 8:
                        break

        sheets_summary.append({
            "name": sheet_name,
            "rows": max_row,
            "cols": max_col,
            "cell_count": cell_count,
            "formula_count": formula_count,
            "number_count": number_count,
            "text_count": text_count,
            "empty_count": empty_count,
            "anomaly_count": anomaly_count,
            "key_metrics": key_metrics,
        })

    formula_pct = round((total_formulas / total_cells * 100) if total_cells > 0 else 0, 1)
    empty_pct = round((total_empty / total_cells * 100) if total_cells > 0 else 0, 1)
    anomaly_pct = round((total_anomalies / total_cells * 100) if total_cells > 0 else 0, 1)

    quality_score = 100
    if empty_pct > 30:
        quality_score -= 20
    elif empty_pct > 15:
        quality_score -= 10
    if formula_pct < 10:
        quality_score -= 15
    if anomaly_pct > 5:
        quality_score -= 20
    elif anomaly_pct > 2:
        quality_score -= 10
    if len(workbook_data.named_ranges) == 0:
        quality_score -= 5
    quality_score = max(0, quality_score)

    cross_sheet_deps = []
    if graph:
        seen_deps = set()
        for node in graph.nodes():
            if "!" not in node:
                continue
            src_sheet = node.split("!")[0]
            for succ in graph.successors(node):
                if "!" in succ:
                    dst_sheet = succ.split("!")[0]
                    if src_sheet != dst_sheet:
                        pair = (src_sheet, dst_sheet)
                        if pair not in seen_deps:
                            seen_deps.add(pair)
                            cross_sheet_deps.append({"from": src_sheet, "to": dst_sheet})

    return {
        "workbook_uuid": uuid,
        "sheet_count": len(workbook_data.sheets),
        "total_cells": total_cells,
        "total_formulas": total_formulas,
        "total_anomalies": total_anomalies,
        "formula_pct": formula_pct,
        "empty_pct": empty_pct,
        "quality_score": quality_score,
        "named_ranges": len(workbook_data.named_ranges),
        "sheets": sheets_summary,
        "cross_sheet_deps": cross_sheet_deps,
    }


@router.get("/{uuid}/download")
async def download_workbook(
    uuid: str,
    data_cache=Depends(get_workbook_data_cache),
):
    """Export current workbook state (with AI modifications) as .xlsx."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    import io
    from openpyxl import Workbook

    workbook_data = data_cache[uuid]
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    for sheet_name in workbook_data.sheets:
        ws = wb.create_sheet(title=sheet_name)
        sheet_cells = [
            c for c in workbook_data.cells.values()
            if c.sheet_name == sheet_name
        ]
        for cd in sheet_cells:
            cell = ws.cell(row=cd.row, column=cd.col)
            if cd.formula:
                cell.value = f"={cd.formula}"
            else:
                cell.value = cd.value

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="workbook_{uuid[:8]}.xlsx"'},
    )


@router.post("/{uuid}/cell-edit")
async def cell_edit(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
    states=Depends(get_workbook_states),
    audit_trails=Depends(get_audit_trails),
):
    """Direct cell edit from the grid UI (user types a value or formula)."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    sheet = payload.get("sheet", "")
    cell = payload.get("cell", "")  # e.g. "B4"
    value = payload.get("value", "")

    if not sheet or not cell:
        raise HTTPException(status_code=400, detail="sheet and cell are required.")

    from parser.xlsx_parser import CellData, _infer_data_type
    from openpyxl.utils import column_index_from_string

    full_addr = f"{sheet}!{cell}"
    workbook_data = data_cache[uuid]
    workbook_state = states.get(uuid, {})

    # Capture old value
    old_value = None
    if full_addr in workbook_data.cells:
        old_value = workbook_data.cells[full_addr].value

    # Detect formula vs value
    is_formula = isinstance(value, str) and value.startswith("=")
    formula_str = value[1:] if is_formula else None

    # Parse numeric
    parsed_value = value
    if not is_formula and isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            parsed_value = None
        else:
            try:
                parsed_value = int(stripped)
            except ValueError:
                try:
                    parsed_value = float(stripped)
                except ValueError:
                    parsed_value = stripped

    # Write to workbook_data
    if full_addr in workbook_data.cells:
        workbook_data.cells[full_addr].value = parsed_value
        workbook_data.cells[full_addr].formula = formula_str
        workbook_data.cells[full_addr].data_type = _infer_data_type(parsed_value, formula_str)
        workbook_data.cells[full_addr].is_hardcoded = not is_formula and isinstance(parsed_value, (int, float))
    else:
        col_part = "".join(c for c in cell if c.isalpha())
        row_part = "".join(c for c in cell if c.isdigit())
        if col_part and row_part:
            workbook_data.cells[full_addr] = CellData(
                cell_address=full_addr,
                sheet_name=sheet,
                value=parsed_value,
                formula=formula_str,
                data_type=_infer_data_type(parsed_value, formula_str),
                named_range=None,
                row=int(row_part),
                col=column_index_from_string(col_part),
                is_hardcoded=not is_formula and isinstance(parsed_value, (int, float)),
                is_merged=False,
                merge_master=None,
            )

    # Update state
    workbook_state[full_addr] = {
        "value": parsed_value,
        "formula": formula_str,
        "data_type": _infer_data_type(parsed_value, formula_str),
    }

    # Log to audit trail
    from analysis.audit_trail import AuditTrail
    if uuid not in audit_trails:
        audit_trails[uuid] = AuditTrail()
    audit_trails[uuid].record_change(
        cell=full_addr,
        sheet=sheet,
        old_value=old_value,
        new_value=parsed_value,
        old_formula=None,
        new_formula=formula_str,
        reason="Manual cell edit",
        approved_by="user",
    )

    return {
        "cell": full_addr,
        "old_value": old_value,
        "new_value": parsed_value,
        "formula": formula_str,
    }


@router.post("/{uuid}/revert")
async def revert_changes(
    uuid: str,
    changes: list = Body(...),
    data_cache=Depends(get_workbook_data_cache),
    states=Depends(get_workbook_states),
):
    """Revert cells to their previous values (called when user rejects AI changes)."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from parser.xlsx_parser import _infer_data_type

    workbook_data = data_cache[uuid]
    workbook_state = states.get(uuid, {})
    reverted = 0

    for change in changes:
        cell_addr = change.get("cell", "")
        old_val = change.get("before")
        if cell_addr in workbook_data.cells:
            workbook_data.cells[cell_addr].value = old_val
            workbook_data.cells[cell_addr].data_type = _infer_data_type(old_val, None)
            reverted += 1
        if cell_addr in workbook_state:
            workbook_state[cell_addr]["value"] = old_val

    return {"reverted": reverted}


@router.post("/{uuid}/confirm-delete")
async def confirm_delete(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
    states=Depends(get_workbook_states),
    graphs=Depends(get_workbook_graphs),
):
    """Execute a confirmed delete_range (second approval step)."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    sheet = payload.get("sheet", "")
    range_str = payload.get("range", "")

    from agent.tools import ExcelTools
    tools = ExcelTools(
        workbook_state=states.get(uuid, {}),
        graph=graphs.get(uuid),
        workbook_data=data_cache[uuid],
    )
    result = tools.delete_range(sheet, range_str, confirmed=True)
    return {"success": result.success, "data": result.data, "error": result.error}


@router.post("/{uuid}/delta", response_model=DeltaResponse)
async def delta_update(
    uuid: str,
    request: DeltaRequest,
    embedder=Depends(get_embedder),
    chroma=Depends(get_chroma),
    lsh=Depends(get_lsh),
    graphs=Depends(get_workbook_graphs),
    data_cache=Depends(get_workbook_data_cache),
):
    if uuid not in graphs:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    start = time.time()
    graph = graphs[uuid]
    workbook_data = data_cache[uuid]

    context_text = " ".join(request.changed_cells)
    context_embedding = embedder.embed_single(context_text)
    candidate_chunk_ids = lsh.find_candidates(context_embedding, hamming_threshold=2)

    chunk_builder = ChunkBuilder()
    all_chunks = chunk_builder.build_chunks(workbook_data, graph, uuid)
    affected_chunks = [c for c in all_chunks if c.chunk_id in set(candidate_chunk_ids)]

    if not affected_chunks:
        affected_chunks = all_chunks[:3]

    texts = [c.text for c in affected_chunks]
    if texts:
        embeddings = embedder.embed(texts)
        chunk_records = []
        for i, chunk in enumerate(affected_chunks):
            lsh.add(chunk.chunk_id, embeddings[i])
            chunk_records.append(
                ChunkRecord(
                    chunk_id=chunk.chunk_id,
                    text=chunk.text,
                    embedding=embeddings[i],
                    metadata=chunk.metadata,
                )
            )
        chroma.delta_upsert(uuid, candidate_chunk_ids, chunk_records)

    elapsed_ms = int((time.time() - start) * 1000)
    return DeltaResponse(chunks_updated=len(affected_chunks), time_ms=elapsed_ms)


@router.post("/{uuid}/state", response_model=StateResponse)
async def update_state(
    uuid: str,
    request: StateRequest,
    states=Depends(get_workbook_states),
):
    if uuid not in states:
        states[uuid] = {}
    states[uuid].update(request.cells)
    return StateResponse(updated=len(request.cells))


@router.get("/{uuid}/anomalies", response_model=AnomaliesResponse)
async def get_anomalies(
    uuid: str,
    graphs=Depends(get_workbook_graphs),
):
    if uuid not in graphs:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    graph = graphs[uuid]
    anomalies = []
    for node in graph.nodes():
        nd = graph.nodes[node]
        if not nd.get("is_anomaly"):
            continue
        score = nd.get("anomaly_score", 0.0)
        if score < -0.1:
            severity = "high"
        elif score < 0.0:
            severity = "medium"
        else:
            severity = "low"
        anomalies.append(
            AnomalyItem(
                cell=node,
                value=nd.get("value"),
                anomaly_score=score,
                formula=nd.get("formula"),
                sheet_name=nd.get("sheet_name"),
                named_range=nd.get("named_range"),
                severity=severity,
            )
        )
    anomalies.sort(key=lambda x: x.anomaly_score)
    return AnomaliesResponse(anomalies=anomalies)


@router.get("/{uuid}/grid")
async def get_grid(
    uuid: str,
    data_cache=Depends(get_workbook_data_cache),
):
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl
    workbook_data = data_cache[uuid]

    sheet_bounds: Dict[str, Dict[str, int]] = {}
    for cell in workbook_data.cells.values():
        s = cell.sheet_name
        if s not in sheet_bounds:
            sheet_bounds[s] = {"max_row": 0, "max_col": 0}
        if cell.row > sheet_bounds[s]["max_row"]:
            sheet_bounds[s]["max_row"] = cell.row
        if cell.col > sheet_bounds[s]["max_col"]:
            sheet_bounds[s]["max_col"] = cell.col

    sheets_out = []
    for sheet_name in workbook_data.sheets:
        if sheet_name not in sheet_bounds:
            continue
        max_row = min(sheet_bounds[sheet_name]["max_row"], 200)
        max_col = min(sheet_bounds[sheet_name]["max_col"], 26)
        col_headers = [gcl(c) for c in range(1, max_col + 1)]

        cell_lookup = {}
        for cell in workbook_data.cells.values():
            if cell.sheet_name == sheet_name:
                cell_lookup[(cell.row, cell.col)] = cell

        rows = []
        for r in range(1, max_row + 1):
            row_cells = []
            for c in range(1, max_col + 1):
                cd = cell_lookup.get((r, c))
                if cd is None:
                    row_cells.append({"v": "", "f": False, "t": "empty"})
                else:
                    val = cd.value
                    if val is None:
                        display = cd.formula or ""
                    elif isinstance(val, float):
                        display = f"{val:,.2f}" if abs(val) >= 1000 else str(round(val, 6)).rstrip('0').rstrip('.')
                    elif isinstance(val, bool):
                        display = "TRUE" if val else "FALSE"
                    else:
                        display = str(val)
                    cell_out = {
                        "v": display,
                        "f": bool(cd.formula),
                        "t": cd.data_type,
                        "nr": cd.named_range or "",
                    }
                    if cd.style:
                        sd = cd.style.to_dict()
                        if sd:
                            cell_out["s"] = sd
                    if cd.is_merged:
                        cell_out["mg"] = True
                    if cd.merge_master:
                        cell_out["mm"] = cd.merge_master
                    row_cells.append(cell_out)
            rows.append(row_cells)

        sheets_out.append({
            "name": sheet_name,
            "colHeaders": col_headers,
            "rows": rows,
        })

    return {"sheets": sheets_out}


@router.get("/{uuid}/validate-formulas")
async def validate_formulas(
    uuid: str,
    data_cache=Depends(get_workbook_data_cache),
):
    """Run formula consistency validation on the workbook."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from analysis.formula_validator import FormulaValidator
    validator = FormulaValidator(data_cache[uuid])
    report = validator.run()

    return {
        "consistency_score": report.consistency_score,
        "formula_count": report.formula_count,
        "total_issues": report.total_issues,
        "errors": report.errors,
        "warnings": report.warnings,
        "info": report.info,
        "issues": [
            {
                "severity": i.severity,
                "category": i.category,
                "cell": i.cell,
                "sheet": i.sheet,
                "message": i.message,
                "expected": i.expected,
                "actual": i.actual,
            }
            for i in report.issues[:100]
        ],
    }


@router.post("/{uuid}/diff")
async def diff_workbooks(
    uuid: str,
    file_b: UploadFile = File(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Compare uploaded workbook (A) against a second file (B). Returns structural + value diff."""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    if not file_b.filename or not file_b.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        content = await file_b.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        from analysis.version_diff import VersionDiffer

        parser_b = XLSXParser()
        wb_b = parser_b.parse(tmp_path)
        wb_a = data_cache[uuid]

        differ = VersionDiffer()
        report = differ.diff_workbooks(wb_a, wb_b, label_a="current", label_b=file_b.filename or "uploaded")

        return {
            "summary": report.summary,
            "total_changes": report.total_changes,
            "sheets_added": report.sheets_added,
            "sheets_removed": report.sheets_removed,
            "sheet_diffs": [
                {
                    "sheet": sd.sheet_name,
                    "added": sd.cells_added,
                    "removed": sd.cells_removed,
                    "modified": sd.cells_modified,
                    "formula_changes": sd.formula_changes,
                    "value_changes": sd.value_changes,
                    "max_delta_pct": sd.max_delta_pct,
                }
                for sd in report.sheet_diffs
            ],
            "cell_diffs": [
                {
                    "cell": cd.cell,
                    "sheet": cd.sheet,
                    "type": cd.change_type,
                    "old_value": cd.old_value,
                    "new_value": cd.new_value,
                    "old_formula": cd.old_formula,
                    "new_formula": cd.new_formula,
                    "delta": cd.delta,
                    "delta_pct": cd.delta_pct,
                    "impact": cd.impact_score,
                }
                for cd in report.cell_diffs[:500]
            ],
            "high_impact": [
                {
                    "cell": cd.cell,
                    "sheet": cd.sheet,
                    "type": cd.change_type,
                    "old_value": cd.old_value,
                    "new_value": cd.new_value,
                    "delta_pct": cd.delta_pct,
                    "impact": cd.impact_score,
                }
                for cd in report.high_impact_changes[:50]
            ],
        }
    finally:
        os.unlink(tmp_path)


@router.get("/{uuid}/explain/{cell:path}", response_model=ExplainResponse)
async def explain_cell(
    uuid: str,
    cell: str,
    graphs=Depends(get_workbook_graphs),
    data_cache=Depends(get_workbook_data_cache),
    states=Depends(get_workbook_states),
):
    if uuid not in graphs:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    graph = graphs[uuid]
    workbook_data = data_cache[uuid]
    tools = ExcelTools(
        workbook_state=states.get(uuid, {}),
        graph=graph,
        workbook_data=workbook_data,
    )
    result = tools.explain_formula(cell)
    if not result.success:
        raise HTTPException(status_code=404, detail=result.error)

    data = result.data
    return ExplainResponse(
        cell=data.get("cell", cell),
        value=data.get("value"),
        formula=data.get("formula"),
        named_range=data.get("named_range"),
        cluster=data.get("cluster"),
        pagerank=data.get("pagerank", 0.0),
        depends_on=data.get("depends_on", []),
    )


@router.post("/{uuid}/format-cells")
async def format_cells(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Apply formatting to a range of cells.
    payload: { cells: ["Sheet1!A1", ...], style: { b: true, fc: "#FF0000", ... } }
    """
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from parser.xlsx_parser import CellStyle

    cell_addrs = payload.get("cells", [])
    style_patch = payload.get("style", {})
    workbook_data = data_cache[uuid]

    from parser.xlsx_parser import CellData, _infer_data_type
    from openpyxl.utils import column_index_from_string

    updated = 0
    for addr in cell_addrs:
        cd = workbook_data.cells.get(addr)
        if not cd:
            # Create a placeholder CellData for cells not yet in cache
            try:
                parts = addr.split("!")
                if len(parts) != 2:
                    continue
                sheet_name, cell_ref = parts[0], parts[1]
                import re
                m = re.match(r"([A-Za-z]+)(\d+)", cell_ref)
                if not m:
                    continue
                col_idx = column_index_from_string(m.group(1))
                row_idx = int(m.group(2))
                cd = CellData(
                    cell_address=addr,
                    sheet_name=sheet_name,
                    value=None,
                    formula=None,
                    data_type="empty",
                    named_range=None,
                    row=row_idx,
                    col=col_idx,
                    is_hardcoded=False,
                    is_merged=False,
                    merge_master=None,
                )
                workbook_data.cells[addr] = cd
            except Exception:
                continue
        if cd.style is None:
            cd.style = CellStyle()
        s = cd.style
        if "b" in style_patch: s.bold = bool(style_patch["b"])
        if "i" in style_patch: s.italic = bool(style_patch["i"])
        if "u" in style_patch: s.underline = bool(style_patch["u"])
        if "fs" in style_patch: s.font_size = float(style_patch["fs"])
        if "fn" in style_patch: s.font_name = str(style_patch["fn"])
        if "fc" in style_patch: s.font_color = str(style_patch["fc"])
        if "bg" in style_patch: s.bg_color = str(style_patch["bg"])
        if "ha" in style_patch: s.h_align = str(style_patch["ha"])
        if "va" in style_patch: s.v_align = str(style_patch["va"])
        if "wt" in style_patch: s.wrap_text = bool(style_patch["wt"])
        if "ind" in style_patch: s.indent = int(style_patch["ind"])
        if "nf" in style_patch: s.number_format = str(style_patch["nf"])
        if "bt" in style_patch: s.border_top = str(style_patch["bt"])
        if "bb" in style_patch: s.border_bottom = str(style_patch["bb"])
        if "bl" in style_patch: s.border_left = str(style_patch["bl"])
        if "br" in style_patch: s.border_right = str(style_patch["br"])
        if "bc" in style_patch: s.border_color = str(style_patch["bc"])
        updated += 1

    return {"updated": updated}


@router.post("/{uuid}/insert-row")
async def insert_row(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Insert a row at given position. payload: { sheet, row (1-based), count? }"""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl
    from parser.xlsx_parser import CellData, _infer_data_type

    sheet = payload.get("sheet", "")
    row_num = int(payload.get("row", 1))
    count = int(payload.get("count", 1))
    workbook_data = data_cache[uuid]

    cells_to_shift = sorted(
        [c for c in workbook_data.cells.values() if c.sheet_name == sheet and c.row >= row_num],
        key=lambda c: (-c.row, c.col),
    )

    for cd in cells_to_shift:
        old_addr = cd.cell_address
        del workbook_data.cells[old_addr]
        cd.row += count
        col_letter = gcl(cd.col)
        cd.cell_address = f"{sheet}!{col_letter}{cd.row}"
        workbook_data.cells[cd.cell_address] = cd

    return {"inserted": count, "at_row": row_num}


@router.post("/{uuid}/insert-col")
async def insert_col(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Insert column at position. payload: { sheet, col (1-based), count? }"""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl
    from parser.xlsx_parser import CellData, _infer_data_type

    sheet = payload.get("sheet", "")
    col_num = int(payload.get("col", 1))
    count = int(payload.get("count", 1))
    workbook_data = data_cache[uuid]

    cells_to_shift = sorted(
        [c for c in workbook_data.cells.values() if c.sheet_name == sheet and c.col >= col_num],
        key=lambda c: (c.row, -c.col),
    )

    for cd in cells_to_shift:
        old_addr = cd.cell_address
        del workbook_data.cells[old_addr]
        cd.col += count
        col_letter = gcl(cd.col)
        cd.cell_address = f"{sheet}!{col_letter}{cd.row}"
        workbook_data.cells[cd.cell_address] = cd

    return {"inserted": count, "at_col": col_num}


@router.post("/{uuid}/delete-row")
async def delete_row(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Delete row(s). payload: { sheet, row (1-based), count? }"""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl

    sheet = payload.get("sheet", "")
    row_num = int(payload.get("row", 1))
    count = int(payload.get("count", 1))
    workbook_data = data_cache[uuid]

    to_remove = [
        addr for addr, c in workbook_data.cells.items()
        if c.sheet_name == sheet and row_num <= c.row < row_num + count
    ]
    for addr in to_remove:
        del workbook_data.cells[addr]

    cells_to_shift = sorted(
        [c for c in workbook_data.cells.values() if c.sheet_name == sheet and c.row >= row_num + count],
        key=lambda c: (c.row, c.col),
    )

    for cd in cells_to_shift:
        old_addr = cd.cell_address
        del workbook_data.cells[old_addr]
        cd.row -= count
        col_letter = gcl(cd.col)
        cd.cell_address = f"{sheet}!{col_letter}{cd.row}"
        workbook_data.cells[cd.cell_address] = cd

    return {"deleted": count, "at_row": row_num}


@router.post("/{uuid}/delete-col")
async def delete_col(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Delete column(s). payload: { sheet, col (1-based), count? }"""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl

    sheet = payload.get("sheet", "")
    col_num = int(payload.get("col", 1))
    count = int(payload.get("count", 1))
    workbook_data = data_cache[uuid]

    to_remove = [
        addr for addr, c in workbook_data.cells.items()
        if c.sheet_name == sheet and col_num <= c.col < col_num + count
    ]
    for addr in to_remove:
        del workbook_data.cells[addr]

    cells_to_shift = sorted(
        [c for c in workbook_data.cells.values() if c.sheet_name == sheet and c.col >= col_num + count],
        key=lambda c: (c.row, c.col),
    )

    for cd in cells_to_shift:
        old_addr = cd.cell_address
        del workbook_data.cells[old_addr]
        cd.col -= count
        col_letter = gcl(cd.col)
        cd.cell_address = f"{sheet}!{col_letter}{cd.row}"
        workbook_data.cells[cd.cell_address] = cd

    return {"deleted": count, "at_col": col_num}


@router.post("/{uuid}/data-validation")
async def set_data_validation(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Set data validation on a cell range.
    payload: { cells: ["Sheet1!A1",...], type: "list"|"number"|"text_length",
               values?: ["a","b"], min?: 0, max?: 100, message?: "...", error_message?: "..." }
    """
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    cells = payload.get("cells", [])
    rule = {
        "type": payload.get("type", "list"),
        "values": payload.get("values", []),
        "min": payload.get("min"),
        "max": payload.get("max"),
        "message": payload.get("message", ""),
        "error_message": payload.get("error_message", "Invalid input"),
    }

    workbook_data = data_cache[uuid]
    if not hasattr(workbook_data, "validations"):
        workbook_data.validations = {}

    for addr in cells:
        workbook_data.validations[addr] = rule

    return {"updated": len(cells), "rule": rule}


@router.get("/{uuid}/data-validation")
async def get_data_validation(
    uuid: str,
    data_cache=Depends(get_workbook_data_cache),
):
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")
    workbook_data = data_cache[uuid]
    validations = getattr(workbook_data, "validations", {})
    return {"validations": validations}


@router.post("/{uuid}/named-ranges")
async def set_named_range(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Define a named range. payload: { name: "MyRange", range: "Sheet1!A1:B10" }"""
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    name = payload.get("name", "").strip()
    range_str = payload.get("range", "").strip()
    if not name or not range_str:
        raise HTTPException(status_code=400, detail="name and range are required.")

    workbook_data = data_cache[uuid]
    workbook_data.named_ranges[name] = range_str

    return {"name": name, "range": range_str, "total": len(workbook_data.named_ranges)}


@router.delete("/{uuid}/named-ranges/{name}")
async def delete_named_range(
    uuid: str,
    name: str,
    data_cache=Depends(get_workbook_data_cache),
):
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")
    workbook_data = data_cache[uuid]
    if name in workbook_data.named_ranges:
        del workbook_data.named_ranges[name]
    return {"deleted": name, "total": len(workbook_data.named_ranges)}


@router.get("/{uuid}/named-ranges")
async def get_named_ranges(
    uuid: str,
    data_cache=Depends(get_workbook_data_cache),
):
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")
    return {"named_ranges": data_cache[uuid].named_ranges}


@router.post("/{uuid}/goal-seek")
async def goal_seek(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Goal Seek: find input value that makes target cell equal goal.
    payload: { target_cell: "Sheet1!B10", goal: 1000, changing_cell: "Sheet1!A1" }
    Uses simple bisection method.
    """
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from parser.xlsx_parser import _infer_data_type

    target_addr = payload.get("target_cell", "")
    goal = float(payload.get("goal", 0))
    changing_addr = payload.get("changing_cell", "")
    workbook_data = data_cache[uuid]

    if target_addr not in workbook_data.cells or changing_addr not in workbook_data.cells:
        raise HTTPException(status_code=404, detail="Cell not found.")

    changing_cell = workbook_data.cells[changing_addr]
    original_value = changing_cell.value

    def get_target_value(input_val):
        changing_cell.value = input_val
        target = workbook_data.cells.get(target_addr)
        if target and target.value is not None:
            try:
                return float(target.value)
            except (ValueError, TypeError):
                return None
        return None

    lo, hi = -1e6, 1e6
    result = original_value
    best_diff = float("inf")

    for _ in range(100):
        mid = (lo + hi) / 2.0
        val = get_target_value(mid)
        if val is None:
            break
        diff = abs(val - goal)
        if diff < best_diff:
            best_diff = diff
            result = mid
        if diff < 0.0001:
            break
        if val < goal:
            lo = mid
        else:
            hi = mid

    changing_cell.value = result
    changing_cell.data_type = _infer_data_type(result, None)

    return {
        "changing_cell": changing_addr,
        "result_value": result,
        "target_cell": target_addr,
        "achieved_value": get_target_value(result),
        "goal": goal,
        "difference": best_diff,
    }


@router.post("/{uuid}/text-to-columns")
async def text_to_columns(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Split a column's text by delimiter into multiple columns.
    payload: { sheet, col (1-based), delimiter: "," | ";" | " " | "\\t" }
    """
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl
    from parser.xlsx_parser import CellData, CellStyle, _infer_data_type

    sheet = payload.get("sheet", "")
    col_num = int(payload.get("col", 1))
    delimiter = payload.get("delimiter", ",")
    workbook_data = data_cache[uuid]

    cells_in_col = sorted(
        [c for c in workbook_data.cells.values()
         if c.sheet_name == sheet and c.col == col_num],
        key=lambda c: c.row,
    )

    max_parts = 0
    split_data = []
    for cd in cells_in_col:
        val = str(cd.value) if cd.value is not None else ""
        parts = val.split(delimiter)
        parts = [p.strip() for p in parts]
        max_parts = max(max_parts, len(parts))
        split_data.append((cd.row, parts))

    created = 0
    for row_num, parts in split_data:
        for i, part in enumerate(parts):
            target_col = col_num + i
            col_letter = gcl(target_col)
            addr = f"{sheet}!{col_letter}{row_num}"
            parsed = part
            try:
                parsed = int(part)
            except ValueError:
                try:
                    parsed = float(part)
                except ValueError:
                    pass

            if addr in workbook_data.cells:
                workbook_data.cells[addr].value = parsed
                workbook_data.cells[addr].data_type = _infer_data_type(parsed, None)
            else:
                workbook_data.cells[addr] = CellData(
                    cell_address=addr, sheet_name=sheet, value=parsed,
                    formula=None, data_type=_infer_data_type(parsed, None),
                    named_range=None, row=row_num, col=target_col,
                    is_hardcoded=isinstance(parsed, (int, float)),
                    is_merged=False, merge_master=None,
                )
            created += 1

    return {"split_into_columns": max_parts, "cells_written": created}


@router.post("/{uuid}/remove-duplicates")
async def remove_duplicates(
    uuid: str,
    payload: dict = Body(...),
    data_cache=Depends(get_workbook_data_cache),
):
    """Remove duplicate rows based on specified columns.
    payload: { sheet, columns: [1, 2] (1-based col indices, empty = all), keep_first: true }
    """
    if uuid not in data_cache:
        raise HTTPException(status_code=404, detail=f"Workbook {uuid} not found.")

    from openpyxl.utils import get_column_letter as gcl

    sheet = payload.get("sheet", "")
    columns = payload.get("columns", [])
    keep_first = payload.get("keep_first", True)
    workbook_data = data_cache[uuid]

    sheet_cells = [c for c in workbook_data.cells.values() if c.sheet_name == sheet]
    if not sheet_cells:
        return {"removed": 0}

    max_row = max(c.row for c in sheet_cells)
    max_col = max(c.col for c in sheet_cells)

    cell_lookup = {}
    for c in sheet_cells:
        cell_lookup[(c.row, c.col)] = c

    check_cols = columns if columns else list(range(1, max_col + 1))
    seen = set()
    rows_to_remove = []

    for r in range(2, max_row + 1):
        key_parts = []
        for col in check_cols:
            cd = cell_lookup.get((r, col))
            key_parts.append(str(cd.value) if cd and cd.value is not None else "")
        key = tuple(key_parts)
        if key in seen:
            rows_to_remove.append(r)
        else:
            seen.add(key)

    removed = 0
    for r in rows_to_remove:
        addrs_to_del = [
            addr for addr, c in workbook_data.cells.items()
            if c.sheet_name == sheet and c.row == r
        ]
        for addr in addrs_to_del:
            del workbook_data.cells[addr]
        removed += 1

    return {"removed": removed, "remaining_rows": max_row - removed}
