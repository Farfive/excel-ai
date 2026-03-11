import io
import os
import pytest
import openpyxl
from openpyxl.utils import get_column_letter


@pytest.fixture
def sample_xlsx_path(tmp_path):
    wb = openpyxl.Workbook()

    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_assumptions["A1"] = 0.085
    ws_assumptions["A2"] = 0.032
    ws_assumptions["A3"] = 0.19

    ws_dcf = wb.create_sheet("DCF")
    ws_dcf["B2"] = "=Assumptions!A1*100"
    ws_dcf["B3"] = "=SUM(C2:C5)"
    ws_dcf["C2"] = 1000
    ws_dcf["C3"] = 1200
    ws_dcf["C4"] = 1400
    ws_dcf["C5"] = 1600

    wb.create_named_range("discount_rate", ws_assumptions, "A1")

    file_path = str(tmp_path / "sample.xlsx")
    wb.save(file_path)
    return file_path


@pytest.fixture
def empty_xlsx_path(tmp_path):
    wb = openpyxl.Workbook()
    file_path = str(tmp_path / "empty.xlsx")
    wb.save(file_path)
    return file_path
