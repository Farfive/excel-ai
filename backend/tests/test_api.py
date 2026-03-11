import io
import pytest
import pytest_asyncio
import openpyxl
from httpx import AsyncClient, ASGITransport
from unittest.mock import patch, MagicMock


def make_xlsx_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = 0.085
    ws["A2"] = 0.032
    ws_dcf = wb.create_sheet("DCF")
    ws_dcf["B2"] = "=Assumptions!A1*100"
    ws_dcf["C2"] = 1000
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture
def mock_embedder():
    m = MagicMock()
    m._model = MagicMock()
    m.embed.return_value = [[0.1] * 768]
    m.embed_single.return_value = [0.1] * 768
    m.load.return_value = None
    return m


@pytest.fixture
def mock_ollama():
    m = MagicMock()
    m.is_running = pytest.mark.asyncio(lambda: True)
    import asyncio
    async def _is_running():
        return True
    m.is_running = _is_running
    async def _close():
        pass
    m.close = _close
    return m


@pytest.fixture
def mock_chroma():
    m = MagicMock()
    m.upsert_chunks.return_value = None
    m.query.return_value = []
    m.workbook_exists.return_value = False
    m._get_client.return_value = MagicMock()
    return m


@pytest.fixture
def app_with_mocks(mock_embedder, mock_ollama, mock_chroma):
    import api.dependencies as deps
    deps.embedder = mock_embedder
    deps.ollama = mock_ollama
    deps.chroma = mock_chroma

    from api.main import app
    return app


@pytest.mark.asyncio
async def test_health_returns_ok(app_with_mocks):
    async with AsyncClient(transport=ASGITransport(app=app_with_mocks), base_url="http://test") as client:
        response = await client.get("/health")
    assert response.status_code == 200
    data = response.json()
    assert "status" in data
    assert "components" in data


@pytest.mark.asyncio
async def test_upload_returns_workbook_uuid(app_with_mocks, mock_embedder):
    mock_embedder.embed.return_value = [[0.1] * 768 for _ in range(20)]
    xlsx_bytes = make_xlsx_bytes()
    async with AsyncClient(transport=ASGITransport(app=app_with_mocks), base_url="http://test") as client:
        response = await client.post(
            "/workbook/upload",
            files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert response.status_code == 200
    data = response.json()
    assert "workbook_uuid" in data
    assert "cell_count" in data


@pytest.mark.asyncio
async def test_upload_non_xlsx_returns_400(app_with_mocks):
    async with AsyncClient(transport=ASGITransport(app=app_with_mocks), base_url="http://test") as client:
        response = await client.post(
            "/workbook/upload",
            files={"file": ("test.csv", b"col1,col2\n1,2", "text/csv")},
        )
    assert response.status_code == 400


@pytest.mark.asyncio
async def test_anomalies_endpoint_works(app_with_mocks, mock_embedder):
    mock_embedder.embed.return_value = [[0.1] * 768 for _ in range(20)]
    xlsx_bytes = make_xlsx_bytes()
    async with AsyncClient(transport=ASGITransport(app=app_with_mocks), base_url="http://test") as client:
        upload_resp = await client.post(
            "/workbook/upload",
            files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_resp.status_code == 200
        uuid = upload_resp.json()["workbook_uuid"]

        anomaly_resp = await client.get(f"/workbook/{uuid}/anomalies")
    assert anomaly_resp.status_code == 200
    data = anomaly_resp.json()
    assert "anomalies" in data


@pytest.mark.asyncio
async def test_delta_update_works(app_with_mocks, mock_embedder):
    mock_embedder.embed.return_value = [[0.1] * 768 for _ in range(20)]
    xlsx_bytes = make_xlsx_bytes()
    async with AsyncClient(transport=ASGITransport(app=app_with_mocks), base_url="http://test") as client:
        upload_resp = await client.post(
            "/workbook/upload",
            files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert upload_resp.status_code == 200
        uuid = upload_resp.json()["workbook_uuid"]

        delta_resp = await client.post(
            f"/workbook/{uuid}/delta",
            json={"changed_cells": ["Assumptions!A1", "Assumptions!A2"]},
        )
    assert delta_resp.status_code == 200
    data = delta_resp.json()
    assert "chunks_updated" in data
