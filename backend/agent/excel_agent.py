import difflib
import json
import logging
import re
from collections import defaultdict
from typing import Any, AsyncGenerator, Dict, List, Optional

from pydantic import BaseModel, validator

from agent.ollama_client import OllamaClient
from agent.tools import ExcelTools, ToolResult
from rag.retrieval import RAGRetriever
from analysis.sensitivity import SensitivityAnalyzer
from analysis.integrity import IntegrityChecker
from analysis.smart_suggestions import SmartSuggestionsEngine
from analysis.scenarios import ScenarioManager
from parser.xlsx_parser import WorkbookData

logger = logging.getLogger(__name__)

VALID_TOOLS = {
    'read_range', 'write_range', 'write_formula', 'add_column', 'add_row',
    'create_sheet', 'delete_range', 'get_dependencies', 'find_anomalies',
    'explain_formula', 'generate_change_log', 'get_audit_trail', 'pivot_table',
    'validate_formulas', 'run_sensitivity', 'run_integrity_check',
    'run_smart_suggestions', 'create_scenario', 'compare_scenarios',
}

CELL_ADDR_RE = re.compile(r'^[A-Za-z]+\d+(:[A-Za-z]+\d+)?$')


class ToolStepModel(BaseModel):
    step: int
    tool: str
    args: dict
    reason: str

    @validator('tool')
    def tool_must_exist(cls, v: str) -> str:
        if v not in VALID_TOOLS:
            raise ValueError(f'Unknown tool: {v}. Valid: {sorted(VALID_TOOLS)}')
        return v


class PlanModel(BaseModel):
    steps: List[ToolStepModel]

SYSTEM_PROMPT = """You are an expert financial model analyst AI assistant embedded in an Excel workbook viewer.
You have DIRECT ACCESS to the workbook data. You can read, write, create sheets, add columns/rows, delete data, write formulas, and run analyses.

CRITICAL RULES:
1. You MUST use the CELL INDEX provided below to find the EXACT cell addresses. Never guess cell addresses.
2. When the user says "change X to Y", find X in the cell index, identify the correct sheet and cell address, then use write_range.
3. Always use the EXACT sheet name from the cell index (case-sensitive).
4. When writing values, pass them as a list: "values": [0.69] for single cell, "values": [0.1, 0.2, 0.3] for ranges.
5. Respond in the SAME LANGUAGE as the user (Polish → Polish, English → English).
6. After writing, briefly confirm what was changed and the cell address.
7. For DELETE operations: ALWAYS use confirmed=false first to show preview. The system will handle the confirmation flow.
8. For formulas: use write_formula, NOT write_range. Do NOT include the leading "=" in the formula string.

MULTI-SHEET AWARENESS:
- The CELL INDEX below shows ALL sheets in the workbook, not just one.
- The XREF section at the bottom shows cross-sheet formula dependencies (e.g. Revenue→P&L means P&L has formulas referencing Revenue).
- When a user request affects data that flows to other sheets (via XREF), consider downstream impacts.
- You can create multi-step plans that span multiple sheets in a single operation.
- When reading or writing, always specify the correct sheet name for each step.
- If a change in one sheet should propagate (e.g. changing an assumption that affects Revenue, P&L, DCF), explain the downstream impact.

AVAILABLE TOOLS:

Data Reading:
- read_range(sheet, range): Read cell values. range like "B4" or "B4:H4".
- get_dependencies(cell): Get upstream/downstream dependencies (full address like "Assumptions!B4").
- explain_formula(cell): Explain a formula and its dependencies.

Data Writing:
- write_range(sheet, range, values): Write values to cells. values is a list.
- write_formula(sheet, cell, formula): Write a formula to a cell. formula WITHOUT leading "=". Example: "SUM(B2:B10)"
- add_column(sheet, after_col, header, values): Add a new column after the specified column. values is optional list.
- add_row(sheet, row_num, values): Write data into a row. values = {"A": val1, "B": val2, ...}
- create_sheet(name): Create a new empty sheet in the workbook.

Data Deletion (DANGEROUS — requires double confirmation):
- delete_range(sheet, range, confirmed): Delete cells. ALWAYS call with confirmed=false first to preview.

Aggregation & Pivot:
- pivot_table(source_sheet, target_sheet, row_field, col_field, value_field, agg): Create a pivot/summary sheet.
  row_field/col_field/value_field = column HEADER text from row 1. agg = sum|avg|count|min|max.
  col_field is optional (omit for 1D grouping).

Analysis:
- find_anomalies(sheet): Find statistical anomalies.
- validate_formulas(): Check formula consistency across the workbook (patterns, gaps, hardcoded values).
- run_sensitivity(args): Run sensitivity analysis.
- run_integrity_check(): Check model integrity.
- get_audit_trail(): Show history of all AI actions in this session.

FEW-SHOT EXAMPLES:

User: "Zmień Revenue Growth Rate na 0.25"
Thinking: Cell index shows Revenue Growth Rate is in Assumptions!B4 (row label in column A, value in B4=0.15)
Plan: [{"step": 1, "tool": "write_range", "args": {"sheet": "Assumptions", "range": "B4", "values": [0.25]}, "reason": "User requested Revenue Growth Rate change to 0.25"}]

User: "Add a Notes column to the Assumptions sheet"
Plan: [{"step": 1, "tool": "add_column", "args": {"sheet": "Assumptions", "after_col": "K", "header": "Notes"}, "reason": "Add Notes column after last column"}]

User: "Create a new sheet called Summary"
Plan: [{"step": 1, "tool": "create_sheet", "args": {"name": "Summary"}, "reason": "Create Summary sheet"}]

User: "Add a formula that sums Revenue B2:B10"
Plan: [{"step": 1, "tool": "write_formula", "args": {"sheet": "Revenue", "cell": "B11", "formula": "SUM(B2:B10)"}, "reason": "Add SUM formula"}]

User: "Delete row 5 in Assumptions"
Thinking: Deletion is dangerous. I must preview first with confirmed=false.
Plan: [{"step": 1, "tool": "delete_range", "args": {"sheet": "Assumptions", "range": "A5:K5", "confirmed": false}, "reason": "Preview data in row 5 before deletion"}]

User: "What is the current WACC?"
Plan: [{"step": 1, "tool": "read_range", "args": {"sheet": "Assumptions", "range": "B11"}, "reason": "Read current WACC value"}]

User: "Add a new row with Q1 data: Revenue=1000, Costs=500"
Plan: [{"step": 1, "tool": "add_row", "args": {"sheet": "Revenue", "row_num": 12, "values": {"A": "Q1", "B": 1000, "C": 500}}, "reason": "Add Q1 data row"}]

User: "Show me a summary of revenue by year"
Plan: [{"step": 1, "tool": "pivot_table", "args": {"source_sheet": "Revenue", "target_sheet": "Revenue_Summary", "row_field": "Year", "value_field": "Revenue", "agg": "sum"}, "reason": "Aggregate revenue by year into summary sheet"}]

User: "Check if all formulas are consistent"
Plan: [{"step": 1, "tool": "validate_formulas", "args": {}, "reason": "Run formula consistency validation"}]

User: "Change Revenue Growth Rate to 20% and show me the impact on Net Income in P&L"
Thinking: Revenue Growth Rate is in Assumptions!B4. Changing it will cascade through Revenue→COGS→P&L via XREF. I should write the value then read the downstream result.
Plan: [{"step": 1, "tool": "write_range", "args": {"sheet": "Assumptions", "range": "B4", "values": [0.20]}, "reason": "Update Revenue Growth Rate to 20%"}, {"step": 2, "tool": "read_range", "args": {"sheet": "P&L", "range": "B15:K15"}, "reason": "Read Net Income row to show downstream impact of the assumption change"}]
"""


KPI_KEYWORDS = [
    'revenue', 'net income', 'ebit', 'ebitda', 'gross profit', 'gross margin',
    'net margin', 'ebit margin', 'operating income', 'free cash flow', 'fcf',
    'enterprise value', 'equity value', 'price per share', 'eps', 'npv', 'irr',
    'total assets', 'total liabilities', 'cash', 'wacc', 'terminal value',
]


class ExcelAgent:
    def __init__(
        self,
        ollama: OllamaClient,
        retriever: RAGRetriever,
        tools: ExcelTools,
        workbook_data: Optional[WorkbookData] = None,
        scenario_manager: Optional[Any] = None,
    ) -> None:
        self.ollama = ollama
        self.retriever = retriever
        self.tools = tools
        self.workbook_data = workbook_data or tools.workbook_data
        self.scenario_manager = scenario_manager

    def _compute_impact(self, plan: List[Dict]) -> Dict:
        """Traverse dependency graph from write_range targets to estimate impact on KPIs.
        Returns dict with: cells_affected (int), kpis (list of {label, cell, current_value, note}).
        """
        graph = self.tools.graph
        wb = self.workbook_data

        # 1. Collect all cells being written
        written_cells = set()
        written_values: Dict[str, Any] = {}
        for step in plan:
            if step.get("tool") not in ("write_range", "write_formula"):
                continue
            args = step.get("args", {})
            sheet = args.get("sheet", "")
            rng = args.get("range") or args.get("cell", "")
            if sheet and rng:
                addr = f"{sheet}!{rng.upper()}"
                written_cells.add(addr)
                vals = args.get("values")
                if isinstance(vals, list) and vals:
                    written_values[addr] = vals[0]
                elif isinstance(vals, dict):
                    written_values.update(vals)
                elif args.get("formula"):
                    written_values[addr] = f"={args['formula']}"

        if not written_cells:
            return {"cells_affected": 0, "kpis": [], "written": []}

        # 2. BFS forward through dependency graph to find all downstream cells
        affected: set = set()
        queue = list(written_cells)
        visited = set(written_cells)
        while queue:
            node = queue.pop(0)
            for successor in graph.successors(node) if graph.has_node(node) else []:
                if successor not in visited:
                    visited.add(successor)
                    affected.add(successor)
                    queue.append(successor)

        # 3. Among affected cells, find KPI-labelled ones (check row-label in col A)
        kpis = []
        seen_labels = set()
        for addr in sorted(affected):
            cd = wb.cells.get(addr)
            if not cd or cd.value is None:
                continue
            # Try to find row label (col A of same row)
            label_addr = f"{cd.sheet_name}!A{cd.row}"
            label_cd = wb.cells.get(label_addr)
            label = str(label_cd.value).strip() if label_cd and label_cd.value else ""
            label_lower = label.lower()

            is_kpi = any(kw in label_lower for kw in KPI_KEYWORDS)
            if not is_kpi or label_lower in seen_labels:
                continue
            seen_labels.add(label_lower)

            current_val = cd.value
            formatted = None
            if isinstance(current_val, float):
                if abs(current_val) < 1:
                    formatted = f"{current_val:.1%}"
                elif abs(current_val) > 1_000_000:
                    formatted = f"${current_val/1_000_000:.1f}M"
                elif abs(current_val) > 1_000:
                    formatted = f"${current_val:,.0f}"
                else:
                    formatted = f"{current_val:.2f}"
            elif isinstance(current_val, int):
                formatted = f"{current_val:,}" if abs(current_val) > 999 else str(current_val)
            else:
                formatted = str(current_val)[:20]

            kpis.append({
                "label": label,
                "cell": addr,
                "current_value": formatted,
                "sheet": cd.sheet_name,
            })
            if len(kpis) >= 8:
                break

        # 4. Summarise written cells for display
        written_summary = []
        for addr, val in written_values.items():
            written_summary.append({"cell": addr, "new_value": str(val)[:30]})

        return {
            "cells_affected": len(affected),
            "kpis": kpis,
            "written": written_summary,
        }

    def build_cell_index(self, max_rows_per_sheet: int = 40) -> str:
        """Build a SheetCompressor-style cell index with 3-layer context.

        Layer 1 — Local: cell labels + adjacent values per row
        Layer 2 — Table: structural anchors (header rows, boundary rows like SUM/TOTAL)
        Layer 3 — Document: cross-sheet formula refs, named ranges

        Fits ~3500 chars for typical financial models.
        """
        from openpyxl.utils import get_column_letter
        wb = self.workbook_data
        lines: List[str] = [f"CELL INDEX | Sheets: {', '.join(wb.sheets)}"]

        cross_sheet_refs: List[str] = []

        for sheet_name in wb.sheets:
            sheet_cells = [
                c for c in wb.cells.values()
                if c.sheet_name == sheet_name and not c.is_hidden
            ]
            if not sheet_cells:
                continue

            max_row = max((c.row for c in sheet_cells), default=0)
            max_col = max((c.col for c in sheet_cells), default=0)
            max_col = min(max_col, 20)

            cell_lookup: Dict[tuple, Any] = {}
            for c in sheet_cells:
                if c.col <= max_col:
                    cell_lookup[(c.row, c.col)] = c

            lines.append(f"\n[{sheet_name}] {max_row}r x {max_col}c")

            # --- Detect header row (first row with 2+ text values) ---
            header_row = 1
            for r in range(1, min(5, max_row + 1)):
                text_count = sum(
                    1 for c in range(1, max_col + 1)
                    if cell_lookup.get((r, c)) and isinstance(cell_lookup[(r, c)].value, str)
                )
                if text_count >= 2:
                    header_row = r
                    break

            # --- Structural anchors: header row ---
            hdrs = []
            for c in range(1, max_col + 1):
                cd = cell_lookup.get((header_row, c))
                if cd and cd.value is not None:
                    hdrs.append(f"{get_column_letter(c)}={str(cd.value)[:15]}")
            if hdrs:
                lines.append(f"H(r{header_row}): {', '.join(hdrs)}")

            # --- Layer 1: Row labels + values ---
            display_rows = min(max_row, max_rows_per_sheet)
            row_entries = []
            boundary_rows = []
            for r in range(header_row + 1, display_rows + 1):
                label_cell = cell_lookup.get((r, 1))
                if label_cell and label_cell.value is not None:
                    label = str(label_cell.value)[:18]

                    # Detect boundary rows (Total, Subtotal, SUM)
                    label_lower = label.lower()
                    is_boundary = any(kw in label_lower for kw in ['total', 'subtotal', 'net ', 'gross '])

                    val_info = ""
                    for vc in range(2, min(max_col + 1, 4)):
                        vcd = cell_lookup.get((r, vc))
                        if vcd:
                            cl = get_column_letter(vc)
                            if vcd.formula:
                                val_info = f"{cl}{r}=f"
                            elif vcd.value is not None:
                                v = vcd.value
                                vs = str(round(v, 4)) if isinstance(v, float) else str(v)[:12]
                                val_info = f"{cl}{r}={vs}"
                            break

                    entry = f"r{r}:A={label}" + (f",{val_info}" if val_info else "")
                    if is_boundary:
                        entry = f"*{entry}"
                        boundary_rows.append(r)
                    row_entries.append(entry)

            if row_entries:
                lines.append("DATA: " + " | ".join(row_entries))

            # --- Layer 2: Table boundaries ---
            if boundary_rows:
                lines.append(f"BOUNDS: rows {','.join(str(r) for r in boundary_rows[:6])}")

            # --- Formula patterns (dominant per column) ---
            formula_cols: Dict[int, Dict[str, List[int]]] = defaultdict(lambda: defaultdict(list))
            for c in sheet_cells:
                if c.formula and c.col <= max_col:
                    pattern = re.sub(r'\d+', 'N', c.formula)
                    formula_cols[c.col][pattern].append(c.row)

                    # Layer 3: detect cross-sheet references
                    if '!' in c.formula and len(cross_sheet_refs) < 20:
                        ref_match = re.search(r"([A-Za-z ]+)!", c.formula)
                        if ref_match:
                            ref_sheet = ref_match.group(1).strip("'")
                            cross_sheet_refs.append(f"{c.cell_address}→{ref_sheet}")

            fp_lines = []
            for col_idx in sorted(formula_cols.keys()):
                cl = get_column_letter(col_idx)
                dominant = max(formula_cols[col_idx].items(), key=lambda x: len(x[1]))
                pat, rows = dominant
                if len(rows) >= 2:
                    fp_lines.append(f"{cl}:={pat}(r{min(rows)}-{max(rows)})")
            if fp_lines:
                lines.append(f"FP: {', '.join(fp_lines[:8])}")

            # --- Named ranges ---
            named = [(c.named_range, c.cell_address) for c in sheet_cells if c.named_range]
            if named:
                lines.append(f"NR: {', '.join(f'{n}={a}' for n, a in named[:10])}")

        # --- Layer 3: Cross-sheet formula references ---
        if cross_sheet_refs:
            lines.append(f"\nXREF: {' | '.join(cross_sheet_refs[:20])}")

        result = "\n".join(lines)
        max_chars = 5000
        if len(result) > max_chars:
            result = result[:max_chars] + "\n...(truncated)"
        return result

    def _try_parse_explicit_table(self, query: str) -> Optional[List[Dict]]:
        """If the query explicitly lists cell=value assignments, build a plan directly.
        Handles patterns like: Create sheet 'X'. Write A1=Year B1=Revenue A2=2021 B2=144
        Returns a plan list or None if pattern not detected.
        """
        # Detect create sheet
        sheet_match = re.search(r"(?:create|new)\s+sheet\s+['\"]?(\w[\w\s]*?)['\"]?[\.\s]", query, re.IGNORECASE)
        sheet_name = sheet_match.group(1).strip() if sheet_match else None

        # Find all CELL=VALUE pairs: e.g. A1=Year, B2=155.5, C3=0.08
        cell_val_pattern = re.compile(r'\b([A-Z]{1,3}\d{1,4})=([^\s]+)', re.IGNORECASE)
        pairs = cell_val_pattern.findall(query)
        if len(pairs) < 3:
            return None  # Not enough explicit assignments — let LLM handle it

        steps: List[Dict] = []
        step_num = 0

        if sheet_name:
            step_num += 1
            steps.append({"step": step_num, "tool": "create_sheet", "args": {"name": sheet_name}, "reason": f"Create sheet {sheet_name}"})

        sheet = sheet_name or (self.workbook_data.sheets[0] if self.workbook_data.sheets else "Sheet1")

        # Coerce values
        def coerce(v: str):
            try:
                f = float(v)
                return int(f) if f == int(f) else f
            except ValueError:
                return v.strip("'\"")

        # Group by row to write row-by-row (more efficient)
        from collections import defaultdict
        rows: dict = defaultdict(dict)
        col_order: dict = defaultdict(list)
        for cell, val in pairs:
            col_match = re.match(r'^([A-Za-z]+)(\d+)$', cell)
            if col_match:
                col = col_match.group(1).upper()
                row = int(col_match.group(2))
                rows[row][col] = coerce(val)
                if col not in col_order[row]:
                    col_order[row].append(col)

        from openpyxl.utils import column_index_from_string, get_column_letter
        for row_num in sorted(rows.keys()):
            cols_in_row = sorted(rows[row_num].keys(), key=lambda c: column_index_from_string(c))
            if not cols_in_row:
                continue
            # Check if columns are contiguous
            col_indices = [column_index_from_string(c) for c in cols_in_row]
            is_contiguous = (max(col_indices) - min(col_indices) + 1) == len(col_indices)

            if is_contiguous and len(cols_in_row) > 1:
                start_col = get_column_letter(min(col_indices))
                end_col = get_column_letter(max(col_indices))
                range_str = f"{start_col}{row_num}:{end_col}{row_num}"
                vals = [rows[row_num][c] for c in cols_in_row]
                step_num += 1
                steps.append({"step": step_num, "tool": "write_range",
                               "args": {"sheet": sheet, "range": range_str, "values": vals},
                               "reason": f"Write row {row_num}"})
            else:
                for col in cols_in_row:
                    step_num += 1
                    steps.append({"step": step_num, "tool": "write_range",
                                  "args": {"sheet": sheet, "range": f"{col}{row_num}", "values": [rows[row_num][col]]},
                                  "reason": f"Write {col}{row_num}"})

        # Also detect explicit formula patterns: "write formula =SUM(A1:A5) in B1" or "formula =X in CELL"
        formula_pattern = re.compile(
            r'(?:write\s+)?formula\s+=?([A-Z]+\([^)]+\))\s+in\s+([A-Z]{1,3}\d{1,4})',
            re.IGNORECASE
        )
        for fm in formula_pattern.finditer(query):
            formula_str = fm.group(1).strip()
            cell_addr = fm.group(2).upper()
            step_num += 1
            steps.append({"step": step_num, "tool": "write_formula",
                           "args": {"sheet": sheet, "cell": cell_addr, "formula": formula_str},
                           "reason": f"Write formula {formula_str} in {cell_addr}"})

        logger.info(f"_try_parse_explicit_table: built {len(steps)} steps from {len(pairs)} cell assignments")
        return steps

    async def plan(self, query: str, context: str, cell_index: str) -> List[Dict]:
        # Fast path: if query has explicit CELL=VALUE assignments, build plan directly
        explicit_plan = self._try_parse_explicit_table(query)
        if explicit_plan:
            try:
                plan_model = PlanModel(steps=[ToolStepModel(**s) for s in explicit_plan])
                logger.info(f"Explicit table plan: {len(plan_model.steps)} steps (bypassed LLM)")
                return [s.dict() for s in plan_model.steps]
            except Exception as e:
                logger.warning(f"Explicit plan validation failed: {e}, falling back to LLM")

        system = (
            SYSTEM_PROMPT
            + "\n" + cell_index
            + "\n\nReturn ONLY a valid JSON array of steps. Each step has: "
            "step (int), tool (string), args (object), reason (string). "
            "No markdown, no explanation, just the JSON array. "
            "IMPORTANT: Use EXACT sheet names and cell addresses from the CELL INDEX above."
        )
        messages: List[Dict[str, str]] = [
            {
                "role": "user",
                "content": (
                    f"RAG Context:\n{context}\n\n"
                    f"User request: {query}\n\n"
                    "Return ONLY a JSON array. Format example for writing a table row by row:\n"
                    '[{"step":1,"tool":"create_sheet","args":{"name":"Sheet1"},"reason":"..."},'
                    '{"step":2,"tool":"write_range","args":{"sheet":"Sheet1","range":"A1:C1","values":["Year","Revenue","Growth"]},"reason":"header row"},'
                    '{"step":3,"tool":"write_range","args":{"sheet":"Sheet1","range":"A2:C2","values":[2021,144,0]},"reason":"data row 1"},'
                    '{"step":4,"tool":"write_formula","args":{"sheet":"Sheet1","cell":"B1","formula":"SUM(A1:A5)"},"reason":"formula"}]\n'
                    "CRITICAL RULES:\n"
                    "1. NEVER use write_table — always use write_range with one row at a time.\n"
                    "2. args MUST use exact keys: name / sheet / range / values / cell / formula.\n"
                    "3. values is always a flat list (one row of values per step).\n"
                    "4. No markdown fences. Output ONLY the JSON array."
                ),
            }
        ]

        last_error: Optional[str] = None
        for attempt in range(3):
            try:
                response = await self.ollama.chat(
                    messages=messages, system=system, temperature=0.0,
                )
                parsed = self._parse_plan_json(response)
                logger.info(f"_parse_plan_json: {len(parsed)} raw steps: {[s.get('tool') for s in parsed]}")
                parsed = self._normalize_plan(parsed)
                logger.info(f"_normalize_plan: {len(parsed)} steps: {[s.get('tool') for s in parsed]}")
                plan_model = PlanModel(steps=[ToolStepModel(**s) for s in parsed])
                logger.info(f"Plan validated OK (attempt {attempt + 1}): {len(plan_model.steps)} steps")
                return [s.dict() for s in plan_model.steps]
            except Exception as e:
                last_error = str(e)
                resp_preview = repr(response[:200]) if 'response' in dir() and response else repr(response if 'response' in dir() else "<unset>")
                logger.warning(f"Plan validation failed (attempt {attempt + 1}/3): {last_error} | response={resp_preview}")
                messages.append({"role": "assistant", "content": response if 'response' in dir() else ""})
                messages.append({
                    "role": "user",
                    "content": (
                        f"Your response was invalid. Error: {last_error}\n"
                        "Fix the JSON and try again. Return ONLY a valid JSON array."
                    ),
                })

        logger.error(f"Plan generation failed after 3 attempts: {last_error}")
        fallback_sheet = self.workbook_data.sheets[0] if self.workbook_data.sheets else "Sheet1"
        return [{"step": 1, "tool": "read_range", "args": {"sheet": fallback_sheet, "range": "A1"}, "reason": f"Fallback: {last_error}"}]

    def _parse_plan_json(self, response: str) -> List[Dict]:
        """Extract JSON array from LLM response, stripping markdown fences."""
        cleaned = response.strip()
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.MULTILINE)
        cleaned = re.sub(r"\s*```$", "", cleaned, flags=re.MULTILINE)
        cleaned = cleaned.strip()

        # Find start of outer JSON array using bracket matching
        start = cleaned.find('[')
        candidate = cleaned[start:] if start != -1 else cleaned

        # 1) Try full parse
        try:
            data = json.loads(candidate)
            if isinstance(data, list):
                return data
            if isinstance(data, dict) and "steps" in data:
                return data["steps"]
            return [data]
        except json.JSONDecodeError:
            pass

        # 2) Truncated recovery: strip incomplete last step objects one by one
        # Find last complete '}' at top level (depth 1 inside outer array)
        # Strategy: keep chopping from last '}' until we can close with ']'
        probe = candidate
        for _ in range(20):
            last_brace = probe.rfind('}')
            if last_brace == -1:
                break
            attempt = probe[:last_brace + 1] + ']'
            try:
                data = json.loads(attempt)
                if isinstance(data, list) and data:
                    return data
            except json.JSONDecodeError:
                pass
            probe = probe[:last_brace]

        # 3) Last resort: extract complete step objects via nested-brace matching
        steps = []
        i = 0
        while i < len(candidate):
            if candidate[i] == '{':
                depth, j = 0, i
                while j < len(candidate):
                    if candidate[j] == '{': depth += 1
                    elif candidate[j] == '}':
                        depth -= 1
                        if depth == 0:
                            try:
                                obj = json.loads(candidate[i:j+1])
                                if "tool" in obj:
                                    steps.append(obj)
                            except json.JSONDecodeError:
                                pass
                            break
                    j += 1
                i = j + 1
            else:
                i += 1
        if steps:
            return steps

        raise json.JSONDecodeError("Could not parse plan JSON", candidate, 0)

    def _validate_step_args(self, step: Dict) -> Optional[ToolResult]:
        """Pre-validate step arguments. Returns error ToolResult or None if OK."""
        tool_name = step.get("tool", "")
        args = step.get("args", {})

        # --- Sheet name validation with fuzzy matching ---
        # Skip validation for sheets that will be created (target_sheet in pivot_table, name in create_sheet)
        skip_keys: set = set()
        if tool_name == "pivot_table":
            skip_keys.add("target_sheet")
        if tool_name == "create_sheet":
            skip_keys.add("name")

        sheet_keys = ["sheet", "source_sheet", "target_sheet"]
        for key in sheet_keys:
            if key in skip_keys:
                continue
            sheet = args.get(key)
            if sheet and sheet not in self.workbook_data.sheets:
                closest = difflib.get_close_matches(sheet, self.workbook_data.sheets, n=1, cutoff=0.5)
                if closest:
                    old_sheet = sheet
                    args[key] = closest[0]
                    logger.info(f"Auto-corrected sheet name '{old_sheet}' -> '{closest[0]}'")
                else:
                    return ToolResult(
                        tool_name=tool_name, success=False, data=None,
                        error=f"Sheet '{sheet}' not found. Available sheets: {self.workbook_data.sheets}",
                    )

        # --- Cell address format validation ---
        cell_range = args.get("range") or args.get("cell")
        if cell_range and tool_name in ('read_range', 'write_range', 'write_formula', 'delete_range'):
            if not CELL_ADDR_RE.match(cell_range.strip()):
                return ToolResult(
                    tool_name=tool_name, success=False, data=None,
                    error=f"Invalid cell address format: '{cell_range}'. Expected e.g. 'B4' or 'A1:D10'.",
                )

        # --- values must be a list for write_range ---
        if tool_name == "write_range":
            values = args.get("values")
            if values is not None and not isinstance(values, list):
                args["values"] = [values]
                logger.info(f"Auto-wrapped scalar value into list: {args['values']}")

        return None

    def _normalize_plan(self, steps: List[Dict]) -> List[Dict]:
        """Normalize all steps in a plan before Pydantic validation."""
        from openpyxl.utils import get_column_letter
        last_created_sheet: Optional[str] = None
        normalized = []
        step_counter = 0

        for i, s in enumerate(steps):
            if not isinstance(s, dict):
                continue
            if "reason" not in s:
                s["reason"] = ""
            tool = s.get("tool", "")
            args = s.get("args", {})
            if not isinstance(args, dict):
                args = {}

            # write_table / set_table aliases → expand into individual write_range steps
            if tool in ("write_table", "set_table", "fill_table", "write_data"):
                args = self._normalize_args("write_range", args)
                sheet = args.get("sheet") or last_created_sheet or "Sheet1"
                values_2d = args.get("values", [])
                range_str = args.get("range", "A1")
                # Parse start cell
                start_match = re.match(r"^([A-Za-z]+)(\d+)", range_str)
                start_col = 1
                start_row = 1
                if start_match:
                    for ci, ch in enumerate(start_match.group(1).upper()):
                        start_col = start_col * 26 + (ord(ch) - ord('A') + 1)
                    start_col -= 26  # undo extra mult on first iter
                    start_row = int(start_match.group(2))
                    # Simple single-letter col
                    col_str = start_match.group(1).upper()
                    start_col = sum((ord(c) - ord('A') + 1) * (26 ** (len(col_str)-1-j)) for j, c in enumerate(col_str))
                if isinstance(values_2d, list) and values_2d:
                    for ri, row_vals in enumerate(values_2d):
                        if not isinstance(row_vals, list):
                            row_vals = [row_vals]
                        for ci, val in enumerate(row_vals):
                            col_letter = get_column_letter(start_col + ci)
                            cell_addr = f"{col_letter}{start_row + ri}"
                            step_counter += 1
                            normalized.append({
                                "step": step_counter,
                                "tool": "write_range",
                                "args": {"sheet": sheet, "range": cell_addr, "values": [val]},
                                "reason": s.get("reason", ""),
                            })
                continue

            args = self._normalize_args(tool, args)
            # Track last created sheet
            if tool == "create_sheet":
                name = args.get("name")
                if name:
                    last_created_sheet = name
            # Inject missing sheet
            if tool in ("write_range", "write_formula", "read_range", "delete_range"):
                if "sheet" not in args:
                    candidate = args.pop("name", None) or last_created_sheet
                    if candidate:
                        args["sheet"] = candidate
                        logger.info(f"Auto-injected sheet '{candidate}' into step {i+1}")
            s["args"] = args
            step_counter += 1
            s["step"] = step_counter
            normalized.append(s)
        return normalized

    def _normalize_args(self, tool_name: str, args: Dict) -> Dict:
        """Normalize arg key aliases from local models that use non-standard key names."""
        # create_sheet: name — ONLY for create_sheet, never steal 'sheet' from write ops
        if tool_name == "create_sheet":
            for alt in ("sheetName", "sheet_name"):
                if "name" not in args and alt in args:
                    args["name"] = args.pop(alt)
                    break
        # write_range / read_range / delete_range / write_formula: range or cell
        range_alts = ("cellAddress", "cellIndex", "cell_range", "cellRange",
                      "cells", "cell_address", "address", "cell_ref", "cellRef")
        if tool_name == "write_formula":
            for alt in range_alts:
                if "cell" not in args and alt in args:
                    args["cell"] = args.pop(alt)
                    break
        else:
            for alt in range_alts:
                if "range" not in args and alt in args:
                    args["range"] = args.pop(alt)
                    break
        # sheet key aliases
        for alt in ("sheetName", "sheet_name", "worksheet", "sheetname"):
            if "sheet" not in args and alt in args:
                args["sheet"] = args.pop(alt)
                break
        # write_range: values aliases + ensure list + coerce string numerics
        if tool_name == "write_range":
            for alt in ("value", "data", "cell_values", "cellValues"):
                if "values" not in args and alt in args:
                    args["values"] = args.pop(alt)
                    break
            if "values" in args and not isinstance(args["values"], list):
                args["values"] = [args["values"]]
            if "values" in args:
                coerced = []
                for v in args["values"]:
                    if isinstance(v, str) and not v.startswith("="):
                        try:
                            coerced.append(int(v) if "." not in v else float(v))
                        except (ValueError, TypeError):
                            coerced.append(v)
                    else:
                        coerced.append(v)
                args["values"] = coerced
        # write_formula: strip leading = if present (model sometimes adds it)
        if tool_name == "write_formula" and "formula" in args:
            f = args["formula"]
            if isinstance(f, str) and f.startswith("="):
                args["formula"] = f[1:]
        return args

    def execute_step(self, step: Dict) -> ToolResult:
        tool_name = step.get("tool", "")
        args = self._normalize_args(tool_name, step.get("args", {}))
        step["args"] = args

        validation_error = self._validate_step_args(step)
        if validation_error is not None:
            return validation_error

        dispatch = {
            "read_range": lambda: self.tools.read_range(
                args.get("sheet", "Sheet1"), args.get("range", "A1")
            ),
            "write_range": lambda: self.tools.write_range(
                args.get("sheet", "Sheet1"),
                args.get("range", "A1"),
                args.get("values", {}),
            ),
            "write_formula": lambda: self.tools.write_formula(
                args.get("sheet", "Sheet1"),
                args.get("cell", "A1"),
                args.get("formula", ""),
            ),
            "add_column": lambda: self.tools.add_column(
                args.get("sheet", "Sheet1"),
                args.get("after_col", "A"),
                args.get("header", ""),
                args.get("values"),
            ),
            "add_row": lambda: self.tools.add_row(
                args.get("sheet", "Sheet1"),
                args.get("row_num", 1),
                args.get("values", {}),
            ),
            "create_sheet": lambda: self.tools.create_sheet(
                args.get("name", "NewSheet"),
            ),
            "delete_range": lambda: self.tools.delete_range(
                args.get("sheet", "Sheet1"),
                args.get("range", "A1"),
                args.get("confirmed", False),
            ),
            "get_dependencies": lambda: self.tools.get_dependencies(args.get("cell", "")),
            "find_anomalies": lambda: self.tools.find_anomalies(args.get("sheet")),
            "explain_formula": lambda: self.tools.explain_formula(args.get("cell", "")),
            "generate_change_log": lambda: self.tools.generate_change_log(args.get("changes", [])),
            "get_audit_trail": lambda: self.tools.get_audit_trail(),
            "pivot_table": lambda: self.tools.pivot_table(
                args.get("source_sheet", "Sheet1"),
                args.get("target_sheet", "Summary"),
                args.get("row_field", ""),
                args.get("col_field"),
                args.get("value_field", ""),
                args.get("agg", "sum"),
            ),
            "validate_formulas": lambda: self.tools.validate_formulas(),
            "run_sensitivity": lambda: self._run_sensitivity(args),
            "run_integrity_check": lambda: self._run_integrity_check(),
            "run_smart_suggestions": lambda: self._run_smart_suggestions(),
            "create_scenario": lambda: self._create_scenario(args),
            "compare_scenarios": lambda: self._compare_scenarios(),
        }

        fn = dispatch.get(tool_name)
        if fn is None:
            return ToolResult(
                tool_name=tool_name,
                success=False,
                data=None,
                error=f"Unknown tool: {tool_name}",
            )
        return fn()

    def _run_sensitivity(self, args: Dict) -> ToolResult:
        try:
            analyzer = SensitivityAnalyzer(self.tools.graph, self.tools.workbook_data)
            report = analyzer.run(
                max_inputs=args.get("max_inputs", 15),
                max_outputs=args.get("max_outputs", 8),
            )
            return ToolResult(
                tool_name="run_sensitivity",
                success=True,
                data={
                    "input_cells_tested": report.input_cells_tested,
                    "output_cells_monitored": report.output_cells_monitored,
                    "top_drivers": report.top_drivers,
                    "tornado_chart_data": report.tornado_chart_data,
                    "data_points": len(report.results),
                },
            )
        except Exception as e:
            logger.error(f"run_sensitivity failed: {e}")
            return ToolResult(tool_name="run_sensitivity", success=False, data=None, error=str(e))

    def _run_integrity_check(self) -> ToolResult:
        try:
            checker = IntegrityChecker(self.tools.graph, self.tools.workbook_data)
            report = checker.run()
            return ToolResult(
                tool_name="run_integrity_check",
                success=True,
                data={
                    "model_health_score": report.model_health_score,
                    "total_issues": report.total_issues,
                    "critical": report.critical,
                    "warning": report.warning,
                    "info": report.info,
                    "summary": report.summary,
                    "issues": [
                        {"severity": i.severity, "category": i.category, "cell": i.cell,
                         "message": i.message, "suggestion": i.suggestion}
                        for i in report.issues[:20]
                    ],
                },
            )
        except Exception as e:
            logger.error(f"run_integrity_check failed: {e}")
            return ToolResult(tool_name="run_integrity_check", success=False, data=None, error=str(e))

    def _run_smart_suggestions(self) -> ToolResult:
        try:
            engine = SmartSuggestionsEngine(self.tools.graph, self.tools.workbook_data)
            report = engine.run()
            return ToolResult(
                tool_name="run_smart_suggestions",
                success=True,
                data={
                    "model_maturity_score": report.model_maturity_score,
                    "total_suggestions": report.total_suggestions,
                    "suggestions": [
                        {"priority": s.priority, "category": s.category, "title": s.title,
                         "description": s.description, "suggested_action": s.suggested_action,
                         "affected_cells": s.affected_cells[:5]}
                        for s in report.suggestions[:15]
                    ],
                },
            )
        except Exception as e:
            logger.error(f"run_smart_suggestions failed: {e}")
            return ToolResult(tool_name="run_smart_suggestions", success=False, data=None, error=str(e))

    def _create_scenario(self, args: Dict) -> ToolResult:
        try:
            if not self.scenario_manager:
                self.scenario_manager = ScenarioManager(self.tools.graph, self.tools.workbook_data)
            mgr = self.scenario_manager
            if "Base Case" not in mgr.scenarios:
                mgr.create_base_case()
            name = args.get("name", "Custom")
            desc = args.get("description", "")
            pct = args.get("perturbation_pct")
            if pct is not None:
                s = mgr.create_perturbation_scenario(name, desc, float(pct))
            else:
                overrides = args.get("input_overrides", {})
                s = mgr.create_scenario(name, desc, overrides)
            return ToolResult(
                tool_name="create_scenario",
                success=True,
                data={
                    "name": s.name,
                    "description": s.description,
                    "input_count": len(s.input_overrides),
                    "output_count": len(s.computed_outputs),
                    "top_outputs": dict(list(s.computed_outputs.items())[:10]),
                },
            )
        except Exception as e:
            logger.error(f"create_scenario failed: {e}")
            return ToolResult(tool_name="create_scenario", success=False, data=None, error=str(e))

    def _compare_scenarios(self) -> ToolResult:
        try:
            if not self.scenario_manager:
                return ToolResult(
                    tool_name="compare_scenarios", success=False, data=None,
                    error="No scenarios created yet. Use create_scenario first.",
                )
            report = self.scenario_manager.compare()
            return ToolResult(
                tool_name="compare_scenarios",
                success=True,
                data={
                    "scenario_count": len(report.scenarios),
                    "summary": report.summary,
                    "comparisons": [
                        {"output_cell": c.output_cell, "base_value": c.base_value,
                         "deltas": c.deltas, "delta_pcts": c.delta_pcts}
                        for c in report.comparisons[:15]
                    ],
                },
            )
        except Exception as e:
            logger.error(f"compare_scenarios failed: {e}")
            return ToolResult(tool_name="compare_scenarios", success=False, data=None, error=str(e))

    async def reflect(self, step: Dict, result: ToolResult) -> tuple[bool, str]:
        system = (
            "You are a financial model safety checker. "
            "Evaluate whether a proposed change to an Excel financial model is reasonable. "
            "Return ONLY valid JSON: {\"ok\": true/false, \"concern\": \"explanation\"}"
        )
        step_desc = json.dumps(step, default=str)
        result_desc = json.dumps(result.data, default=str)[:500]
        messages = [
            {
                "role": "user",
                "content": (
                    f"Step executed: {step_desc}\n"
                    f"Result: {result_desc}\n\n"
                    "Is this change reasonable for a financial model? "
                    "Flag if values jump unreasonably (e.g. IRR changes by >50pp, values become negative unexpectedly)."
                ),
            }
        ]
        try:
            response = await self.ollama.chat(messages=messages, system=system, temperature=0.1)
            cleaned = re.sub(r"```(?:json)?", "", response).strip()
            data = json.loads(cleaned)
            ok = bool(data.get("ok", True))
            concern = str(data.get("concern", ""))
            return ok, concern
        except Exception as e:
            logger.warning(f"Reflect failed: {e}")
            return True, ""

    async def execute(
        self,
        query: str,
        workbook_uuid: str,
        approved_plan: Optional[List[Dict]] = None,
        mode: str = "plan",
    ) -> AsyncGenerator[Dict, None]:
        """Execute a query. mode='agent' auto-executes, mode='plan' stops for approval, mode='ask' read-only Q&A."""
        yield {"event": "retrieving", "data": {"message": "Retrieving relevant context..."}}

        try:
            chunks = await self.retriever.retrieve(query, workbook_uuid, k=15, workbook_data=self.workbook_data)
            context = self.retriever.build_context(chunks, query, max_chars=12000)
        except Exception as e:
            logger.error(f"Retrieval failed: {e}")
            context = "No context available."
            chunks = []

        cell_index = self.build_cell_index()

        # --- Ask mode: read-only Q&A, no planning, no tool execution ---
        if mode == "ask":
            yield {"event": "answer_start", "data": {}}
            ask_system = (
                "You are a read-only Excel AI assistant. You answer questions about the workbook "
                "but you NEVER modify any cells or data. Respond in the same language as the user. "
                "Be concise and precise. Use the provided context and cell index."
            )
            ask_messages = [
                {
                    "role": "user",
                    "content": (
                        f"Workbook context:\n{context}\n\n"
                        f"Cell index:\n{cell_index}\n\n"
                        f"Question: {query}"
                    ),
                }
            ]
            accumulated = ""
            async for chunk in self.ollama.stream_chat(messages=ask_messages, system=ask_system):
                accumulated += chunk
                yield {"event": "answer", "data": {"chunk": chunk}}
            yield {"event": "done", "data": {"message": "Complete", "full_answer": accumulated}}
            return

        if approved_plan is None:
            yield {"event": "planning", "data": {"message": "Generating execution plan..."}}
            plan = await self.plan(query, context, cell_index)

            if mode == "plan":
                impact = self._compute_impact(plan)
                yield {"event": "plan_ready", "data": {"plan": plan, "requiresApproval": True, "impact_preview": impact}}
                return
            # Agent mode: auto-execute, no approval needed
        else:
            plan = self._normalize_plan(approved_plan)

        yield {"event": "executing_plan", "data": {"message": f"Executing {len(plan)} steps...", "step_count": len(plan)}}

        execution_results = []
        for step in plan:
            step_num = step.get("step", "?")
            tool_name = step.get("tool", "")
            reason = step.get("reason", "")
            args = step.get("args", {})

            yield {
                "event": "executing",
                "data": {
                    "step": step_num,
                    "tool": tool_name,
                    "args": args,
                    "reason": reason,
                    "status": "running",
                },
            }

            result = self.execute_step(step)

            # --- Post-execution verify + auto-retry for write operations ---
            if result.success and tool_name in ('write_range', 'write_formula'):
                verify_sheet = args.get('sheet', '')
                verify_addr = args.get('range') or args.get('cell', '')
                if verify_sheet and verify_addr:
                    full_addr = f"{verify_sheet}!{verify_addr.upper()}"
                    state_data = self.tools.workbook_state.get(full_addr)
                    cell_data = self.tools.workbook_data.cells.get(full_addr)

                    if state_data is None and cell_data is None:
                        logger.warning(f"Post-verify FAIL: {full_addr} not found — retrying step")
                        retry_result = self.execute_step(step)
                        if retry_result.success:
                            result = retry_result
                            logger.info(f"Post-verify retry OK: {full_addr}")
                        else:
                            logger.error(f"Post-verify retry FAILED: {full_addr}: {retry_result.error}")
                    else:
                        actual = state_data if state_data is not None else (cell_data.value if cell_data else None)
                        expected = args.get('values', [None])[0] if tool_name == 'write_range' else args.get('formula', '')
                        if expected is not None and actual is not None:
                            actual_str = str(actual.get('value', actual)) if isinstance(actual, dict) else str(actual)
                            if str(expected) not in actual_str and actual_str not in str(expected):
                                logger.warning(f"Post-verify MISMATCH: {full_addr} expected={expected} actual={actual_str}")
                            else:
                                logger.info(f"Post-verify OK: {full_addr} = {actual_str}")
                        else:
                            logger.info(f"Post-verify OK: {full_addr} = {actual}")

            execution_results.append({"step": step_num, "tool": tool_name, "success": result.success, "data": result.data, "error": result.error})

            yield {
                "event": "tool_result",
                "data": {
                    "step": step_num,
                    "tool": tool_name,
                    "args": args,
                    "success": result.success,
                    "data": result.data,
                    "error": result.error,
                    "status": "done" if result.success else "error",
                },
            }

        yield {"event": "answer_start", "data": {}}

        results_summary = json.dumps(execution_results, default=str)[:800]
        answer_system = (
            "You are an Excel AI assistant. Summarize tool execution results concisely. "
            "Respond in the same language as the user's question."
        )
        answer_messages = [
            {
                "role": "user",
                "content": (
                    f"Request: {query}\n\n"
                    f"Results: {results_summary}\n\n"
                    "Summarize what was done. If data was read, present it clearly. "
                    "If cells were written, confirm the values."
                ),
            }
        ]

        accumulated = ""
        async for chunk in self.ollama.stream_chat(
            messages=answer_messages, system=answer_system
        ):
            accumulated += chunk
            yield {"event": "answer", "data": {"chunk": chunk}}

        yield {"event": "done", "data": {"message": "Complete", "full_answer": accumulated}}
