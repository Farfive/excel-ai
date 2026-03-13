import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional

import networkx as nx

from parser.xlsx_parser import WorkbookData

logger = logging.getLogger(__name__)


@dataclass
class ToolResult:
    tool_name: str
    success: bool
    data: Any
    error: Optional[str] = None


class ExcelTools:
    def __init__(
        self,
        workbook_state: Dict[str, Any],
        graph: nx.DiGraph,
        workbook_data: WorkbookData,
    ) -> None:
        self.workbook_state = workbook_state
        self.graph = graph
        self.workbook_data = workbook_data
        self.audit_trail: List[Dict[str, Any]] = []

    def _log_action(self, tool: str, args: Dict[str, Any], result: ToolResult) -> None:
        self.audit_trail.append({
            "timestamp": datetime.utcnow().isoformat(),
            "tool": tool,
            "args": args,
            "success": result.success,
            "error": result.error,
        })

    def read_range(self, sheet: str, range_str: str) -> ToolResult:
        try:
            results = {}
            cells_to_read = self._resolve_range(sheet, range_str)
            for addr in cells_to_read:
                if addr in self.workbook_state:
                    results[addr] = self.workbook_state[addr]
                elif addr in self.workbook_data.cells:
                    cell = self.workbook_data.cells[addr]
                    results[addr] = {
                        "value": cell.value,
                        "formula": cell.formula,
                        "data_type": cell.data_type,
                    }
                else:
                    results[addr] = None
            return ToolResult(tool_name="read_range", success=True, data=results)
        except Exception as e:
            logger.error(f"read_range failed: {e}")
            return ToolResult(tool_name="read_range", success=False, data=None, error=str(e))

    def write_range(self, sheet: str, range_str: str, values: Any) -> ToolResult:
        try:
            from openpyxl.utils import get_column_letter, column_index_from_string
            topo_order = self.graph.graph.get("topological_order", [])
            topo_index = {node: i for i, node in enumerate(topo_order)}

            # Normalize values: list → zip with resolved cells; single value → broadcast
            if isinstance(values, list):
                resolved = self._resolve_range(sheet, range_str)
                # Flatten 2D list (list of rows) into 1D
                flat: List[Any] = []
                for item in values:
                    if isinstance(item, list):
                        flat.extend(item)
                    else:
                        flat.append(item)
                values_dict: Dict[str, Any] = {}
                for addr, val in zip(resolved, flat):
                    short_addr = addr.split("!")[-1]
                    values_dict[short_addr] = val
                values = values_dict
            elif not isinstance(values, dict):
                # scalar broadcast across entire range
                resolved = self._resolve_range(sheet, range_str)
                values = {addr.split("!")[-1]: values for addr in resolved}

            sorted_values = sorted(
                values.items(),
                key=lambda item: topo_index.get(item[0], 999999),
            )

            from parser.xlsx_parser import CellData, _infer_data_type

            instructions = []
            diff: List[Dict[str, Any]] = []

            for addr, val in sorted_values:
                # Build full address key e.g. "Assumptions!B4"
                full_addr = addr if "!" in addr else f"{sheet}!{addr}"

                # Capture BEFORE value
                before: Any = None
                if full_addr in self.workbook_data.cells:
                    before = self.workbook_data.cells[full_addr].value
                elif full_addr in self.workbook_state:
                    before = self.workbook_state[full_addr].get("value")

                instructions.append({
                    "cell": addr,
                    "value": val,
                    "topo_index": topo_index.get(addr, 999999),
                })

                diff.append({
                    "cell": full_addr,
                    "before": before,
                    "after": val,
                })

                # Detect formula strings (=...) and handle correctly
                is_formula = isinstance(val, str) and val.startswith("=")
                formula_str = val[1:] if is_formula else None
                stored_value = val if is_formula else val

                # Mutate workbook_data.cells so /grid reflects changes
                if full_addr in self.workbook_data.cells:
                    self.workbook_data.cells[full_addr].value = stored_value
                    self.workbook_data.cells[full_addr].formula = formula_str
                    self.workbook_data.cells[full_addr].data_type = _infer_data_type(stored_value, formula_str)
                    self.workbook_data.cells[full_addr].is_hardcoded = isinstance(val, (int, float))
                else:
                    col_part = "".join(c for c in addr if c.isalpha())
                    row_part = "".join(c for c in addr if c.isdigit())
                    if col_part and row_part:
                        col_idx = column_index_from_string(col_part)
                        row_idx = int(row_part)
                        self.workbook_data.cells[full_addr] = CellData(
                            cell_address=full_addr,
                            sheet_name=sheet,
                            value=stored_value,
                            formula=formula_str,
                            data_type=_infer_data_type(stored_value, formula_str),
                            named_range=None,
                            row=row_idx,
                            col=col_idx,
                            is_hardcoded=isinstance(val, (int, float)),
                            is_merged=False,
                            merge_master=None,
                        )

                self.workbook_state[full_addr] = {"value": stored_value, "formula": formula_str, "data_type": _infer_data_type(stored_value, formula_str)}

            return ToolResult(
                tool_name="write_range",
                success=True,
                data={"ordered_writes": instructions, "count": len(instructions), "diff": diff},
            )
        except Exception as e:
            logger.error(f"write_range failed: {e}")
            return ToolResult(tool_name="write_range", success=False, data=None, error=str(e))

    def get_dependencies(self, cell: str) -> ToolResult:
        try:
            if cell not in self.graph:
                return ToolResult(
                    tool_name="get_dependencies",
                    success=False,
                    data=None,
                    error=f"Cell {cell} not found in graph",
                )

            upstream_1 = list(self.graph.predecessors(cell))
            upstream_2 = []
            for u in upstream_1:
                upstream_2.extend(self.graph.predecessors(u))

            downstream_1 = list(self.graph.successors(cell))
            downstream_2 = []
            for d in downstream_1:
                downstream_2.extend(self.graph.successors(d))

            node_data = dict(self.graph.nodes[cell])

            def _enrich(addr: str) -> Dict:
                nd = dict(self.graph.nodes.get(addr, {}))
                return {
                    "address": addr,
                    "value": nd.get("value"),
                    "formula": nd.get("formula"),
                    "named_range": nd.get("named_range"),
                    "cluster_name": nd.get("cluster_name"),
                }

            return ToolResult(
                tool_name="get_dependencies",
                success=True,
                data={
                    "cell": cell,
                    "value": node_data.get("value"),
                    "formula": node_data.get("formula"),
                    "named_range": node_data.get("named_range"),
                    "cluster": node_data.get("cluster_name"),
                    "upstream_1hop": [_enrich(a) for a in upstream_1],
                    "upstream_2hop": [_enrich(a) for a in upstream_2],
                    "downstream_1hop": [_enrich(a) for a in downstream_1],
                    "downstream_2hop": [_enrich(a) for a in downstream_2],
                },
            )
        except Exception as e:
            logger.error(f"get_dependencies failed: {e}")
            return ToolResult(tool_name="get_dependencies", success=False, data=None, error=str(e))

    def find_anomalies(self, sheet: Optional[str] = None) -> ToolResult:
        try:
            anomalies = []
            for node in self.graph.nodes():
                nd = self.graph.nodes[node]
                if not nd.get("is_anomaly"):
                    continue
                if sheet and nd.get("sheet_name") != sheet:
                    continue
                anomalies.append({
                    "cell": node,
                    "value": nd.get("value"),
                    "anomaly_score": nd.get("anomaly_score", 0.0),
                    "formula": nd.get("formula"),
                    "sheet_name": nd.get("sheet_name"),
                    "named_range": nd.get("named_range"),
                })
            anomalies.sort(key=lambda x: x["anomaly_score"])
            return ToolResult(tool_name="find_anomalies", success=True, data={"anomalies": anomalies})
        except Exception as e:
            logger.error(f"find_anomalies failed: {e}")
            return ToolResult(tool_name="find_anomalies", success=False, data=None, error=str(e))

    def explain_formula(self, cell: str) -> ToolResult:
        try:
            if cell not in self.graph:
                if cell in self.workbook_data.cells:
                    cd = self.workbook_data.cells[cell]
                    return ToolResult(
                        tool_name="explain_formula",
                        success=True,
                        data={
                            "cell": cell,
                            "value": cd.value,
                            "formula": cd.formula,
                            "named_range": cd.named_range,
                            "depends_on": [],
                        },
                    )
                return ToolResult(
                    tool_name="explain_formula",
                    success=False,
                    data=None,
                    error=f"Cell {cell} not found",
                )

            nd = dict(self.graph.nodes[cell])
            depends_on = []
            for pred in self.graph.predecessors(cell):
                pnd = dict(self.graph.nodes.get(pred, {}))
                depends_on.append({
                    "address": pred,
                    "value": pnd.get("value"),
                    "named_range": pnd.get("named_range"),
                    "formula": pnd.get("formula"),
                })

            return ToolResult(
                tool_name="explain_formula",
                success=True,
                data={
                    "cell": cell,
                    "value": nd.get("value"),
                    "formula": nd.get("formula"),
                    "named_range": nd.get("named_range"),
                    "cluster": nd.get("cluster_name"),
                    "pagerank": nd.get("pagerank", 0.0),
                    "depends_on": depends_on,
                },
            )
        except Exception as e:
            logger.error(f"explain_formula failed: {e}")
            return ToolResult(tool_name="explain_formula", success=False, data=None, error=str(e))

    def generate_change_log(self, changes: List[Dict]) -> ToolResult:
        try:
            date_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
            lines = [
                f"## Change Log — {date_str}\n",
                "| Cell | Old Value | New Value | Reason |",
                "|------|-----------|-----------|--------|",
            ]
            for change in changes:
                cell = change.get("cell", "")
                old_val = change.get("old_value", "")
                new_val = change.get("new_value", "")
                reason = change.get("reason", "")
                lines.append(f"| {cell} | {old_val} | {new_val} | {reason} |")
            markdown = "\n".join(lines)
            return ToolResult(tool_name="generate_change_log", success=True, data={"markdown": markdown})
        except Exception as e:
            logger.error(f"generate_change_log failed: {e}")
            return ToolResult(tool_name="generate_change_log", success=False, data=None, error=str(e))

    def add_column(self, sheet: str, after_col: str, header: str, values: Optional[List[Any]] = None) -> ToolResult:
        """Insert a new column after `after_col` with a header and optional values."""
        try:
            from openpyxl.utils import column_index_from_string, get_column_letter
            from parser.xlsx_parser import CellData, _infer_data_type

            after_idx = column_index_from_string(after_col.upper())
            new_col_idx = after_idx + 1
            new_col_letter = get_column_letter(new_col_idx)

            # Find max row in this sheet
            sheet_cells = [c for c in self.workbook_data.cells.values() if c.sheet_name == sheet]
            max_row = max((c.row for c in sheet_cells), default=1) if sheet_cells else 1

            diff: List[Dict[str, Any]] = []

            # Write header at row 1
            header_addr = f"{sheet}!{new_col_letter}1"
            self.workbook_data.cells[header_addr] = CellData(
                cell_address=header_addr, sheet_name=sheet, value=header,
                formula=None, data_type="text", named_range=None,
                row=1, col=new_col_idx, is_hardcoded=True, is_merged=False, merge_master=None,
            )
            diff.append({"cell": header_addr, "before": None, "after": header})

            # Write values starting at row 2
            if values:
                for i, val in enumerate(values):
                    row_num = i + 2
                    addr = f"{sheet}!{new_col_letter}{row_num}"
                    self.workbook_data.cells[addr] = CellData(
                        cell_address=addr, sheet_name=sheet, value=val,
                        formula=None, data_type=_infer_data_type(val, None),
                        named_range=None, row=row_num, col=new_col_idx,
                        is_hardcoded=True, is_merged=False, merge_master=None,
                    )
                    diff.append({"cell": addr, "before": None, "after": val})

            result = ToolResult(tool_name="add_column", success=True,
                data={"column": new_col_letter, "header": header, "rows_written": len(values or []) + 1, "diff": diff})
            self._log_action("add_column", {"sheet": sheet, "after_col": after_col, "header": header}, result)
            return result
        except Exception as e:
            logger.error(f"add_column failed: {e}")
            return ToolResult(tool_name="add_column", success=False, data=None, error=str(e))

    def add_row(self, sheet: str, row_num: int, values: Dict[str, Any]) -> ToolResult:
        """Insert data into a specific row. values = {"A": val1, "B": val2, ...}"""
        try:
            from parser.xlsx_parser import CellData, _infer_data_type
            from openpyxl.utils import column_index_from_string

            diff: List[Dict[str, Any]] = []

            for col_letter, val in values.items():
                col_idx = column_index_from_string(col_letter.upper())
                addr = f"{sheet}!{col_letter.upper()}{row_num}"
                before = self.workbook_data.cells[addr].value if addr in self.workbook_data.cells else None

                self.workbook_data.cells[addr] = CellData(
                    cell_address=addr, sheet_name=sheet, value=val,
                    formula=None, data_type=_infer_data_type(val, None),
                    named_range=None, row=row_num, col=col_idx,
                    is_hardcoded=True, is_merged=False, merge_master=None,
                )
                self.workbook_state[addr] = {"value": val, "formula": None, "data_type": _infer_data_type(val, None)}
                diff.append({"cell": addr, "before": before, "after": val})

            result = ToolResult(tool_name="add_row", success=True,
                data={"sheet": sheet, "row": row_num, "cells_written": len(values), "diff": diff})
            self._log_action("add_row", {"sheet": sheet, "row_num": row_num, "values": values}, result)
            return result
        except Exception as e:
            logger.error(f"add_row failed: {e}")
            return ToolResult(tool_name="add_row", success=False, data=None, error=str(e))

    def create_sheet(self, name: str) -> ToolResult:
        """Create a new empty sheet in the workbook."""
        try:
            if name in self.workbook_data.sheets:
                return ToolResult(tool_name="create_sheet", success=False, data=None,
                    error=f"Sheet '{name}' already exists")
            self.workbook_data.sheets.append(name)
            result = ToolResult(tool_name="create_sheet", success=True,
                data={"sheet": name, "total_sheets": len(self.workbook_data.sheets), "diff": []})
            self._log_action("create_sheet", {"name": name}, result)
            return result
        except Exception as e:
            logger.error(f"create_sheet failed: {e}")
            return ToolResult(tool_name="create_sheet", success=False, data=None, error=str(e))

    def delete_range(self, sheet: str, range_str: str, confirmed: bool = False) -> ToolResult:
        """Delete cells in a range. If confirmed=False, returns PREVIEW only (no deletion).
        Frontend must show preview to user, then call again with confirmed=True."""
        try:
            cells_to_delete = self._resolve_range(sheet, range_str)
            preview: List[Dict[str, Any]] = []

            for addr in cells_to_delete:
                if addr in self.workbook_data.cells:
                    cd = self.workbook_data.cells[addr]
                    preview.append({
                        "cell": addr,
                        "value": cd.value,
                        "formula": cd.formula,
                        "data_type": cd.data_type,
                    })

            if not confirmed:
                return ToolResult(tool_name="delete_range", success=True,
                    data={
                        "preview": preview,
                        "count": len(preview),
                        "requires_confirmation": True,
                        "message": f"⚠️ This will DELETE {len(preview)} cells in {sheet}!{range_str}. Review the data below and confirm.",
                    })

            # Actually delete
            diff: List[Dict[str, Any]] = []
            for addr in cells_to_delete:
                if addr in self.workbook_data.cells:
                    before = self.workbook_data.cells[addr].value
                    del self.workbook_data.cells[addr]
                    diff.append({"cell": addr, "before": before, "after": None})
                if addr in self.workbook_state:
                    del self.workbook_state[addr]

            result = ToolResult(tool_name="delete_range", success=True,
                data={"deleted": len(diff), "diff": diff})
            self._log_action("delete_range", {"sheet": sheet, "range": range_str}, result)
            return result
        except Exception as e:
            logger.error(f"delete_range failed: {e}")
            return ToolResult(tool_name="delete_range", success=False, data=None, error=str(e))

    def write_formula(self, sheet: str, cell: str, formula: str) -> ToolResult:
        """Write a formula to a cell. Formula should NOT include the leading '='."""
        try:
            from openpyxl.utils import column_index_from_string
            from parser.xlsx_parser import CellData

            addr = f"{sheet}!{cell.upper()}"
            before = self.workbook_data.cells[addr].value if addr in self.workbook_data.cells else None
            before_formula = self.workbook_data.cells[addr].formula if addr in self.workbook_data.cells else None

            col_part = "".join(c for c in cell if c.isalpha())
            row_part = "".join(c for c in cell if c.isdigit())
            col_idx = column_index_from_string(col_part.upper()) if col_part else 1
            row_idx = int(row_part) if row_part else 1

            clean_formula = formula.lstrip("=")

            self.workbook_data.cells[addr] = CellData(
                cell_address=addr, sheet_name=sheet, value=f"={clean_formula}",
                formula=clean_formula, data_type="formula",
                named_range=None, row=row_idx, col=col_idx,
                is_hardcoded=False, is_merged=False, merge_master=None,
            )
            self.workbook_state[addr] = {"value": f"={clean_formula}", "formula": clean_formula, "data_type": "formula"}

            diff = [{"cell": addr, "before": before_formula or before, "after": f"={clean_formula}"}]

            result = ToolResult(tool_name="write_formula", success=True,
                data={"cell": addr, "formula": f"={clean_formula}", "diff": diff})
            self._log_action("write_formula", {"sheet": sheet, "cell": cell, "formula": formula}, result)
            return result
        except Exception as e:
            logger.error(f"write_formula failed: {e}")
            return ToolResult(tool_name="write_formula", success=False, data=None, error=str(e))

    def pivot_table(
        self,
        source_sheet: str,
        target_sheet: str,
        row_field: str,
        col_field: Optional[str] = None,
        value_field: str = "",
        agg: str = "sum",
    ) -> ToolResult:
        """Generate a pivot-table-like summary sheet from source data.
        
        row_field / col_field / value_field are column HEADERS (first row labels).
        agg is one of: sum, avg, count, min, max.
        """
        try:
            from collections import defaultdict
            from parser.xlsx_parser import CellData

            # --- locate header row in source sheet ---
            header_cells = {}
            for c in self.workbook_data.cells.values():
                if c.sheet_name == source_sheet and c.row == 1 and c.value is not None:
                    header_cells[str(c.value).strip().lower()] = c.col

            if not header_cells:
                return ToolResult(tool_name="pivot_table", success=False, data=None,
                    error=f"No headers found in row 1 of sheet '{source_sheet}'")

            row_col = header_cells.get(row_field.strip().lower())
            val_col = header_cells.get(value_field.strip().lower()) if value_field else None
            cross_col = header_cells.get(col_field.strip().lower()) if col_field else None

            if row_col is None:
                return ToolResult(tool_name="pivot_table", success=False, data=None,
                    error=f"Row field '{row_field}' not found in headers: {list(header_cells.keys())}")

            # --- read source data ---
            max_row = max((c.row for c in self.workbook_data.cells.values() if c.sheet_name == source_sheet), default=1)
            cell_lookup = {}
            for c in self.workbook_data.cells.values():
                if c.sheet_name == source_sheet:
                    cell_lookup[(c.row, c.col)] = c

            if cross_col is not None:
                # 2D pivot: row_field x col_field → value
                buckets: Dict[str, Dict[str, list]] = defaultdict(lambda: defaultdict(list))
                cross_labels: set = set()
                for r in range(2, max_row + 1):
                    row_label_cell = cell_lookup.get((r, row_col))
                    cross_label_cell = cell_lookup.get((r, cross_col))
                    val_cell = cell_lookup.get((r, val_col)) if val_col else None
                    if row_label_cell is None or row_label_cell.value is None:
                        continue
                    rl = str(row_label_cell.value)
                    cl = str(cross_label_cell.value) if cross_label_cell and cross_label_cell.value is not None else "Other"
                    cross_labels.add(cl)
                    v = val_cell.value if val_cell and isinstance(val_cell.value, (int, float)) else (1 if agg == "count" else 0)
                    buckets[rl][cl].append(v)

                sorted_cross = sorted(cross_labels)
                sorted_rows = sorted(buckets.keys())
            else:
                # 1D pivot: row_field → aggregation of value
                buckets_1d: Dict[str, list] = defaultdict(list)
                for r in range(2, max_row + 1):
                    row_label_cell = cell_lookup.get((r, row_col))
                    val_cell = cell_lookup.get((r, val_col)) if val_col else None
                    if row_label_cell is None or row_label_cell.value is None:
                        continue
                    rl = str(row_label_cell.value)
                    v = val_cell.value if val_cell and isinstance(val_cell.value, (int, float)) else (1 if agg == "count" else 0)
                    buckets_1d[rl].append(v)
                sorted_rows = sorted(buckets_1d.keys())
                sorted_cross = []

            # --- aggregate ---
            def _agg(vals: list) -> float:
                if not vals:
                    return 0.0
                if agg == "sum":
                    return sum(vals)
                elif agg == "avg":
                    return sum(vals) / len(vals)
                elif agg == "count":
                    return float(len(vals))
                elif agg == "min":
                    return min(vals)
                elif agg == "max":
                    return max(vals)
                return sum(vals)

            # --- create target sheet ---
            create_result = self.create_sheet(target_sheet)
            if not create_result.success and "already exists" not in (create_result.error or ""):
                return create_result

            from openpyxl.utils import get_column_letter
            diff: List[Dict[str, Any]] = []

            # Write headers
            header_row_label = row_field
            addr = f"{target_sheet}!A1"
            self.workbook_data.cells[addr] = CellData(
                cell_address=addr, sheet_name=target_sheet, value=header_row_label,
                formula=None, data_type="text", named_range=None, row=1, col=1,
                is_hardcoded=True, is_merged=False, merge_master=None,
            )
            self.workbook_state[addr] = {"value": header_row_label}
            diff.append({"cell": addr, "before": None, "after": header_row_label})

            if sorted_cross:
                for ci, cl in enumerate(sorted_cross):
                    col_letter = get_column_letter(ci + 2)
                    addr = f"{target_sheet}!{col_letter}1"
                    self.workbook_data.cells[addr] = CellData(
                        cell_address=addr, sheet_name=target_sheet, value=cl,
                        formula=None, data_type="text", named_range=None, row=1, col=ci + 2,
                        is_hardcoded=True, is_merged=False, merge_master=None,
                    )
                    self.workbook_state[addr] = {"value": cl}
                    diff.append({"cell": addr, "before": None, "after": cl})
            else:
                val_header = f"{agg.upper()}({value_field})" if value_field else "Count"
                addr = f"{target_sheet}!B1"
                self.workbook_data.cells[addr] = CellData(
                    cell_address=addr, sheet_name=target_sheet, value=val_header,
                    formula=None, data_type="text", named_range=None, row=1, col=2,
                    is_hardcoded=True, is_merged=False, merge_master=None,
                )
                self.workbook_state[addr] = {"value": val_header}
                diff.append({"cell": addr, "before": None, "after": val_header})

            # Write data rows
            for ri, rl in enumerate(sorted_rows):
                r = ri + 2
                addr = f"{target_sheet}!A{r}"
                self.workbook_data.cells[addr] = CellData(
                    cell_address=addr, sheet_name=target_sheet, value=rl,
                    formula=None, data_type="text", named_range=None, row=r, col=1,
                    is_hardcoded=True, is_merged=False, merge_master=None,
                )
                self.workbook_state[addr] = {"value": rl}
                diff.append({"cell": addr, "before": None, "after": rl})

                if sorted_cross:
                    for ci, cl in enumerate(sorted_cross):
                        val = round(_agg(buckets[rl].get(cl, [])), 4)
                        col_letter = get_column_letter(ci + 2)
                        addr = f"{target_sheet}!{col_letter}{r}"
                        self.workbook_data.cells[addr] = CellData(
                            cell_address=addr, sheet_name=target_sheet, value=val,
                            formula=None, data_type="number", named_range=None, row=r, col=ci + 2,
                            is_hardcoded=True, is_merged=False, merge_master=None,
                        )
                        self.workbook_state[addr] = {"value": val}
                        diff.append({"cell": addr, "before": None, "after": val})
                else:
                    val = round(_agg(buckets_1d[rl]), 4)
                    addr = f"{target_sheet}!B{r}"
                    self.workbook_data.cells[addr] = CellData(
                        cell_address=addr, sheet_name=target_sheet, value=val,
                        formula=None, data_type="number", named_range=None, row=r, col=2,
                        is_hardcoded=True, is_merged=False, merge_master=None,
                    )
                    self.workbook_state[addr] = {"value": val}
                    diff.append({"cell": addr, "before": None, "after": val})

            result = ToolResult(tool_name="pivot_table", success=True,
                data={
                    "target_sheet": target_sheet,
                    "rows": len(sorted_rows),
                    "columns": len(sorted_cross) if sorted_cross else 1,
                    "aggregation": agg,
                    "diff": diff,
                })
            self._log_action("pivot_table", {
                "source_sheet": source_sheet, "target_sheet": target_sheet,
                "row_field": row_field, "col_field": col_field,
                "value_field": value_field, "agg": agg,
            }, result)
            return result
        except Exception as e:
            logger.error(f"pivot_table failed: {e}")
            return ToolResult(tool_name="pivot_table", success=False, data=None, error=str(e))

    def validate_formulas(self) -> ToolResult:
        """Run formula consistency checks on the entire workbook."""
        try:
            from analysis.formula_validator import FormulaValidator
            validator = FormulaValidator(self.workbook_data)
            report = validator.run()
            result = ToolResult(tool_name="validate_formulas", success=True,
                data={
                    "consistency_score": report.consistency_score,
                    "formula_count": report.formula_count,
                    "total_issues": report.total_issues,
                    "errors": report.errors,
                    "warnings": report.warnings,
                    "info": report.info,
                    "issues": [
                        {
                            "severity": i.severity,
                            "category": i.category,
                            "cell": i.cell,
                            "sheet": i.sheet,
                            "message": i.message,
                            "expected": i.expected,
                            "actual": i.actual,
                        }
                        for i in report.issues[:30]
                    ],
                })
            self._log_action("validate_formulas", {}, result)
            return result
        except Exception as e:
            logger.error(f"validate_formulas failed: {e}")
            return ToolResult(tool_name="validate_formulas", success=False, data=None, error=str(e))

    def get_audit_trail(self) -> ToolResult:
        """Return the full audit trail of all AI actions in this session."""
        return ToolResult(tool_name="get_audit_trail", success=True,
            data={"actions": self.audit_trail, "count": len(self.audit_trail)})

    def _resolve_range(self, sheet: str, range_str: str) -> List[str]:
        import re
        RANGE_PAT = re.compile(r"^([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)$", re.IGNORECASE)
        CELL_PAT = re.compile(r"^([A-Z]+)([0-9]+)$", re.IGNORECASE)

        range_str = range_str.strip()
        m = RANGE_PAT.match(range_str)
        if m:
            col_start = self._col_to_num(m.group(1))
            row_start = int(m.group(2))
            col_end = self._col_to_num(m.group(3))
            row_end = int(m.group(4))
            result = []
            for r in range(row_start, row_end + 1):
                for c in range(col_start, col_end + 1):
                    result.append(f"{sheet}!{self._num_to_col(c)}{r}")
            return result

        m2 = CELL_PAT.match(range_str)
        if m2:
            return [f"{sheet}!{range_str.upper()}"]

        return [f"{sheet}!{range_str}"]

    @staticmethod
    def _col_to_num(col: str) -> int:
        col = col.upper()
        result = 0
        for ch in col:
            result = result * 26 + (ord(ch) - ord("A") + 1)
        return result

    @staticmethod
    def _num_to_col(n: int) -> str:
        result = ""
        while n > 0:
            n, rem = divmod(n - 1, 26)
            result = chr(ord("A") + rem) + result
        return result
