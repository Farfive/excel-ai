import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)


@dataclass
class CellStyle:
    bold: bool = False
    italic: bool = False
    underline: bool = False
    font_size: float = 11.0
    font_name: str = "Calibri"
    font_color: str = "#000000"
    bg_color: str = ""
    h_align: str = ""
    v_align: str = ""
    wrap_text: bool = False
    indent: int = 0
    number_format: str = "General"
    border_top: str = ""
    border_bottom: str = ""
    border_left: str = ""
    border_right: str = ""
    border_color: str = "#000000"

    def to_dict(self) -> Dict[str, Any]:
        d: Dict[str, Any] = {}
        if self.bold: d["b"] = True
        if self.italic: d["i"] = True
        if self.underline: d["u"] = True
        d["fs"] = self.font_size
        d["fn"] = self.font_name
        if self.font_color != "#000000": d["fc"] = self.font_color
        if self.bg_color: d["bg"] = self.bg_color
        if self.h_align: d["ha"] = self.h_align
        if self.v_align: d["va"] = self.v_align
        if self.wrap_text: d["wt"] = True
        if self.indent: d["ind"] = self.indent
        if self.number_format != "General": d["nf"] = self.number_format
        if self.border_top: d["bt"] = self.border_top
        if self.border_bottom: d["bb"] = self.border_bottom
        if self.border_left: d["bl"] = self.border_left
        if self.border_right: d["br"] = self.border_right
        if self.border_color != "#000000": d["bc"] = self.border_color
        return d


def _extract_style(cell) -> CellStyle:
    s = CellStyle()
    try:
        font = cell.font
        if font:
            s.bold = bool(font.bold)
            s.italic = bool(font.italic)
            s.underline = font.underline not in (None, "none", False)
            if font.size: s.font_size = float(font.size)
            if font.name: s.font_name = font.name
            if font.color and font.color.rgb and isinstance(font.color.rgb, str) and len(font.color.rgb) >= 6:
                rgb = font.color.rgb
                if len(rgb) == 8: rgb = rgb[2:]
                s.font_color = f"#{rgb}"
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb and isinstance(fill.fgColor.rgb, str):
            rgb = fill.fgColor.rgb
            if rgb != "00000000" and len(rgb) >= 6:
                if len(rgb) == 8: rgb = rgb[2:]
                s.bg_color = f"#{rgb}"
        align = cell.alignment
        if align:
            if align.horizontal: s.h_align = align.horizontal
            if align.vertical: s.v_align = align.vertical
            if align.wrap_text: s.wrap_text = True
            if align.indent: s.indent = int(align.indent)
        if cell.number_format and cell.number_format != "General":
            s.number_format = cell.number_format
        border = cell.border
        if border:
            if border.top and border.top.style: s.border_top = border.top.style
            if border.bottom and border.bottom.style: s.border_bottom = border.bottom.style
            if border.left and border.left.style: s.border_left = border.left.style
            if border.right and border.right.style: s.border_right = border.right.style
    except Exception:
        pass
    return s


@dataclass
class CellData:
    cell_address: str
    sheet_name: str
    value: Any
    formula: Optional[str]
    data_type: str
    named_range: Optional[str]
    row: int
    col: int
    is_hardcoded: bool
    is_merged: bool
    merge_master: Optional[str]
    is_hidden: bool = False
    style: Optional[CellStyle] = None


@dataclass
class WorkbookData:
    cells: Dict[str, CellData] = field(default_factory=dict)
    sheets: List[str] = field(default_factory=list)
    named_ranges: Dict[str, str] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)


def _infer_data_type(value: Any, formula: Optional[str]) -> str:
    if formula:
        return "formula"
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "text"
    return "date"


class XLSXParser:
    def parse(self, file_path: str) -> WorkbookData:
        workbook_data = WorkbookData()
        workbook_data.metadata["has_circular_refs"] = False

        try:
            wb_formula = openpyxl.load_workbook(file_path, data_only=False)
        except Exception as e:
            logger.error(f"Failed to open workbook for formulas: {e}")
            raise

        try:
            wb_values = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open workbook for values: {e}")
            raise

        named_range_map: Dict[str, str] = {}
        try:
            for name, defn in wb_formula.defined_names.items():
                for title, coord in defn.destinations:
                    workbook_data.named_ranges[name] = f"{title}!{coord}"
                    if title and coord:
                        named_range_map[f"{title}!{coord}"] = name
        except Exception as e:
            logger.warning(f"Error parsing named ranges: {e}")

        workbook_data.sheets = wb_formula.sheetnames

        for sheet_name in wb_formula.sheetnames:
            ws_formula = wb_formula[sheet_name]
            ws_values = wb_values[sheet_name] if sheet_name in wb_values.sheetnames else None
            is_hidden = ws_formula.sheet_state != "visible"

            merged_cells_map: Dict[str, str] = {}
            for merged_range in ws_formula.merged_cells.ranges:
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col,
                )
                master_col_letter = get_column_letter(min_col)
                master_addr = f"{sheet_name}!{master_col_letter}{min_row}"
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        col_letter = get_column_letter(c)
                        cell_addr = f"{sheet_name}!{col_letter}{r}"
                        if cell_addr != master_addr:
                            merged_cells_map[cell_addr] = master_addr

            for row in ws_formula.iter_rows():
                for cell in row:
                    col_letter = get_column_letter(cell.column)
                    cell_address = f"{sheet_name}!{col_letter}{cell.row}"

                    if cell_address in merged_cells_map:
                        cell_data = CellData(
                            cell_address=cell_address,
                            sheet_name=sheet_name,
                            value=None,
                            formula=None,
                            data_type="empty",
                            named_range=None,
                            row=cell.row,
                            col=cell.column,
                            is_hardcoded=False,
                            is_merged=True,
                            merge_master=merged_cells_map[cell_address],
                            is_hidden=is_hidden,
                            style=_extract_style(cell),
                        )
                        workbook_data.cells[cell_address] = cell_data
                        continue

                    formula_val = cell.value
                    formula_str: Optional[str] = None

                    if isinstance(formula_val, str) and formula_val.startswith("="):
                        formula_str = formula_val
                        value = None
                        if ws_values and cell.row <= ws_values.max_row and cell.column <= ws_values.max_column:
                            val_cell = ws_values.cell(row=cell.row, column=cell.column)
                            value = val_cell.value
                    else:
                        value = formula_val

                    if value is None and formula_str is None:
                        continue

                    data_type = _infer_data_type(value, formula_str)
                    is_hardcoded = isinstance(value, (int, float)) and formula_str is None

                    nr_key = cell_address
                    named_range = named_range_map.get(nr_key)

                    is_merged_master = any(
                        str(mr).startswith(f"{col_letter}{cell.row}")
                        for mr in ws_formula.merged_cells.ranges
                    )

                    cell_data = CellData(
                        cell_address=cell_address,
                        sheet_name=sheet_name,
                        value=value,
                        formula=formula_str,
                        data_type=data_type,
                        named_range=named_range,
                        row=cell.row,
                        col=cell.column,
                        is_hardcoded=is_hardcoded,
                        is_merged=is_merged_master,
                        merge_master=None,
                        is_hidden=is_hidden,
                        style=_extract_style(cell),
                    )
                    workbook_data.cells[cell_address] = cell_data

        workbook_data.metadata["sheet_count"] = len(workbook_data.sheets)
        workbook_data.metadata["cell_count"] = len(workbook_data.cells)
        workbook_data.metadata["named_range_count"] = len(workbook_data.named_ranges)

        logger.info(
            f"Parsed workbook: {workbook_data.metadata['cell_count']} cells, "
            f"{len(workbook_data.sheets)} sheets"
        )
        return workbook_data
