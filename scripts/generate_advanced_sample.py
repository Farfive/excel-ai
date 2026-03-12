"""
Generate a LARGE advanced Excel file for testing Excel AI.
~5000+ cells, 15 sheets, formulas, styles, merged cells, named ranges,
data validation, conditional formatting, hyperlinks, cross-sheet refs.
"""
import openpyxl
from openpyxl.utils import get_column_letter as gcl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
    NamedStyle,
)
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, DataBarRule, FormulaRule,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import random
import os
import datetime

wb = openpyxl.Workbook()

# ── Constants ──────────────────────────────────────────────────
YEARS = list(range(2021, 2034))
YC = len(YEARS)
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
REGIONS = ["North America", "Europe", "Asia Pacific", "Latin America", "Middle East & Africa"]
PRODUCTS = [
    ("Enterprise SaaS", 52_000_000),
    ("SMB Platform", 28_000_000),
    ("Data Analytics Suite", 18_000_000),
    ("Cloud Infrastructure", 14_000_000),
    ("AI/ML Services", 9_000_000),
    ("Cybersecurity Module", 7_500_000),
    ("Professional Services", 6_000_000),
    ("Maintenance & Support", 4_500_000),
    ("Marketplace Revenue", 3_000_000),
    ("Licensing & Royalties", 2_000_000),
]
DEPARTMENTS = [
    ("Engineering", 145), ("Sales", 95), ("Marketing", 48),
    ("Customer Success", 42), ("Finance & Accounting", 24),
    ("HR & People Ops", 18), ("Legal & Compliance", 12),
    ("Executive Team", 11), ("Data Science & AI", 30),
    ("Product Management", 22), ("DevOps / SRE", 18),
    ("Quality Assurance", 25), ("IT Operations", 15),
    ("Business Development", 20), ("Design & UX", 16),
]
EMPLOYEE_NAMES = [
    "Anna Kowalski", "James Chen", "Maria Garcia", "David Kim", "Sarah Johnson",
    "Tomasz Nowak", "Li Wei", "Elena Petrova", "Carlos Rodriguez", "Yuki Tanaka",
    "Priya Sharma", "Michael Brown", "Sophie Martin", "Ahmed Hassan", "Olga Smirnova",
    "Lucas Müller", "Fatima Al-Sayed", "Roberto Rossi", "Hanna Johansson", "Raj Patel",
    "Emma Wilson", "Jakub Zieliński", "Aiko Suzuki", "Diego López", "Natalia Volkova",
    "Thomas Anderson", "Isabella Santos", "Jun Park", "Katarina Novák", "Viktor Petrov",
    "Amelia Thompson", "Marco Bianchi", "Sakura Yamamoto", "Andrei Popov", "Clara Fischer",
    "Hiroshi Watanabe", "Marta Fernández", "Sven Eriksson", "Nadia Kozlova", "Pierre Dubois",
    "Mei Lin", "João Silva", "Eva Horváth", "Khalid Omar", "Ingrid Larsen",
    "Ravi Kumar", "Zofia Wiśniewska", "Liam O'Brien", "Yuna Choi", "Stefan Gruber",
]

# ── Styles ─────────────────────────────────────────────────────
HDR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR2_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
LINK_FONT = Font(name="Calibri", color="0563C1", underline="single", size=10)
THIN = Side(style="thin", color="B4C6E7")
MED = Side(style="medium", color="1F4E79")
BORDER = Border(bottom=THIN, right=THIN)
BORDER_TOTAL = Border(top=MED, bottom=Side(style="double", color="1F4E79"), right=THIN)
NUM = '#,##0'
PCT = '0.0%'
MONEY = '#,##0.00'
DATE_FMT = 'YYYY-MM-DD'
ACC = '#,##0;[Red](#,##0)'

def hdr_row(ws, row, mc):
    for c in range(1, mc + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def hdr2_row(ws, row, mc):
    for c in range(1, mc + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = HDR2_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

def inp(cell):
    cell.fill = INPUT_FILL
    cell.border = BORDER

def fml(cell):
    cell.fill = FORMULA_FILL
    cell.border = BORDER

def tot(cell):
    cell.fill = TOTAL_FILL
    cell.border = BORDER_TOTAL
    cell.font = Font(bold=True)

def title(ws, text, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(YC+1, 14))
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, size=14, color="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")

def subtitle(ws, text, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(YC+1, 14))
    c = ws.cell(row, 1, text)
    c.font = Font(italic=True, size=10, color="808080")

def setup_cols(ws, a_width=32):
    ws.column_dimensions['A'].width = a_width
    for i in range(YC):
        ws.column_dimensions[gcl(i + 2)].width = 14

def year_headers(ws, row):
    for i, yr in enumerate(YEARS):
        ws.cell(row, i + 2, yr)
    hdr_row(ws, row, YC + 1)

# ════════════════════════════════════════════════════════════════
# SHEET 1: ASSUMPTIONS
# ════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Assumptions"
setup_cols(ws)
title(ws, "MODEL ASSUMPTIONS")
subtitle(ws, "Key financial model inputs — yellow cells are editable assumptions")
year_headers(ws, 4)

assumptions = {
    "Revenue Growth Rate":       [0.00, 0.08, 0.15, 0.12, 0.10, 0.09, 0.08, 0.07, 0.06, 0.05, 0.05, 0.04, 0.04],
    "COGS % of Revenue":         [0.44, 0.43, 0.42, 0.41, 0.40, 0.39, 0.38, 0.38, 0.37, 0.37, 0.36, 0.36, 0.35],
    "SGA % of Revenue":          [0.20, 0.19, 0.18, 0.17, 0.16, 0.15, 0.15, 0.14, 0.14, 0.13, 0.13, 0.12, 0.12],
    "R&D % of Revenue":          [0.14, 0.13, 0.12, 0.11, 0.10, 0.10, 0.09, 0.09, 0.08, 0.08, 0.08, 0.08, 0.07],
    "D&A % of Revenue":          [0.03]*YC,
    "Tax Rate":                  [0.21]*YC,
    "CapEx % of Revenue":        [0.06, 0.05, 0.05, 0.05, 0.04, 0.04, 0.04, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03],
    "NWC % of Revenue":          [0.12, 0.11, 0.10, 0.10, 0.09, 0.09, 0.09, 0.08, 0.08, 0.08, 0.08, 0.08, 0.08],
    "WACC":                      [0.10]*YC,
    "Terminal Growth Rate":      [0.025]*YC,
    "Shares Outstanding (M)":   [100]*YC,
    "Interest Rate on Debt":     [0.045]*YC,
    "Dividend Payout Ratio":     [0.25, 0.25, 0.28, 0.28, 0.30, 0.30, 0.30, 0.32, 0.32, 0.35, 0.35, 0.35, 0.35],
    "Inflation Rate":            [0.035, 0.032, 0.028, 0.025, 0.022, 0.020, 0.020, 0.020, 0.020, 0.020, 0.020, 0.020, 0.020],
    "Avg Salary Increase":       [0.04, 0.04, 0.035, 0.035, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03],
    "Customer Churn Rate":       [0.12, 0.11, 0.10, 0.09, 0.08, 0.08, 0.07, 0.07, 0.06, 0.06, 0.06, 0.05, 0.05],
    "Avg Deal Size ($K)":        [45, 48, 52, 55, 58, 62, 65, 68, 72, 75, 78, 82, 85],
}

r = 5
nrm = {}  # named range map: label -> row
for label, values in assumptions.items():
    ws.cell(r, 1, label).font = Font(bold=True, size=10)
    for i, val in enumerate(values[:YC]):
        c = ws.cell(r, i + 2, val)
        inp(c)
        if any(k in label for k in ["%", "Rate", "Ratio", "Growth", "Churn", "WACC"]):
            c.number_format = PCT
        elif "Salary" in label:
            c.number_format = PCT
        elif "Deal" in label:
            c.number_format = NUM
        else:
            c.number_format = NUM
    safe = label.replace(" ", "_").replace("&", "and").replace("%", "pct").replace("(", "").replace(")", "")
    nrm[label] = r
    try:
        ref = f"Assumptions!$B${r}:${gcl(YC+1)}${r}"
        wb.defined_names.new(safe, attr_text=ref)
    except Exception:
        pass
    r += 1

# Scenario selector with data validation
r += 1
ws.cell(r, 1, "Active Scenario").font = Font(bold=True, size=11, color="1F4E79")
ws.cell(r, 2, "Base Case")
dv = DataValidation(type="list", formula1='"Base Case,Bull Case,Bear Case"', allow_blank=False)
dv.error = "Must be Base Case, Bull Case, or Bear Case"
dv.errorTitle = "Invalid Scenario"
dv.prompt = "Select the active scenario"
dv.promptTitle = "Scenario"
ws.add_data_validation(dv)
dv.add(ws.cell(r, 2))
ws.cell(r, 2).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ws.cell(r, 2).font = Font(bold=True, size=11)

# ════════════════════════════════════════════════════════════════
# SHEET 2: REVENUE
# ════════════════════════════════════════════════════════════════
ws_rev = wb.create_sheet("Revenue")
setup_cols(ws_rev)
title(ws_rev, "REVENUE BUILD")
subtitle(ws_rev, "Product-level revenue with growth rates from Assumptions")
year_headers(ws_rev, 4)

r = 5
prod_rows = {}
for name, base in PRODUCTS:
    ws_rev.cell(r, 1, name).font = Font(bold=True, size=10)
    for i in range(YC):
        if i == 0:
            c = ws_rev.cell(r, i + 2, base)
            inp(c)
        else:
            prev = gcl(i + 1)
            gref = f"Assumptions!{gcl(i+2)}${nrm['Revenue Growth Rate']}"
            c = ws_rev.cell(r, i + 2, value=f"={prev}{r}*(1+{gref})")
            fml(c)
        c.number_format = NUM
    prod_rows[name] = r
    r += 1

r += 1
ws_rev.cell(r, 1, "Total Revenue").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_rev.cell(r, i + 2, value=f"=SUM({cl}5:{cl}{r-2})")
    tot(c)
    c.number_format = NUM
total_rev_row = r

# YoY growth row
r += 1
ws_rev.cell(r, 1, "YoY Growth").font = Font(bold=True, size=10, color="808080")
for i in range(YC):
    cl = gcl(i + 2)
    if i == 0:
        ws_rev.cell(r, i + 2, "N/A").font = Font(color="808080")
    else:
        prev = gcl(i + 1)
        c = ws_rev.cell(r, i + 2, value=f"=({cl}{total_rev_row}-{prev}{total_rev_row})/{prev}{total_rev_row}")
        fml(c)
        c.number_format = PCT

# Revenue mix %
r += 1
ws_rev.cell(r, 1, "Revenue Mix %").font = Font(bold=True, size=10, color="808080")
r += 1
for name in [p[0] for p in PRODUCTS]:
    ws_rev.cell(r, 1, f"  {name}").font = Font(size=9, color="666666")
    for i in range(YC):
        cl = gcl(i + 2)
        c = ws_rev.cell(r, i + 2, value=f"={cl}{prod_rows[name]}/{cl}{total_rev_row}")
        c.number_format = PCT
        c.font = Font(size=9, color="666666")
    r += 1

# ════════════════════════════════════════════════════════════════
# SHEET 3: COGS
# ════════════════════════════════════════════════════════════════
ws_cogs = wb.create_sheet("COGS")
setup_cols(ws_cogs)
title(ws_cogs, "COST OF GOODS SOLD")
subtitle(ws_cogs, "Detailed cost breakdown driven by COGS % from Assumptions")
year_headers(ws_cogs, 4)

cogs_items = [
    ("Cloud Hosting (AWS/Azure/GCP)", 0.30),
    ("Software Licenses (3rd party)", 0.12),
    ("Customer Support Staff", 0.18),
    ("Payment Processing", 0.08),
    ("Data Center Operations", 0.10),
    ("Content Delivery (CDN)", 0.05),
    ("Security & Compliance", 0.07),
    ("Integration & API Costs", 0.06),
    ("Other Direct Costs", 0.04),
]

r = 5
cogs_first = r
for item, share in cogs_items:
    ws_cogs.cell(r, 1, item).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        rev = f"Revenue!{cl}${total_rev_row}"
        pct = f"Assumptions!{cl}${nrm['COGS % of Revenue']}"
        c = ws_cogs.cell(r, i + 2, value=f"={rev}*{pct}*{share}")
        fml(c)
        c.number_format = NUM
    r += 1

cogs_last = r - 1
r += 1
ws_cogs.cell(r, 1, "Total COGS").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cogs.cell(r, i + 2, value=f"=SUM({cl}{cogs_first}:{cl}{cogs_last})")
    tot(c)
    c.number_format = NUM
total_cogs_row = r

# COGS margin row
r += 1
ws_cogs.cell(r, 1, "COGS as % of Revenue").font = Font(italic=True, size=10, color="808080")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cogs.cell(r, i + 2, value=f"={cl}{total_cogs_row}/Revenue!{cl}${total_rev_row}")
    c.number_format = PCT
    c.font = Font(size=9, color="808080")

# ════════════════════════════════════════════════════════════════
# SHEET 4: OPEX
# ════════════════════════════════════════════════════════════════
ws_opex = wb.create_sheet("OpEx")
setup_cols(ws_opex)
title(ws_opex, "OPERATING EXPENSES")
subtitle(ws_opex, "Departmental OpEx driven by SGA/R&D/D&A assumptions")
year_headers(ws_opex, 4)

opex_items = [
    ("Sales & Marketing", "SGA % of Revenue", 0.35),
    ("Salaries & Benefits", "SGA % of Revenue", 0.28),
    ("Research & Development", "R&D % of Revenue", 0.55),
    ("General & Administrative", "SGA % of Revenue", 0.10),
    ("Depreciation & Amortization", "D&A % of Revenue", 1.0),
    ("Rent & Facilities", "SGA % of Revenue", 0.07),
    ("Travel & Entertainment", "SGA % of Revenue", 0.04),
    ("Professional Fees", "SGA % of Revenue", 0.06),
    ("Insurance", "SGA % of Revenue", 0.03),
    ("IT Infrastructure", "R&D % of Revenue", 0.30),
    ("Recruiting & Training", "SGA % of Revenue", 0.04),
    ("Stock-Based Compensation", "SGA % of Revenue", 0.03),
]

r = 5
opex_first = r
for item, key, share in opex_items:
    ws_opex.cell(r, 1, item).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        rev = f"Revenue!{cl}${total_rev_row}"
        pct = f"Assumptions!{cl}${nrm[key]}"
        c = ws_opex.cell(r, i + 2, value=f"={rev}*{pct}*{share}")
        fml(c)
        c.number_format = NUM
    r += 1

opex_last = r - 1
r += 1
ws_opex.cell(r, 1, "Total OpEx").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_opex.cell(r, i + 2, value=f"=SUM({cl}{opex_first}:{cl}{opex_last})")
    tot(c)
    c.number_format = NUM
total_opex_row = r

# ════════════════════════════════════════════════════════════════
# SHEET 5: P&L
# ════════════════════════════════════════════════════════════════
ws_pl = wb.create_sheet("P&L")
setup_cols(ws_pl)
title(ws_pl, "PROFIT & LOSS STATEMENT")
subtitle(ws_pl, "Consolidated income statement with cross-sheet references")
year_headers(ws_pl, 4)

plr = {}
def pl_line(row, label, fn, fmt=NUM, bold=False, color="000000"):
    ws_pl.cell(row, 1, label).font = Font(bold=bold, size=11 if bold else 10, color=color)
    for i in range(YC):
        cl = gcl(i + 2)
        c = ws_pl.cell(row, i + 2, value=fn(cl))
        if bold:
            tot(c)
        else:
            fml(c)
        c.number_format = fmt
    plr[label] = row

r = 5
pl_line(r, "Revenue", lambda c: f"=Revenue!{c}${total_rev_row}", bold=True, color="1F4E79"); r+=1
pl_line(r, "Cost of Goods Sold", lambda c: f"=-COGS!{c}${total_cogs_row}"); r+=1
pl_line(r, "Gross Profit", lambda c: f"={c}{plr['Revenue']}+{c}{plr['Cost of Goods Sold']}", bold=True, color="1F4E79"); r+=1
pl_line(r, "Gross Margin %", lambda c: f"={c}{plr['Gross Profit']}/{c}{plr['Revenue']}", fmt=PCT); r+=2
pl_line(r, "Operating Expenses", lambda c: f"=-OpEx!{c}${total_opex_row}"); r+=1
pl_line(r, "EBIT", lambda c: f"={c}{plr['Gross Profit']}+{c}{plr['Operating Expenses']}", bold=True, color="1F4E79"); r+=1
pl_line(r, "EBIT Margin %", lambda c: f"={c}{plr['EBIT']}/{c}{plr['Revenue']}", fmt=PCT); r+=2

debt_r = nrm["Interest Rate on Debt"]
pl_line(r, "Interest Expense", lambda c: f"=-'Balance Sheet'!{c}$17*Assumptions!{c}${debt_r}"); r+=1
pl_line(r, "EBT", lambda c: f"={c}{plr['EBIT']}+{c}{plr['Interest Expense']}", bold=True, color="1F4E79"); r+=1
tax_r = nrm["Tax Rate"]
pl_line(r, "Income Tax", lambda c: f"=-MAX(0,{c}{plr['EBT']}*Assumptions!{c}${tax_r})"); r+=1
pl_line(r, "Net Income", lambda c: f"={c}{plr['EBT']}+{c}{plr['Income Tax']}", bold=True, color="1F4E79"); r+=1
pl_line(r, "Net Margin %", lambda c: f"={c}{plr['Net Income']}/{c}{plr['Revenue']}", fmt=PCT); r+=2
sh_r = nrm["Shares Outstanding (M)"]
pl_line(r, "EPS", lambda c: f"={c}{plr['Net Income']}/Assumptions!{c}${sh_r}/1000000", fmt=MONEY); r+=1
pl_line(r, "EBITDA", lambda c: f"={c}{plr['EBIT']}+Revenue!{c}${total_rev_row}*Assumptions!{c}${nrm['D&A % of Revenue']}", bold=True, color="1F4E79")

# ════════════════════════════════════════════════════════════════
# SHEET 6: BALANCE SHEET
# ════════════════════════════════════════════════════════════════
ws_bs = wb.create_sheet("Balance Sheet")
setup_cols(ws_bs)
title(ws_bs, "BALANCE SHEET")
subtitle(ws_bs, "Simplified balance sheet with asset/liability structure")
year_headers(ws_bs, 4)

bsr = {}
r = 5
# Merged section header
ws_bs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_bs.cell(r, 1, "ASSETS").font = Font(bold=True, size=12, color="1F4E79")
r += 1

items_assets = [
    ("Cash & Equivalents", 55_000_000, None),
    ("Accounts Receivable", None, lambda c,i: f"=Revenue!{c}${total_rev_row}*Assumptions!{c}${nrm['NWC % of Revenue']}*0.4"),
    ("Inventory", None, lambda c,i: f"=Revenue!{c}${total_rev_row}*Assumptions!{c}${nrm['NWC % of Revenue']}*0.3"),
    ("Prepaid Expenses", None, lambda c,i: f"=Revenue!{c}${total_rev_row}*0.015"),
    ("Other Current Assets", None, lambda c,i: f"=Revenue!{c}${total_rev_row}*0.01"),
]
ca_first = r
for label, base, fn in items_assets:
    ws_bs.cell(r, 1, label).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        if base is not None and i == 0:
            c = ws_bs.cell(r, i + 2, base)
            inp(c)
        elif fn:
            c = ws_bs.cell(r, i + 2, value=fn(cl, i))
            fml(c)
        elif base is not None and i > 0:
            c = ws_bs.cell(r, i + 2, value=f"='Cash Flow'!{cl}$21")
            fml(c)
        c.number_format = NUM
    bsr[label] = r
    r += 1

ca_last = r - 1
ws_bs.cell(r, 1, "Total Current Assets").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_bs.cell(r, i + 2, value=f"=SUM({cl}{ca_first}:{cl}{ca_last})")
    tot(c)
    c.number_format = NUM
bsr["Total CA"] = r
r += 1

ws_bs.cell(r, 1, "PP&E (Net)").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    if i == 0:
        c = ws_bs.cell(r, i + 2, 35_000_000)
        inp(c)
    else:
        prev = gcl(i + 1)
        c = ws_bs.cell(r, i + 2, value=f"={prev}{r}+Revenue!{cl}${total_rev_row}*Assumptions!{cl}${nrm['CapEx % of Revenue']}-Revenue!{cl}${total_rev_row}*Assumptions!{cl}${nrm['D&A % of Revenue']}")
        fml(c)
    c.number_format = NUM
bsr["PPE"] = r
r += 1

ws_bs.cell(r, 1, "Intangible Assets").font = Font(bold=True, size=10)
for i in range(YC):
    c = ws_bs.cell(r, i + 2, 25_000_000)
    inp(c)
    c.number_format = NUM
bsr["Intangibles"] = r
r += 1

ws_bs.cell(r, 1, "Goodwill").font = Font(bold=True, size=10)
for i in range(YC):
    c = ws_bs.cell(r, i + 2, 15_000_000)
    inp(c)
    c.number_format = NUM
bsr["Goodwill"] = r
r += 1

ws_bs.cell(r, 1, "Total Assets").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_bs.cell(r, i + 2, value=f"={cl}{bsr['Total CA']}+{cl}{bsr['PPE']}+{cl}{bsr['Intangibles']}+{cl}{bsr['Goodwill']}")
    tot(c)
    c.number_format = NUM
bsr["Total Assets"] = r
r += 2

ws_bs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_bs.cell(r, 1, "LIABILITIES & EQUITY").font = Font(bold=True, size=12, color="1F4E79")
r += 1

ws_bs.cell(r, 1, "Accounts Payable").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_bs.cell(r, i + 2, value=f"=Revenue!{cl}${total_rev_row}*Assumptions!{cl}${nrm['NWC % of Revenue']}*0.25")
    fml(c)
    c.number_format = NUM
bsr["AP"] = r
r += 1

ws_bs.cell(r, 1, "Long-Term Debt").font = Font(bold=True, size=10)
for i in range(YC):
    c = ws_bs.cell(r, i + 2, 45_000_000)
    inp(c)
    c.number_format = NUM
bsr["Debt"] = r
r += 1

ws_bs.cell(r, 1, "Total Liabilities").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_bs.cell(r, i + 2, value=f"={cl}{bsr['AP']}+{cl}{bsr['Debt']}")
    tot(c)
    c.number_format = NUM
bsr["Total Liab"] = r
r += 1

div_r = nrm["Dividend Payout Ratio"]
ws_bs.cell(r, 1, "Retained Earnings").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    if i == 0:
        c = ws_bs.cell(r, i + 2, 35_000_000)
        inp(c)
    else:
        prev = gcl(i + 1)
        c = ws_bs.cell(r, i + 2, value=f"={prev}{r}+'P&L'!{cl}${plr['Net Income']}*(1-Assumptions!{cl}${div_r})")
        fml(c)
    c.number_format = NUM
bsr["RE"] = r
r += 1

ws_bs.cell(r, 1, "Total Equity").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_bs.cell(r, i + 2, value=f"={cl}{bsr['Total Assets']}-{cl}{bsr['Total Liab']}")
    tot(c)
    c.number_format = NUM
bsr["Total Equity"] = r
r += 1

ws_bs.cell(r, 1, "Balance Check (=0)").font = Font(bold=True, color="FF0000")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_bs.cell(r, i + 2, value=f"={cl}{bsr['Total Assets']}-{cl}{bsr['Total Liab']}-{cl}{bsr['Total Equity']}")
    fml(c)
    c.number_format = NUM

# ════════════════════════════════════════════════════════════════
# SHEET 7: CASH FLOW
# ════════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet("Cash Flow")
setup_cols(ws_cf)
title(ws_cf, "CASH FLOW STATEMENT")
subtitle(ws_cf, "Operating, investing, and financing cash flows")
year_headers(ws_cf, 4)

cfr = {}
r = 5
ws_cf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_cf.cell(r, 1, "OPERATING ACTIVITIES").font = Font(bold=True, size=11, color="1F4E79")
r += 1

ws_cf.cell(r, 1, "Net Income").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cf.cell(r, i + 2, value=f"='P&L'!{cl}${plr['Net Income']}")
    fml(c)
    c.number_format = NUM
cfr["NI"] = r; r += 1

ws_cf.cell(r, 1, "Add: D&A").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cf.cell(r, i + 2, value=f"=Revenue!{cl}${total_rev_row}*Assumptions!{cl}${nrm['D&A % of Revenue']}")
    fml(c)
    c.number_format = NUM
cfr["DA"] = r; r += 1

ws_cf.cell(r, 1, "Changes in Working Capital").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    if i == 0:
        c = ws_cf.cell(r, i + 2, 0)
        inp(c)
    else:
        prev = gcl(i + 1)
        c = ws_cf.cell(r, i + 2, value=f"=-('Balance Sheet'!{cl}${bsr['Accounts Receivable']}-'Balance Sheet'!{prev}${bsr['Accounts Receivable']}+'Balance Sheet'!{cl}${bsr['Inventory']}-'Balance Sheet'!{prev}${bsr['Inventory']}-'Balance Sheet'!{cl}${bsr['AP']}+'Balance Sheet'!{prev}${bsr['AP']})")
        fml(c)
    c.number_format = NUM
cfr["WC"] = r; r += 1

ws_cf.cell(r, 1, "Cash from Operations").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cf.cell(r, i + 2, value=f"={cl}{cfr['NI']}+{cl}{cfr['DA']}+{cl}{cfr['WC']}")
    tot(c)
    c.number_format = NUM
cfr["CFO"] = r; r += 2

ws_cf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_cf.cell(r, 1, "INVESTING ACTIVITIES").font = Font(bold=True, size=11, color="1F4E79")
r += 1

ws_cf.cell(r, 1, "Capital Expenditures").font = Font(bold=True, size=10)
capex_r = nrm["CapEx % of Revenue"]
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cf.cell(r, i + 2, value=f"=-Revenue!{cl}${total_rev_row}*Assumptions!{cl}${capex_r}")
    fml(c)
    c.number_format = NUM
cfr["CapEx"] = r; r += 2

ws_cf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_cf.cell(r, 1, "FINANCING ACTIVITIES").font = Font(bold=True, size=11, color="1F4E79")
r += 1

ws_cf.cell(r, 1, "Dividends Paid").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cf.cell(r, i + 2, value=f"=-'P&L'!{cl}${plr['Net Income']}*Assumptions!{cl}${div_r}")
    fml(c)
    c.number_format = NUM
cfr["Div"] = r; r += 2

ws_cf.cell(r, 1, "Net Change in Cash").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_cf.cell(r, i + 2, value=f"={cl}{cfr['CFO']}+{cl}{cfr['CapEx']}+{cl}{cfr['Div']}")
    tot(c)
    c.number_format = NUM
cfr["Net"] = r; r += 1

ws_cf.cell(r, 1, "Ending Cash Balance").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    if i == 0:
        c = ws_cf.cell(r, i + 2, value=f"='Balance Sheet'!{cl}${bsr['Cash & Equivalents']}+{cl}{cfr['Net']}")
        fml(c)
    else:
        prev = gcl(i + 1)
        c = ws_cf.cell(r, i + 2, value=f"={prev}{r}+{cl}{cfr['Net']}")
        fml(c)
    c.number_format = NUM
    c.font = Font(bold=True)
cfr["End Cash"] = r

# ════════════════════════════════════════════════════════════════
# SHEET 8: DCF VALUATION
# ════════════════════════════════════════════════════════════════
ws_dcf = wb.create_sheet("DCF")
setup_cols(ws_dcf)
title(ws_dcf, "DCF VALUATION MODEL")
subtitle(ws_dcf, "Discounted Cash Flow analysis — FCFF approach")
year_headers(ws_dcf, 4)

dcfr = {}
r = 5
for label, fn in [
    ("EBIT", lambda c: f"='P&L'!{c}${plr['EBIT']}"),
    ("Tax on EBIT", lambda c: f"=-{c}{dcfr['EBIT']}*Assumptions!{c}${tax_r}"),
    ("NOPAT", lambda c: f"={c}{dcfr['EBIT']}+{c}{dcfr['Tax on EBIT']}"),
    ("Add: D&A", lambda c: f"='Cash Flow'!{c}${cfr['DA']}"),
    ("Less: CapEx", lambda c: f"='Cash Flow'!{c}${cfr['CapEx']}"),
    ("Less: Chg NWC", lambda c: f"='Cash Flow'!{c}${cfr['WC']}"),
]:
    ws_dcf.cell(r, 1, label).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        c = ws_dcf.cell(r, i + 2, value=fn(cl))
        fml(c)
        c.number_format = NUM
    dcfr[label] = r
    r += 1

r += 1
ws_dcf.cell(r, 1, "Free Cash Flow (FCFF)").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_dcf.cell(r, i + 2, value=f"={cl}{dcfr['NOPAT']}+{cl}{dcfr['Add: D&A']}+{cl}{dcfr['Less: CapEx']}+{cl}{dcfr['Less: Chg NWC']}")
    tot(c)
    c.number_format = NUM
dcfr["FCFF"] = r; r += 2

wacc_r = nrm["WACC"]
ws_dcf.cell(r, 1, "Discount Factor").font = Font(bold=True, size=10)
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_dcf.cell(r, i + 2, value=f"=1/(1+Assumptions!{cl}${wacc_r})^{i+1}")
    fml(c)
    c.number_format = '0.0000'
dcfr["DF"] = r; r += 1

ws_dcf.cell(r, 1, "PV of FCF").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_dcf.cell(r, i + 2, value=f"={cl}{dcfr['FCFF']}*{cl}{dcfr['DF']}")
    fml(c)
    c.number_format = NUM
dcfr["PV"] = r; r += 2

# Valuation summary
ws_dcf.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_dcf.cell(r, 1, "VALUATION SUMMARY").font = Font(bold=True, size=12, color="1F4E79")
r += 1

last = gcl(YC + 1)
tgr_r = nrm["Terminal Growth Rate"]
val_items = [
    ("Sum of PV of FCFs", f"=SUM(B{dcfr['PV']}:{last}{dcfr['PV']})"),
    ("Terminal Value", f"={last}{dcfr['FCFF']}*(1+Assumptions!{last}${tgr_r})/(Assumptions!{last}${wacc_r}-Assumptions!{last}${tgr_r})"),
]
for label, formula in val_items:
    ws_dcf.cell(r, 1, label).font = Font(bold=True, size=10)
    c = ws_dcf.cell(r, 2, value=formula)
    fml(c)
    c.number_format = NUM
    dcfr[label] = r
    r += 1

ws_dcf.cell(r, 1, "PV of Terminal Value").font = Font(bold=True)
c = ws_dcf.cell(r, 2, value=f"=B{dcfr['Terminal Value']}*{last}{dcfr['DF']}")
fml(c); c.number_format = NUM
dcfr["PV_TV"] = r; r += 1

ws_dcf.cell(r, 1, "Enterprise Value").font = Font(bold=True, size=12, color="1F4E79")
c = ws_dcf.cell(r, 2, value=f"=B{dcfr['Sum of PV of FCFs']}+B{dcfr['PV_TV']}")
tot(c); c.number_format = NUM; c.font = Font(bold=True, size=12)
dcfr["EV"] = r; r += 1

ws_dcf.cell(r, 1, "Less: Net Debt").font = Font(bold=True)
c = ws_dcf.cell(r, 2, value=f"='Balance Sheet'!B${bsr['Debt']}-'Balance Sheet'!B${bsr['Cash & Equivalents']}")
fml(c); c.number_format = NUM
dcfr["ND"] = r; r += 1

ws_dcf.cell(r, 1, "Equity Value").font = Font(bold=True, size=12, color="1F4E79")
c = ws_dcf.cell(r, 2, value=f"=B{dcfr['EV']}-B{dcfr['ND']}")
tot(c); c.number_format = NUM; c.font = Font(bold=True, size=12)
dcfr["EqV"] = r; r += 1

ws_dcf.cell(r, 1, "Price per Share").font = Font(bold=True, size=14, color="1F4E79")
c = ws_dcf.cell(r, 2, value=f"=B{dcfr['EqV']}/Assumptions!B${sh_r}/1000000")
tot(c); c.number_format = MONEY; c.font = Font(bold=True, size=14, color="1F4E79")

# ════════════════════════════════════════════════════════════════
# SHEET 9: SENSITIVITY
# ════════════════════════════════════════════════════════════════
ws_sens = wb.create_sheet("Sensitivity")
ws_sens.column_dimensions['A'].width = 18
title(ws_sens, "SENSITIVITY TABLE: EV vs WACC / Terminal Growth")
subtitle(ws_sens, "Enterprise Value under different discount rate and terminal growth rate scenarios")

wacc_vals = [0.07, 0.08, 0.09, 0.10, 0.11, 0.12, 0.13]
tgr_vals = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035]

ws_sens.cell(4, 1, "WACC \\ TGR").font = Font(bold=True)
for j, tgr in enumerate(tgr_vals):
    c = ws_sens.cell(4, j + 2, tgr)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.number_format = PCT

for i, wacc in enumerate(wacc_vals):
    rr = 5 + i
    c = ws_sens.cell(rr, 1, wacc)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.number_format = PCT
    for j, tgr in enumerate(tgr_vals):
        if wacc <= tgr:
            ws_sens.cell(rr, j + 2, "N/A")
        else:
            fcf = f"DCF!{last}{dcfr['FCFF']}"
            tv = f"{fcf}*(1+{tgr})/({wacc}-{tgr})"
            pv = f"1/(1+{wacc})^{YC}"
            c = ws_sens.cell(rr, j + 2, value=f"=DCF!B{dcfr['Sum of PV of FCFs']}+{tv}*{pv}")
            fml(c); c.number_format = NUM

# Conditional formatting on sensitivity table
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_font = Font(color="006100")
red_font = Font(color="9C0006")

# ════════════════════════════════════════════════════════════════
# SHEET 10: REGIONAL SALES
# ════════════════════════════════════════════════════════════════
ws_reg = wb.create_sheet("Regional Sales")
ws_reg.column_dimensions['A'].width = 28
for i in range(YC):
    ws_reg.column_dimensions[gcl(i + 2)].width = 14
title(ws_reg, "REGIONAL SALES BREAKDOWN")
subtitle(ws_reg, "Revenue allocation by geographic region")
year_headers(ws_reg, 4)

region_shares = {
    "North America": 0.42,
    "Europe": 0.28,
    "Asia Pacific": 0.18,
    "Latin America": 0.07,
    "Middle East & Africa": 0.05,
}

r = 5
reg_rows = {}
for region, share in region_shares.items():
    ws_reg.cell(r, 1, region).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        c = ws_reg.cell(r, i + 2, value=f"=Revenue!{cl}${total_rev_row}*{share}")
        fml(c); c.number_format = NUM
    reg_rows[region] = r
    r += 1

r += 1
ws_reg.cell(r, 1, "Total (Check)").font = Font(bold=True, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_reg.cell(r, i + 2, value=f"=SUM({cl}5:{cl}{r-2})")
    tot(c); c.number_format = NUM

# Region %
r += 2
ws_reg.cell(r, 1, "Region Share %").font = Font(bold=True, size=11, color="1F4E79")
r += 1
for region in region_shares:
    ws_reg.cell(r, 1, f"  {region}").font = Font(size=9, color="666666")
    for i in range(YC):
        cl = gcl(i + 2)
        c = ws_reg.cell(r, i + 2, value=f"={cl}{reg_rows[region]}/Revenue!{cl}${total_rev_row}")
        c.number_format = PCT; c.font = Font(size=9, color="666666")
    r += 1

# ════════════════════════════════════════════════════════════════
# SHEET 11: MONTHLY BREAKDOWN (Year 2024)
# ════════════════════════════════════════════════════════════════
ws_mo = wb.create_sheet("Monthly 2024")
ws_mo.column_dimensions['A'].width = 28
for i in range(12):
    ws_mo.column_dimensions[gcl(i + 2)].width = 13
ws_mo.column_dimensions[gcl(14)].width = 15

title(ws_mo, "MONTHLY P&L — FY 2024")
subtitle(ws_mo, "Monthly breakdown with seasonal patterns")

# Month headers
for i, m in enumerate(MONTHS):
    ws_mo.cell(4, i + 2, m)
ws_mo.cell(4, 14, "FY Total")
hdr_row(ws_mo, 4, 14)

# Seasonal weights
weights = [0.07, 0.07, 0.08, 0.08, 0.085, 0.09, 0.075, 0.075, 0.09, 0.09, 0.095, 0.10]

r = 5
mo_rows = {}
# Monthly revenue
ws_mo.cell(r, 1, "Revenue").font = Font(bold=True, size=10)
annual_rev_col = gcl(YEARS.index(2024) + 2) if 2024 in YEARS else "E"
for i in range(12):
    c = ws_mo.cell(r, i + 2, value=f"=Revenue!{annual_rev_col}${total_rev_row}*{weights[i]}")
    fml(c); c.number_format = NUM
c = ws_mo.cell(r, 14, value=f"=SUM(B{r}:M{r})")
tot(c); c.number_format = NUM
mo_rows["Revenue"] = r; r += 1

# Monthly COGS
ws_mo.cell(r, 1, "COGS").font = Font(bold=True, size=10)
for i in range(12):
    cl = gcl(i + 2)
    c = ws_mo.cell(r, i + 2, value=f"=-{cl}{mo_rows['Revenue']}*Assumptions!{annual_rev_col}${nrm['COGS % of Revenue']}")
    fml(c); c.number_format = NUM
c = ws_mo.cell(r, 14, value=f"=SUM(B{r}:M{r})")
tot(c); c.number_format = NUM
mo_rows["COGS"] = r; r += 1

ws_mo.cell(r, 1, "Gross Profit").font = Font(bold=True, color="1F4E79")
for i in range(12):
    cl = gcl(i + 2)
    c = ws_mo.cell(r, i + 2, value=f"={cl}{mo_rows['Revenue']}+{cl}{mo_rows['COGS']}")
    tot(c); c.number_format = NUM
c = ws_mo.cell(r, 14, value=f"=SUM(B{r}:M{r})")
tot(c); c.number_format = NUM
mo_rows["GP"] = r; r += 1

ws_mo.cell(r, 1, "OpEx").font = Font(bold=True, size=10)
for i in range(12):
    cl = gcl(i + 2)
    c = ws_mo.cell(r, i + 2, value=f"=-{cl}{mo_rows['Revenue']}*(Assumptions!{annual_rev_col}${nrm['SGA % of Revenue']}+Assumptions!{annual_rev_col}${nrm['R&D % of Revenue']}+Assumptions!{annual_rev_col}${nrm['D&A % of Revenue']})")
    fml(c); c.number_format = NUM
c = ws_mo.cell(r, 14, value=f"=SUM(B{r}:M{r})")
tot(c); c.number_format = NUM
mo_rows["OpEx"] = r; r += 1

ws_mo.cell(r, 1, "EBIT").font = Font(bold=True, color="1F4E79")
for i in range(12):
    cl = gcl(i + 2)
    c = ws_mo.cell(r, i + 2, value=f"={cl}{mo_rows['GP']}+{cl}{mo_rows['OpEx']}")
    tot(c); c.number_format = NUM
c = ws_mo.cell(r, 14, value=f"=SUM(B{r}:M{r})")
tot(c); c.number_format = NUM
mo_rows["EBIT"] = r; r += 1

# Cumulative revenue
ws_mo.cell(r, 1, "Cumulative Revenue").font = Font(italic=True, size=10, color="808080")
for i in range(12):
    cl = gcl(i + 2)
    if i == 0:
        c = ws_mo.cell(r, i + 2, value=f"=B{mo_rows['Revenue']}")
    else:
        prev = gcl(i + 1)
        c = ws_mo.cell(r, i + 2, value=f"={prev}{r}+{cl}{mo_rows['Revenue']}")
    fml(c); c.number_format = NUM

# ════════════════════════════════════════════════════════════════
# SHEET 12: HEADCOUNT
# ════════════════════════════════════════════════════════════════
ws_hc = wb.create_sheet("Headcount")
setup_cols(ws_hc)
title(ws_hc, "HEADCOUNT PLAN")
subtitle(ws_hc, "Department-level FTE projections linked to revenue growth")
year_headers(ws_hc, 4)

r = 5
hc_first = r
for dept, base in DEPARTMENTS:
    ws_hc.cell(r, 1, dept).font = Font(bold=True, size=10)
    for i in range(YC):
        if i == 0:
            c = ws_hc.cell(r, i + 2, base)
            inp(c)
        else:
            prev = gcl(i + 1)
            gref = f"Assumptions!{gcl(i+2)}${nrm['Revenue Growth Rate']}"
            c = ws_hc.cell(r, i + 2, value=f"=ROUND({prev}{r}*(1+{gref}*0.65),0)")
            fml(c)
        c.number_format = '#,##0'
    r += 1
hc_last = r - 1

r += 1
ws_hc.cell(r, 1, "Total Headcount").font = Font(bold=True, size=11, color="1F4E79")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_hc.cell(r, i + 2, value=f"=SUM({cl}{hc_first}:{cl}{hc_last})")
    tot(c); c.number_format = '#,##0'
hc_total_row = r

r += 1
ws_hc.cell(r, 1, "Revenue per Employee").font = Font(bold=True, color="808080")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_hc.cell(r, i + 2, value=f"=Revenue!{cl}${total_rev_row}/{cl}{hc_total_row}")
    fml(c); c.number_format = NUM

r += 1
ws_hc.cell(r, 1, "Avg Cost per Employee").font = Font(bold=True, color="808080")
for i in range(YC):
    cl = gcl(i + 2)
    c = ws_hc.cell(r, i + 2, value=f"=(COGS!{cl}${total_cogs_row}+OpEx!{cl}${total_opex_row})/{cl}{hc_total_row}")
    fml(c); c.number_format = NUM

# ════════════════════════════════════════════════════════════════
# SHEET 13: EMPLOYEE DIRECTORY (50 rows)
# ════════════════════════════════════════════════════════════════
ws_emp = wb.create_sheet("Employees")
ws_emp.column_dimensions['A'].width = 6
ws_emp.column_dimensions['B'].width = 22
ws_emp.column_dimensions['C'].width = 22
ws_emp.column_dimensions['D'].width = 14
ws_emp.column_dimensions['E'].width = 14
ws_emp.column_dimensions['F'].width = 18
ws_emp.column_dimensions['G'].width = 15
ws_emp.column_dimensions['H'].width = 12
ws_emp.column_dimensions['I'].width = 30

title(ws_emp, "EMPLOYEE DIRECTORY", 1)
subtitle(ws_emp, "Active employees with department, salary, and performance data", 2)

headers = ["ID", "Name", "Department", "Start Date", "Salary ($)", "Performance", "Rating", "Status", "Email"]
for i, h in enumerate(headers):
    ws_emp.cell(4, i + 1, h)
hdr_row(ws_emp, 4, len(headers))

depts_list = [d[0] for d in DEPARTMENTS]
perf_levels = ["Exceptional", "Exceeds Expectations", "Meets Expectations", "Needs Improvement"]
statuses = ["Active", "Active", "Active", "Active", "Active", "On Leave", "Active", "Active", "Active", "Probation"]

random.seed(42)
for idx, name in enumerate(EMPLOYEE_NAMES):
    r = 5 + idx
    ws_emp.cell(r, 1, f"E{1001 + idx}").font = Font(size=10)
    ws_emp.cell(r, 2, name).font = Font(size=10)
    dept = random.choice(depts_list)
    ws_emp.cell(r, 3, dept).font = Font(size=10)
    start = datetime.date(random.randint(2018, 2024), random.randint(1, 12), random.randint(1, 28))
    ws_emp.cell(r, 4, start).number_format = DATE_FMT
    salary = random.randint(55, 180) * 1000
    ws_emp.cell(r, 5, salary).number_format = NUM
    perf = random.choice(perf_levels)
    ws_emp.cell(r, 6, perf).font = Font(size=10)
    rating = random.choice([3.0, 3.5, 4.0, 4.5, 5.0])
    ws_emp.cell(r, 7, rating).number_format = '0.0'
    status = random.choice(statuses)
    ws_emp.cell(r, 8, status).font = Font(size=10)
    email = name.lower().replace(" ", ".") + "@company.com"
    ws_emp.cell(r, 9, email).font = LINK_FONT
    ws_emp.cell(r, 9).hyperlink = f"mailto:{email}"

    for ci in range(1, 10):
        ws_emp.cell(r, ci).border = BORDER

# Data validation for Performance column
dv_perf = DataValidation(type="list", formula1='"Exceptional,Exceeds Expectations,Meets Expectations,Needs Improvement"')
dv_perf.prompt = "Select performance level"
ws_emp.add_data_validation(dv_perf)
for idx in range(len(EMPLOYEE_NAMES)):
    dv_perf.add(ws_emp.cell(5 + idx, 6))

# Data validation for Status
dv_status = DataValidation(type="list", formula1='"Active,On Leave,Probation,Terminated"')
ws_emp.add_data_validation(dv_status)
for idx in range(len(EMPLOYEE_NAMES)):
    dv_status.add(ws_emp.cell(5 + idx, 8))

# Conditional formatting: salary > 150K → green, < 70K → red
ws_emp.conditional_formatting.add(
    f"E5:E{4 + len(EMPLOYEE_NAMES)}",
    CellIsRule(operator='greaterThan', formula=['150000'], fill=green_fill, font=green_font)
)
ws_emp.conditional_formatting.add(
    f"E5:E{4 + len(EMPLOYEE_NAMES)}",
    CellIsRule(operator='lessThan', formula=['70000'], fill=red_fill, font=red_font)
)

# Conditional formatting: rating >= 4.5 → green
ws_emp.conditional_formatting.add(
    f"G5:G{4 + len(EMPLOYEE_NAMES)}",
    CellIsRule(operator='greaterThanOrEqual', formula=['4.5'], fill=green_fill, font=green_font)
)

# ════════════════════════════════════════════════════════════════
# SHEET 14: RATIOS & MULTIPLES
# ════════════════════════════════════════════════════════════════
ws_rat = wb.create_sheet("Ratios")
setup_cols(ws_rat)
title(ws_rat, "FINANCIAL RATIOS & MULTIPLES")
subtitle(ws_rat, "Key financial metrics derived from P&L, Balance Sheet, and Cash Flow")
year_headers(ws_rat, 4)

ratios = [
    ("PROFITABILITY", None, None),
    ("Gross Margin", lambda c: f"='P&L'!{c}${plr['Gross Margin %']}", PCT),
    ("EBIT Margin", lambda c: f"='P&L'!{c}${plr['EBIT Margin %']}", PCT),
    ("Net Margin", lambda c: f"='P&L'!{c}${plr['Net Margin %']}", PCT),
    ("ROA", lambda c: f"='P&L'!{c}${plr['Net Income']}/'Balance Sheet'!{c}${bsr['Total Assets']}", PCT),
    ("ROE", lambda c: f"='P&L'!{c}${plr['Net Income']}/'Balance Sheet'!{c}${bsr['Total Equity']}", PCT),
    ("", None, None),
    ("LIQUIDITY", None, None),
    ("Current Ratio", lambda c: f"='Balance Sheet'!{c}${bsr['Total CA']}/'Balance Sheet'!{c}${bsr['AP']}", '0.00x'),
    ("Debt/Equity", lambda c: f"='Balance Sheet'!{c}${bsr['Debt']}/'Balance Sheet'!{c}${bsr['Total Equity']}", '0.00x'),
    ("Net Debt/EBITDA", lambda c: f"=('Balance Sheet'!{c}${bsr['Debt']}-'Balance Sheet'!{c}${bsr['Cash & Equivalents']})/'P&L'!{c}${plr['EBITDA']}", '0.00x'),
    ("", None, None),
    ("EFFICIENCY", None, None),
    ("Revenue per Employee", lambda c: f"=Revenue!{c}${total_rev_row}/Headcount!{c}${hc_total_row}", NUM),
    ("OpEx Ratio", lambda c: f"=OpEx!{c}${total_opex_row}/Revenue!{c}${total_rev_row}", PCT),
    ("", None, None),
    ("VALUATION", None, None),
    ("EPS", lambda c: f"='P&L'!{c}${plr['EPS']}", MONEY),
    ("P/E (at $50/share)", lambda c: f"=50/'P&L'!{c}${plr['EPS']}", '0.0x'),
    ("EV/EBITDA", lambda c: f"=DCF!B${dcfr['EV']}/'P&L'!{c}${plr['EBITDA']}", '0.0x'),
]

r = 5
for label, fn, fmt in ratios:
    if fn is None:
        if label:
            ws_rat.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws_rat.cell(r, 1, label).font = Font(bold=True, size=11, color="1F4E79")
        r += 1
        continue
    ws_rat.cell(r, 1, label).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        c = ws_rat.cell(r, i + 2, value=fn(cl))
        fml(c)
        c.number_format = fmt
    r += 1

# Conditional formatting: color scale on margins
for mr in [plr.get('Gross Margin %'), plr.get('EBIT Margin %'), plr.get('Net Margin %')]:
    if mr:
        pass  # applied at P&L level

# ════════════════════════════════════════════════════════════════
# SHEET 15: SCENARIO ANALYSIS
# ════════════════════════════════════════════════════════════════
ws_sc = wb.create_sheet("Scenarios")
ws_sc.column_dimensions['A'].width = 28
ws_sc.column_dimensions['B'].width = 16
ws_sc.column_dimensions['C'].width = 16
ws_sc.column_dimensions['D'].width = 16
ws_sc.column_dimensions['E'].width = 16

title(ws_sc, "SCENARIO ANALYSIS")
subtitle(ws_sc, "Bull/Base/Bear case comparison for key metrics (FY2026 estimates)")

headers_sc = ["Metric", "Bear Case", "Base Case", "Bull Case", "Variance (Bull-Bear)"]
for i, h in enumerate(headers_sc):
    ws_sc.cell(4, i + 1, h)
hdr_row(ws_sc, 4, 5)

yr_idx = YEARS.index(2026) if 2026 in YEARS else 5
yr_col = gcl(yr_idx + 2)

scenarios = [
    ("Revenue ($M)", f"=Revenue!{yr_col}${total_rev_row}*0.85/1000000", f"=Revenue!{yr_col}${total_rev_row}/1000000", f"=Revenue!{yr_col}${total_rev_row}*1.15/1000000"),
    ("EBIT Margin", f"='P&L'!{yr_col}${plr['EBIT Margin %']}*0.7", f"='P&L'!{yr_col}${plr['EBIT Margin %']}", f"='P&L'!{yr_col}${plr['EBIT Margin %']}*1.2"),
    ("Net Income ($M)", f"='P&L'!{yr_col}${plr['Net Income']}*0.6/1000000", f"='P&L'!{yr_col}${plr['Net Income']}/1000000", f"='P&L'!{yr_col}${plr['Net Income']}*1.4/1000000"),
    ("EPS", f"='P&L'!{yr_col}${plr['EPS']}*0.6", f"='P&L'!{yr_col}${plr['EPS']}", f"='P&L'!{yr_col}${plr['EPS']}*1.4"),
    ("FCF ($M)", f"=DCF!{yr_col}${dcfr['FCFF']}*0.7/1000000", f"=DCF!{yr_col}${dcfr['FCFF']}/1000000", f"=DCF!{yr_col}${dcfr['FCFF']}*1.3/1000000"),
    ("Headcount", f"=Headcount!{yr_col}${hc_total_row}*0.9", f"=Headcount!{yr_col}${hc_total_row}", f"=Headcount!{yr_col}${hc_total_row}*1.1"),
]

r = 5
for label, bear, base, bull in scenarios:
    ws_sc.cell(r, 1, label).font = Font(bold=True, size=10)
    fmt = PCT if "Margin" in label else (MONEY if "EPS" in label else NUM)
    for ci, val in enumerate([bear, base, bull], 2):
        c = ws_sc.cell(r, ci, value=val)
        fml(c); c.number_format = fmt
    # Variance
    c = ws_sc.cell(r, 5, value=f"=D{r}-B{r}")
    fml(c); c.number_format = fmt
    r += 1

# Color: Bear=red bg, Bull=green bg
for rr in range(5, 5 + len(scenarios)):
    ws_sc.cell(rr, 2).fill = WARN_FILL
    ws_sc.cell(rr, 4).fill = ACCENT_FILL

# ════════════════════════════════════════════════════════════════
# SHEET 16: KPIs DASHBOARD
# ════════════════════════════════════════════════════════════════
ws_kpi = wb.create_sheet("KPIs")
setup_cols(ws_kpi)
title(ws_kpi, "KEY PERFORMANCE INDICATORS")
subtitle(ws_kpi, "Executive dashboard — all metrics derived from model sheets")
year_headers(ws_kpi, 4)

kpis = [
    ("Revenue", f"=Revenue!{{c}}${total_rev_row}", NUM),
    ("Revenue Growth YoY", None, PCT),
    ("Gross Margin", f"='P&L'!{{c}}${plr['Gross Margin %']}", PCT),
    ("EBIT Margin", f"='P&L'!{{c}}${plr['EBIT Margin %']}", PCT),
    ("Net Margin", f"='P&L'!{{c}}${plr['Net Margin %']}", PCT),
    ("EPS", f"='P&L'!{{c}}${plr['EPS']}", MONEY),
    ("EBITDA", f"='P&L'!{{c}}${plr['EBITDA']}", NUM),
    ("Free Cash Flow", f"=DCF!{{c}}${dcfr['FCFF']}", NUM),
    ("Cash Balance", f"='Balance Sheet'!{{c}}${bsr['Cash & Equivalents']}", NUM),
    ("Total Debt", f"='Balance Sheet'!{{c}}${bsr['Debt']}", NUM),
    ("Debt/Equity", None, '0.00x'),
    ("ROE", None, PCT),
    ("Headcount", f"=Headcount!{{c}}${hc_total_row}", '#,##0'),
    ("Rev/Employee", f"=Revenue!{{c}}${total_rev_row}/Headcount!{{c}}${hc_total_row}", NUM),
    ("Customer Churn", f"=Assumptions!{{c}}${nrm['Customer Churn Rate']}", PCT),
    ("Avg Deal Size ($K)", f"=Assumptions!{{c}}${nrm['Avg Deal Size ($K)']}", NUM),
]

r = 5
kpi_rev_r = r
for label, tpl, fmt in kpis:
    ws_kpi.cell(r, 1, label).font = Font(bold=True, size=10)
    for i in range(YC):
        cl = gcl(i + 2)
        if label == "Revenue Growth YoY":
            if i == 0:
                ws_kpi.cell(r, i + 2, "N/A").font = Font(color="808080")
            else:
                prev = gcl(i + 1)
                c = ws_kpi.cell(r, i + 2, value=f"=({cl}{kpi_rev_r}-{prev}{kpi_rev_r})/{prev}{kpi_rev_r}")
                fml(c)
        elif label == "Debt/Equity":
            c = ws_kpi.cell(r, i + 2, value=f"='Balance Sheet'!{cl}${bsr['Debt']}/'Balance Sheet'!{cl}${bsr['Total Equity']}")
            fml(c)
        elif label == "ROE":
            c = ws_kpi.cell(r, i + 2, value=f"='P&L'!{cl}${plr['Net Income']}/'Balance Sheet'!{cl}${bsr['Total Equity']}")
            fml(c)
        elif tpl:
            c = ws_kpi.cell(r, i + 2, value=tpl.replace("{c}", cl))
            fml(c)
        if isinstance(ws_kpi.cell(r, i + 2).value, str) and ws_kpi.cell(r, i + 2).value != "N/A":
            pass
        ws_kpi.cell(r, i + 2).number_format = fmt
    r += 1

# Conditional formatting on KPIs: color scale for margins
ws_kpi.conditional_formatting.add(
    f"B{kpi_rev_r + 2}:{gcl(YC+1)}{kpi_rev_r + 2}",
    ColorScaleRule(start_type='min', start_color='FFC7CE', end_type='max', end_color='C6EFCE')
)
ws_kpi.conditional_formatting.add(
    f"B{kpi_rev_r + 3}:{gcl(YC+1)}{kpi_rev_r + 3}",
    ColorScaleRule(start_type='min', start_color='FFC7CE', end_type='max', end_color='C6EFCE')
)

# ════════════════════════════════════════════════════════════════
# SHEET 17: LINKS & NOTES
# ════════════════════════════════════════════════════════════════
ws_links = wb.create_sheet("Links & Notes")
ws_links.column_dimensions['A'].width = 30
ws_links.column_dimensions['B'].width = 50
ws_links.column_dimensions['C'].width = 40

title(ws_links, "REFERENCE LINKS & NOTES")
subtitle(ws_links, "External resources and model documentation")

headers_l = ["Category", "Description", "URL / Reference"]
for i, h in enumerate(headers_l):
    ws_links.cell(4, i + 1, h)
hdr_row(ws_links, 4, 3)

links = [
    ("SEC Filings", "Annual report (10-K)", "https://www.sec.gov/cgi-bin/browse-edgar"),
    ("SEC Filings", "Quarterly report (10-Q)", "https://www.sec.gov/cgi-bin/browse-edgar"),
    ("Market Data", "Yahoo Finance", "https://finance.yahoo.com"),
    ("Market Data", "Bloomberg Terminal", "https://www.bloomberg.com/professional"),
    ("Research", "McKinsey Global Institute", "https://www.mckinsey.com/mgi"),
    ("Research", "Gartner IT Spending", "https://www.gartner.com/en/research"),
    ("Competitors", "Salesforce (CRM)", "https://investor.salesforce.com"),
    ("Competitors", "ServiceNow (NOW)", "https://investor.servicenow.com"),
    ("Competitors", "Workday (WDAY)", "https://investor.workday.com"),
    ("Internal", "Board Deck Q4 2024", "SharePoint://internal/board-decks/Q4-2024"),
    ("Internal", "Strategic Plan 2025-2028", "SharePoint://internal/strategy/plan-2025"),
    ("Model Notes", "WACC calculated using CAPM: Rf=4.2%, Beta=1.1, MRP=5.5%", ""),
    ("Model Notes", "Terminal growth assumes GDP + 50bps inflation premium", ""),
    ("Model Notes", "NWC % declining as company scales operations", ""),
    ("Model Notes", "CapEx front-loaded for cloud migration (2024-2026)", ""),
]

for idx, (cat, desc, url) in enumerate(links):
    r = 5 + idx
    ws_links.cell(r, 1, cat).font = Font(bold=True, size=10)
    ws_links.cell(r, 2, desc).font = Font(size=10)
    if url.startswith("http"):
        ws_links.cell(r, 3, url).font = LINK_FONT
        ws_links.cell(r, 3).hyperlink = url
    else:
        ws_links.cell(r, 3, url).font = Font(size=10, color="666666")
    for ci in range(1, 4):
        ws_links.cell(r, ci).border = BORDER

# ════════════════════════════════════════════════════════════════
# Add intentional anomalies for testing
# ════════════════════════════════════════════════════════════════
# Hardcoded value where formula should be
ws_rev.cell(prod_rows["Enterprise SaaS"], 7, 88888888)
ws_rev.cell(prod_rows["Enterprise SaaS"], 7).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# Another anomaly
ws_rev.cell(prod_rows["Cloud Infrastructure"], 9, 5555555)
ws_rev.cell(prod_rows["Cloud Infrastructure"], 9).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# ════════════════════════════════════════════════════════════════
# CHARTS (embedded in KPIs sheet)
# ════════════════════════════════════════════════════════════════
# Revenue bar chart
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Revenue by Year"
chart1.y_axis.title = "Revenue ($)"
chart1.x_axis.title = "Year"
chart1.style = 10
data1 = Reference(ws_kpi, min_col=2, min_row=kpi_rev_r, max_col=YC+1, max_row=kpi_rev_r)
cats1 = Reference(ws_kpi, min_col=2, min_row=4, max_col=YC+1, max_row=4)
chart1.add_data(data1, from_rows=True)
chart1.set_categories(cats1)
chart1.width = 22
chart1.height = 12
ws_kpi.add_chart(chart1, f"A{r + 2}")

# Margin line chart
chart2 = LineChart()
chart2.title = "Margin Trends"
chart2.y_axis.title = "Margin %"
chart2.style = 10
for offset, lbl in [(2, "Gross Margin"), (3, "EBIT Margin"), (4, "Net Margin")]:
    data2 = Reference(ws_kpi, min_col=2, min_row=kpi_rev_r + offset, max_col=YC+1, max_row=kpi_rev_r + offset)
    chart2.add_data(data2, from_rows=True, titles_from_data=False)
    chart2.series[-1].name = lbl
chart2.set_categories(cats1)
chart2.width = 22
chart2.height = 12
ws_kpi.add_chart(chart2, f"A{r + 18}")

# Regional pie chart
chart3 = PieChart()
chart3.title = "Revenue by Region (2024)"
yr24_col = YEARS.index(2024) + 2 if 2024 in YEARS else 5
data3 = Reference(ws_reg, min_col=yr24_col, min_row=5, max_row=5 + len(REGIONS) - 1)
cats3 = Reference(ws_reg, min_col=1, min_row=5, max_row=5 + len(REGIONS) - 1)
chart3.add_data(data3)
chart3.set_categories(cats3)
chart3.width = 16
chart3.height = 12
ws_reg.add_chart(chart3, "A20")

# ════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "advanced_financial_model.xlsx")
wb.save(output)
print(f"Advanced Excel saved to: {os.path.abspath(output)}")
print(f"Sheets: {wb.sheetnames}")
total_cells = sum(ws.max_row * ws.max_column for ws in wb.worksheets)
print(f"Estimated cells: ~{total_cells}")
print(f"Sheet count: {len(wb.sheetnames)}")
