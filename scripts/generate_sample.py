"""
Generate a large realistic financial model Excel file for testing Excel AI.
Sheets: Assumptions, Revenue, COGS, OpEx, P&L, Balance Sheet, Cash Flow, DCF, Sensitivity
~2000+ cells with formulas, named ranges, cross-sheet references.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
import random
import os

wb = openpyxl.Workbook()

YEARS = list(range(2024, 2034))
YEAR_COUNT = len(YEARS)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
)
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
MONEY_FMT = '#,##0.00'


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def style_input(cell):
    cell.fill = INPUT_FILL
    cell.border = BORDER


def style_formula(cell):
    cell.fill = FORMULA_FILL
    cell.border = BORDER


def col(c):
    return get_column_letter(c)


# ─── ASSUMPTIONS ───
ws = wb.active
ws.title = "Assumptions"
ws.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws.column_dimensions[col(i + 2)].width = 14

ws.cell(1, 1, "ASSUMPTIONS").font = Font(bold=True, size=14)
ws.cell(2, 1, "Key Model Inputs")

# Year headers
for i, yr in enumerate(YEARS):
    ws.cell(3, i + 2, yr)
style_header_row(ws, 3, YEAR_COUNT + 1)

assumptions = {
    "Revenue Growth Rate": [0.15, 0.12, 0.10, 0.08, 0.07, 0.06, 0.05, 0.05, 0.04, 0.04],
    "COGS % of Revenue": [0.42, 0.41, 0.40, 0.39, 0.38, 0.38, 0.37, 0.37, 0.36, 0.36],
    "SGA % of Revenue": [0.18, 0.17, 0.16, 0.15, 0.15, 0.14, 0.14, 0.13, 0.13, 0.13],
    "R&D % of Revenue": [0.12, 0.11, 0.10, 0.10, 0.09, 0.09, 0.08, 0.08, 0.08, 0.08],
    "D&A % of Revenue": [0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03, 0.03],
    "Tax Rate": [0.21, 0.21, 0.21, 0.21, 0.21, 0.21, 0.21, 0.21, 0.21, 0.21],
    "CapEx % of Revenue": [0.05, 0.05, 0.04, 0.04, 0.04, 0.03, 0.03, 0.03, 0.03, 0.03],
    "NWC % of Revenue": [0.10, 0.10, 0.09, 0.09, 0.09, 0.08, 0.08, 0.08, 0.08, 0.08],
    "WACC": [0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10],
    "Terminal Growth Rate": [0.025, 0.025, 0.025, 0.025, 0.025, 0.025, 0.025, 0.025, 0.025, 0.025],
    "Shares Outstanding (M)": [100, 100, 100, 100, 100, 100, 100, 100, 100, 100],
    "Interest Rate on Debt": [0.045, 0.045, 0.045, 0.045, 0.045, 0.045, 0.045, 0.045, 0.045, 0.045],
    "Dividend Payout Ratio": [0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30],
}

row = 4
named_range_map = {}
for label, values in assumptions.items():
    ws.cell(row, 1, label).font = Font(bold=True)
    for i, val in enumerate(values):
        c = ws.cell(row, i + 2, val)
        style_input(c)
        if "%" in label or "Rate" in label or "Ratio" in label or "Growth" in label:
            c.number_format = PCT_FMT
        else:
            c.number_format = NUM_FMT
    safe_name = label.replace(" ", "_").replace("&", "and").replace("%", "pct").replace("(", "").replace(")", "")
    named_range_map[label] = (row, safe_name)
    row += 1

# Define named ranges for first year column values
for label, (r, safe) in named_range_map.items():
    try:
        ref = f"Assumptions!$B${r}:${col(YEAR_COUNT+1)}${r}"
        wb.defined_names.new(safe, attr_text=ref)
    except Exception:
        pass

# ─── REVENUE ───
ws_rev = wb.create_sheet("Revenue")
ws_rev.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_rev.column_dimensions[col(i + 2)].width = 14

ws_rev.cell(1, 1, "REVENUE BUILD").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_rev.cell(3, i + 2, yr)
style_header_row(ws_rev, 3, YEAR_COUNT + 1)

products = [
    ("Product A - Enterprise SaaS", 50000000),
    ("Product B - SMB Platform", 25000000),
    ("Product C - Data Analytics", 15000000),
    ("Product D - Cloud Services", 10000000),
    ("Professional Services", 8000000),
    ("Maintenance & Support", 5000000),
    ("Licensing", 3000000),
    ("Other Revenue", 2000000),
]

row = 4
product_rows = {}
for name, base in products:
    ws_rev.cell(row, 1, name).font = Font(bold=True)
    for i in range(YEAR_COUNT):
        if i == 0:
            c = ws_rev.cell(row, i + 2, base)
            style_input(c)
        else:
            prev_col = col(i + 1)
            growth_ref = f"Assumptions!{col(i + 2)}${named_range_map['Revenue Growth Rate'][0]}"
            formula = f"={prev_col}{row}*(1+{growth_ref})"
            c = ws_rev.cell(row, i + 2, value=formula)
            style_formula(c)
        c.number_format = NUM_FMT
    product_rows[name] = row
    row += 1

# Total Revenue
row += 1
ws_rev.cell(row, 1, "Total Revenue").font = Font(bold=True, size=11, color="2F5496")
for i in range(YEAR_COUNT):
    first_row = 4
    last_row = first_row + len(products) - 1
    c_letter = col(i + 2)
    formula = f"=SUM({c_letter}{first_row}:{c_letter}{last_row})"
    c = ws_rev.cell(row, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
total_rev_row = row

# ─── COGS ───
ws_cogs = wb.create_sheet("COGS")
ws_cogs.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_cogs.column_dimensions[col(i + 2)].width = 14

ws_cogs.cell(1, 1, "COST OF GOODS SOLD").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_cogs.cell(3, i + 2, yr)
style_header_row(ws_cogs, 3, YEAR_COUNT + 1)

cogs_items = [
    "Cloud Hosting & Infrastructure",
    "Software Licenses (3rd party)",
    "Customer Support Staff",
    "Payment Processing Fees",
    "Data Center Operations",
]

row = 4
for item in cogs_items:
    ws_cogs.cell(row, 1, item).font = Font(bold=True)
    for i in range(YEAR_COUNT):
        c_letter = col(i + 2)
        rev_ref = f"Revenue!{c_letter}${total_rev_row}"
        cogs_pct_ref = f"Assumptions!{c_letter}${named_range_map['COGS % of Revenue'][0]}"
        share = round(1.0 / len(cogs_items), 4)
        formula = f"={rev_ref}*{cogs_pct_ref}*{share}"
        c = ws_cogs.cell(row, i + 2, value=formula)
        style_formula(c)
        c.number_format = NUM_FMT
    row += 1

row += 1
ws_cogs.cell(row, 1, "Total COGS").font = Font(bold=True, size=11, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=SUM({c_letter}4:{c_letter}{row - 2})"
    c = ws_cogs.cell(row, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
total_cogs_row = row

# ─── OPEX ───
ws_opex = wb.create_sheet("OpEx")
ws_opex.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_opex.column_dimensions[col(i + 2)].width = 14

ws_opex.cell(1, 1, "OPERATING EXPENSES").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_opex.cell(3, i + 2, yr)
style_header_row(ws_opex, 3, YEAR_COUNT + 1)

opex_categories = {
    "Sales & Marketing": "SGA % of Revenue",
    "Salaries & Wages": "SGA % of Revenue",
    "Research & Development": "R&D % of Revenue",
    "General & Administrative": "SGA % of Revenue",
    "Depreciation & Amortization": "D&A % of Revenue",
    "Rent & Facilities": "SGA % of Revenue",
    "Travel & Entertainment": "SGA % of Revenue",
    "Professional Fees (Legal, Audit)": "SGA % of Revenue",
    "Insurance": "SGA % of Revenue",
    "IT & Software": "R&D % of Revenue",
}

row = 4
opex_shares = {
    "Sales & Marketing": 0.35,
    "Salaries & Wages": 0.25,
    "Research & Development": 0.60,
    "General & Administrative": 0.10,
    "Depreciation & Amortization": 1.0,
    "Rent & Facilities": 0.08,
    "Travel & Entertainment": 0.05,
    "Professional Fees (Legal, Audit)": 0.07,
    "Insurance": 0.05,
    "IT & Software": 0.40,
}

for item, assumption_key in opex_categories.items():
    ws_opex.cell(row, 1, item).font = Font(bold=True)
    share = opex_shares.get(item, 0.1)
    for i in range(YEAR_COUNT):
        c_letter = col(i + 2)
        rev_ref = f"Revenue!{c_letter}${total_rev_row}"
        pct_ref = f"Assumptions!{c_letter}${named_range_map[assumption_key][0]}"
        formula = f"={rev_ref}*{pct_ref}*{share}"
        c = ws_opex.cell(row, i + 2, value=formula)
        style_formula(c)
        c.number_format = NUM_FMT
    row += 1

row += 1
ws_opex.cell(row, 1, "Total OpEx").font = Font(bold=True, size=11, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=SUM({c_letter}4:{c_letter}{row - 2})"
    c = ws_opex.cell(row, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
total_opex_row = row

# ─── P&L ───
ws_pl = wb.create_sheet("P&L")
ws_pl.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_pl.column_dimensions[col(i + 2)].width = 14

ws_pl.cell(1, 1, "PROFIT & LOSS STATEMENT").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_pl.cell(3, i + 2, yr)
style_header_row(ws_pl, 3, YEAR_COUNT + 1)

pl_rows = {}

def pl_line(ws, row_num, label, formula_fn, fmt=NUM_FMT, bold=False):
    ws.cell(row_num, 1, label).font = Font(bold=bold, size=11 if bold else 10, color="2F5496" if bold else "000000")
    for i in range(YEAR_COUNT):
        c_letter = col(i + 2)
        formula = formula_fn(c_letter)
        c = ws.cell(row_num, i + 2, value=formula)
        style_formula(c)
        c.number_format = fmt
        if bold:
            c.font = Font(bold=True)
    pl_rows[label] = row_num

r = 4
pl_line(ws_pl, r, "Revenue", lambda c: f"=Revenue!{c}${total_rev_row}", bold=True)
r += 1
pl_line(ws_pl, r, "Cost of Goods Sold", lambda c: f"=-COGS!{c}${total_cogs_row}")
r += 1
pl_line(ws_pl, r, "Gross Profit", lambda c: f"={c}{pl_rows['Revenue']}+{c}{pl_rows['Cost of Goods Sold']}", bold=True)
r += 1
pl_line(ws_pl, r, "Gross Margin", lambda c: f"={c}{pl_rows['Gross Profit']}/{c}{pl_rows['Revenue']}", fmt=PCT_FMT)
r += 2
pl_line(ws_pl, r, "Operating Expenses", lambda c: f"=-OpEx!{c}${total_opex_row}")
r += 1
pl_line(ws_pl, r, "EBIT", lambda c: f"={c}{pl_rows['Gross Profit']}+{c}{pl_rows['Operating Expenses']}", bold=True)
r += 1
pl_line(ws_pl, r, "EBIT Margin", lambda c: f"={c}{pl_rows['EBIT']}/{c}{pl_rows['Revenue']}", fmt=PCT_FMT)
r += 2
debt_row_ref = named_range_map["Interest Rate on Debt"][0]
pl_line(ws_pl, r, "Interest Expense", lambda c: f"=-'Balance Sheet'!{c}$16*Assumptions!{c}${debt_row_ref}")
r += 1
pl_line(ws_pl, r, "EBT", lambda c: f"={c}{pl_rows['EBIT']}+{c}{pl_rows['Interest Expense']}", bold=True)
r += 1
tax_ref = named_range_map["Tax Rate"][0]
pl_line(ws_pl, r, "Income Tax", lambda c: f"=-MAX(0,{c}{pl_rows['EBT']}*Assumptions!{c}${tax_ref})")
r += 1
pl_line(ws_pl, r, "Net Income", lambda c: f"={c}{pl_rows['EBT']}+{c}{pl_rows['Income Tax']}", bold=True)
r += 1
pl_line(ws_pl, r, "Net Margin", lambda c: f"={c}{pl_rows['Net Income']}/{c}{pl_rows['Revenue']}", fmt=PCT_FMT)
r += 2
shares_ref = named_range_map["Shares Outstanding (M)"][0]
pl_line(ws_pl, r, "EPS", lambda c: f"={c}{pl_rows['Net Income']}/Assumptions!{c}${shares_ref}/1000000", fmt=MONEY_FMT)

# ─── BALANCE SHEET ───
ws_bs = wb.create_sheet("Balance Sheet")
ws_bs.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_bs.column_dimensions[col(i + 2)].width = 14

ws_bs.cell(1, 1, "BALANCE SHEET").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_bs.cell(3, i + 2, yr)
style_header_row(ws_bs, 3, YEAR_COUNT + 1)

bs_rows = {}
r = 4

# Assets
ws_bs.cell(r, 1, "ASSETS").font = Font(bold=True, size=12, color="2F5496")
r += 1
ws_bs.cell(r, 1, "Cash & Equivalents").font = Font(bold=True)
for i in range(YEAR_COUNT):
    if i == 0:
        c = ws_bs.cell(r, i + 2, 50000000)
        style_input(c)
    else:
        c = ws_bs.cell(r, i + 2, value=f"='Cash Flow'!{col(i+2)}$19")
        style_formula(c)
    c.number_format = NUM_FMT
bs_rows["Cash"] = r
r += 1

ws_bs.cell(r, 1, "Accounts Receivable").font = Font(bold=True)
nwc_ref = named_range_map["NWC % of Revenue"][0]
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${nwc_ref}*0.4"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
bs_rows["AR"] = r
r += 1

ws_bs.cell(r, 1, "Inventory").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${nwc_ref}*0.3"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
bs_rows["Inventory"] = r
r += 1

ws_bs.cell(r, 1, "Prepaid Expenses").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=Revenue!{c_letter}${total_rev_row}*0.02"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
bs_rows["Prepaid"] = r
r += 1

ws_bs.cell(r, 1, "Total Current Assets").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=SUM({c_letter}{bs_rows['Cash']}:{c_letter}{bs_rows['Prepaid']})"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
bs_rows["Total CA"] = r
r += 1

capex_ref = named_range_map["CapEx % of Revenue"][0]
ws_bs.cell(r, 1, "PP&E (Net)").font = Font(bold=True)
for i in range(YEAR_COUNT):
    if i == 0:
        c = ws_bs.cell(r, i + 2, 30000000)
        style_input(c)
    else:
        c_letter = col(i + 2)
        prev = col(i + 1)
        da_ref = named_range_map["D&A % of Revenue"][0]
        formula = f"={prev}{r}+Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${capex_ref}-Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${da_ref}"
        c = ws_bs.cell(r, i + 2, value=formula)
        style_formula(c)
    c.number_format = NUM_FMT
bs_rows["PPE"] = r
r += 1

ws_bs.cell(r, 1, "Intangible Assets").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c = ws_bs.cell(r, i + 2, 20000000)
    style_input(c)
    c.number_format = NUM_FMT
bs_rows["Intangibles"] = r
r += 1

ws_bs.cell(r, 1, "Total Assets").font = Font(bold=True, size=11, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{bs_rows['Total CA']}+{c_letter}{bs_rows['PPE']}+{c_letter}{bs_rows['Intangibles']}"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
bs_rows["Total Assets"] = r
r += 2

# Liabilities
ws_bs.cell(r, 1, "LIABILITIES & EQUITY").font = Font(bold=True, size=12, color="2F5496")
r += 1
ws_bs.cell(r, 1, "Accounts Payable").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${nwc_ref}*0.3"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
bs_rows["AP"] = r
r += 1

ws_bs.cell(r, 1, "Long-Term Debt").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c = ws_bs.cell(r, i + 2, 40000000)
    style_input(c)
    c.number_format = NUM_FMT
bs_rows["Debt"] = r
r += 1

ws_bs.cell(r, 1, "Total Liabilities").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{bs_rows['AP']}+{c_letter}{bs_rows['Debt']}"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
bs_rows["Total Liabilities"] = r
r += 1

ws_bs.cell(r, 1, "Retained Earnings").font = Font(bold=True)
div_ref = named_range_map["Dividend Payout Ratio"][0]
for i in range(YEAR_COUNT):
    if i == 0:
        c = ws_bs.cell(r, i + 2, 30000000)
        style_input(c)
    else:
        c_letter = col(i + 2)
        prev = col(i + 1)
        formula = f"={prev}{r}+'P&L'!{c_letter}${pl_rows['Net Income']}*(1-Assumptions!{c_letter}${div_ref})"
        c = ws_bs.cell(r, i + 2, value=formula)
        style_formula(c)
    c.number_format = NUM_FMT
bs_rows["Retained Earnings"] = r
r += 1

ws_bs.cell(r, 1, "Total Equity").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{bs_rows['Total Assets']}-{c_letter}{bs_rows['Total Liabilities']}"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
bs_rows["Total Equity"] = r
r += 1

ws_bs.cell(r, 1, "Balance Check (should = 0)").font = Font(bold=True, color="FF0000")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{bs_rows['Total Assets']}-{c_letter}{bs_rows['Total Liabilities']}-{c_letter}{bs_rows['Total Equity']}"
    c = ws_bs.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT

# ─── CASH FLOW ───
ws_cf = wb.create_sheet("Cash Flow")
ws_cf.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_cf.column_dimensions[col(i + 2)].width = 14

ws_cf.cell(1, 1, "CASH FLOW STATEMENT").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_cf.cell(3, i + 2, yr)
style_header_row(ws_cf, 3, YEAR_COUNT + 1)

cf_rows = {}
r = 4

ws_cf.cell(r, 1, "OPERATING ACTIVITIES").font = Font(bold=True, size=11, color="2F5496")
r += 1
ws_cf.cell(r, 1, "Net Income").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"='P&L'!{c_letter}${pl_rows['Net Income']}"
    c = ws_cf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
cf_rows["NI"] = r
r += 1

ws_cf.cell(r, 1, "Add: D&A").font = Font(bold=True)
da_ref = named_range_map["D&A % of Revenue"][0]
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${da_ref}"
    c = ws_cf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
cf_rows["DA"] = r
r += 1

ws_cf.cell(r, 1, "Changes in Working Capital").font = Font(bold=True)
for i in range(YEAR_COUNT):
    if i == 0:
        c = ws_cf.cell(r, i + 2, 0)
        style_input(c)
    else:
        c_letter = col(i + 2)
        prev = col(i + 1)
        formula = f"=-('Balance Sheet'!{c_letter}${bs_rows['AR']}-'Balance Sheet'!{prev}${bs_rows['AR']}+'Balance Sheet'!{c_letter}${bs_rows['Inventory']}-'Balance Sheet'!{prev}${bs_rows['Inventory']}-'Balance Sheet'!{c_letter}${bs_rows['AP']}+'Balance Sheet'!{prev}${bs_rows['AP']})"
        c = ws_cf.cell(r, i + 2, value=formula)
        style_formula(c)
    c.number_format = NUM_FMT
cf_rows["WC"] = r
r += 1

ws_cf.cell(r, 1, "Cash from Operations").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{cf_rows['NI']}+{c_letter}{cf_rows['DA']}+{c_letter}{cf_rows['WC']}"
    c = ws_cf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
cf_rows["CFO"] = r
r += 2

ws_cf.cell(r, 1, "INVESTING ACTIVITIES").font = Font(bold=True, size=11, color="2F5496")
r += 1
ws_cf.cell(r, 1, "Capital Expenditures").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=-Revenue!{c_letter}${total_rev_row}*Assumptions!{c_letter}${capex_ref}"
    c = ws_cf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
cf_rows["CapEx"] = r
r += 2

ws_cf.cell(r, 1, "FINANCING ACTIVITIES").font = Font(bold=True, size=11, color="2F5496")
r += 1
ws_cf.cell(r, 1, "Dividends Paid").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=-'P&L'!{c_letter}${pl_rows['Net Income']}*Assumptions!{c_letter}${div_ref}"
    c = ws_cf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
cf_rows["Div"] = r
r += 2

ws_cf.cell(r, 1, "Net Change in Cash").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{cf_rows['CFO']}+{c_letter}{cf_rows['CapEx']}+{c_letter}{cf_rows['Div']}"
    c = ws_cf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
cf_rows["Net Cash"] = r
r += 1

ws_cf.cell(r, 1, "Ending Cash Balance").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    if i == 0:
        c_letter = col(i + 2)
        formula = f"='Balance Sheet'!{c_letter}${bs_rows['Cash']}+{c_letter}{cf_rows['Net Cash']}"
        c = ws_cf.cell(r, i + 2, value=formula)
        style_formula(c)
    else:
        c_letter = col(i + 2)
        prev = col(i + 1)
        formula = f"={prev}{r}+{c_letter}{cf_rows['Net Cash']}"
        c = ws_cf.cell(r, i + 2, value=formula)
        style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
cf_rows["End Cash"] = r

# ─── DCF ───
ws_dcf = wb.create_sheet("DCF")
ws_dcf.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_dcf.column_dimensions[col(i + 2)].width = 14

ws_dcf.cell(1, 1, "DCF VALUATION").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_dcf.cell(3, i + 2, yr)
style_header_row(ws_dcf, 3, YEAR_COUNT + 1)

dcf_rows = {}
r = 4

ws_dcf.cell(r, 1, "EBIT").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"='P&L'!{c_letter}${pl_rows['EBIT']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["EBIT"] = r
r += 1

ws_dcf.cell(r, 1, "Tax on EBIT").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=-{c_letter}{dcf_rows['EBIT']}*Assumptions!{c_letter}${tax_ref}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["Tax"] = r
r += 1

ws_dcf.cell(r, 1, "NOPAT").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{dcf_rows['EBIT']}+{c_letter}{dcf_rows['Tax']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["NOPAT"] = r
r += 1

ws_dcf.cell(r, 1, "Add: D&A").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"='Cash Flow'!{c_letter}${cf_rows['DA']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["DCF_DA"] = r
r += 1

ws_dcf.cell(r, 1, "Less: CapEx").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"='Cash Flow'!{c_letter}${cf_rows['CapEx']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["DCF_CapEx"] = r
r += 1

ws_dcf.cell(r, 1, "Less: Change in NWC").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"='Cash Flow'!{c_letter}${cf_rows['WC']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["DCF_WC"] = r
r += 1

ws_dcf.cell(r, 1, "Free Cash Flow (FCFF)").font = Font(bold=True, size=11, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{dcf_rows['NOPAT']}+{c_letter}{dcf_rows['DCF_DA']}+{c_letter}{dcf_rows['DCF_CapEx']}+{c_letter}{dcf_rows['DCF_WC']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
    c.font = Font(bold=True)
dcf_rows["FCFF"] = r
r += 2

wacc_ref = named_range_map["WACC"][0]
ws_dcf.cell(r, 1, "Discount Factor").font = Font(bold=True)
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    period = i + 1
    formula = f"=1/(1+Assumptions!{c_letter}${wacc_ref})^{period}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = '0.0000'
dcf_rows["DF"] = r
r += 1

ws_dcf.cell(r, 1, "PV of FCF").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"={c_letter}{dcf_rows['FCFF']}*{c_letter}{dcf_rows['DF']}"
    c = ws_dcf.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT
dcf_rows["PV_FCF"] = r
r += 2

# Terminal value and enterprise value
ws_dcf.cell(r, 1, "VALUATION SUMMARY").font = Font(bold=True, size=12, color="2F5496")
r += 1

ws_dcf.cell(r, 1, "Sum of PV of FCFs").font = Font(bold=True)
formula = f"=SUM(B{dcf_rows['PV_FCF']}:{col(YEAR_COUNT+1)}{dcf_rows['PV_FCF']})"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = NUM_FMT
dcf_rows["Sum_PV"] = r
r += 1

tgr_ref = named_range_map["Terminal Growth Rate"][0]
ws_dcf.cell(r, 1, "Terminal Value (Gordon Growth)").font = Font(bold=True)
last_col = col(YEAR_COUNT + 1)
formula = f"={last_col}{dcf_rows['FCFF']}*(1+Assumptions!{last_col}${tgr_ref})/(Assumptions!{last_col}${wacc_ref}-Assumptions!{last_col}${tgr_ref})"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = NUM_FMT
dcf_rows["TV"] = r
r += 1

ws_dcf.cell(r, 1, "PV of Terminal Value").font = Font(bold=True)
formula = f"=B{dcf_rows['TV']}*{last_col}{dcf_rows['DF']}"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = NUM_FMT
dcf_rows["PV_TV"] = r
r += 1

ws_dcf.cell(r, 1, "Enterprise Value").font = Font(bold=True, size=12, color="2F5496")
formula = f"=B{dcf_rows['Sum_PV']}+B{dcf_rows['PV_TV']}"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = NUM_FMT
c.font = Font(bold=True, size=12)
dcf_rows["EV"] = r
r += 1

ws_dcf.cell(r, 1, "Less: Net Debt").font = Font(bold=True)
formula = f"='Balance Sheet'!B${bs_rows['Debt']}-'Balance Sheet'!B${bs_rows['Cash']}"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = NUM_FMT
dcf_rows["Net_Debt"] = r
r += 1

ws_dcf.cell(r, 1, "Equity Value").font = Font(bold=True, size=12, color="2F5496")
formula = f"=B{dcf_rows['EV']}-B{dcf_rows['Net_Debt']}"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = NUM_FMT
c.font = Font(bold=True, size=12)
dcf_rows["Equity_Value"] = r
r += 1

ws_dcf.cell(r, 1, "Price per Share").font = Font(bold=True, size=14, color="2F5496")
formula = f"=B{dcf_rows['Equity_Value']}/Assumptions!B${shares_ref}/1000000"
c = ws_dcf.cell(r, 2, value=formula)
style_formula(c)
c.number_format = MONEY_FMT
c.font = Font(bold=True, size=14, color="2F5496")

# ─── SENSITIVITY TABLE ───
ws_sens = wb.create_sheet("Sensitivity")
ws_sens.column_dimensions['A'].width = 20
ws_sens.cell(1, 1, "SENSITIVITY TABLE: WACC vs Terminal Growth").font = Font(bold=True, size=14)
ws_sens.cell(2, 1, "Enterprise Value at different WACC / TGR combos").font = Font(italic=True, color="666666")

wacc_range = [0.08, 0.09, 0.10, 0.11, 0.12]
tgr_range = [0.01, 0.015, 0.02, 0.025, 0.03, 0.035]

ws_sens.cell(4, 1, "WACC \\ TGR").font = Font(bold=True)
for j, tgr in enumerate(tgr_range):
    c = ws_sens.cell(4, j + 2, tgr)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.number_format = PCT_FMT

for i, wacc in enumerate(wacc_range):
    r = 5 + i
    c = ws_sens.cell(r, 1, wacc)
    c.font = Font(bold=True)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.number_format = PCT_FMT
    for j, tgr in enumerate(tgr_range):
        if wacc <= tgr:
            ws_sens.cell(r, j + 2, "N/A")
        else:
            last_fcf_ref = f"DCF!{last_col}{dcf_rows['FCFF']}"
            tv = f"{last_fcf_ref}*(1+{tgr})/({wacc}-{tgr})"
            pv_factor = f"1/(1+{wacc})^{YEAR_COUNT}"
            formula = f"=DCF!B{dcf_rows['Sum_PV']}+{tv}*{pv_factor}"
            c = ws_sens.cell(r, j + 2, value=formula)
            style_formula(c)
            c.number_format = NUM_FMT

# ─── EMPLOYEES (extra sheet for more data) ───
ws_emp = wb.create_sheet("Headcount")
ws_emp.column_dimensions['A'].width = 25
for i in range(YEAR_COUNT):
    ws_emp.column_dimensions[col(i + 2)].width = 12

ws_emp.cell(1, 1, "HEADCOUNT PLAN").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_emp.cell(3, i + 2, yr)
style_header_row(ws_emp, 3, YEAR_COUNT + 1)

departments = [
    ("Engineering", 120),
    ("Sales", 80),
    ("Marketing", 40),
    ("Customer Success", 35),
    ("Finance & Accounting", 20),
    ("HR & Admin", 15),
    ("Legal", 8),
    ("Executive", 10),
    ("Data Science", 25),
    ("Product Management", 18),
    ("DevOps / SRE", 15),
    ("Quality Assurance", 20),
]

r = 4
dept_first_row = r
for dept, base_hc in departments:
    ws_emp.cell(r, 1, dept).font = Font(bold=True)
    for i in range(YEAR_COUNT):
        if i == 0:
            c = ws_emp.cell(r, i + 2, base_hc)
            style_input(c)
        else:
            prev = col(i + 1)
            growth_ref = f"Assumptions!{col(i+2)}${named_range_map['Revenue Growth Rate'][0]}"
            formula = f"=ROUND({prev}{r}*(1+{growth_ref}*0.7),0)"
            c = ws_emp.cell(r, i + 2, value=formula)
            style_formula(c)
        c.number_format = '#,##0'
    r += 1

dept_last_row = r - 1
r += 1
ws_emp.cell(r, 1, "Total Headcount").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=SUM({c_letter}{dept_first_row}:{c_letter}{dept_last_row})"
    c = ws_emp.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = '#,##0'
    c.font = Font(bold=True)

r += 1
ws_emp.cell(r, 1, "Revenue per Employee").font = Font(bold=True, color="2F5496")
for i in range(YEAR_COUNT):
    c_letter = col(i + 2)
    formula = f"=Revenue!{c_letter}${total_rev_row}/{c_letter}{r-1}"
    c = ws_emp.cell(r, i + 2, value=formula)
    style_formula(c)
    c.number_format = NUM_FMT

# ─── KPI DASHBOARD ───
ws_kpi = wb.create_sheet("KPIs")
ws_kpi.column_dimensions['A'].width = 30
for i in range(YEAR_COUNT):
    ws_kpi.column_dimensions[col(i + 2)].width = 14

ws_kpi.cell(1, 1, "KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=14)
for i, yr in enumerate(YEARS):
    ws_kpi.cell(3, i + 2, yr)
style_header_row(ws_kpi, 3, YEAR_COUNT + 1)

kpis = [
    ("Revenue", f"=Revenue!{{c}}${total_rev_row}", NUM_FMT),
    ("Revenue Growth YoY", None, PCT_FMT),
    ("Gross Margin", f"='P&L'!{{c}}${pl_rows['Gross Margin']}", PCT_FMT),
    ("EBIT Margin", f"='P&L'!{{c}}${pl_rows['EBIT Margin']}", PCT_FMT),
    ("Net Margin", f"='P&L'!{{c}}${pl_rows['Net Margin']}", PCT_FMT),
    ("EPS", f"='P&L'!{{c}}${pl_rows['EPS']}", MONEY_FMT),
    ("Free Cash Flow", f"=DCF!{{c}}${dcf_rows['FCFF']}", NUM_FMT),
    ("Cash Balance", f"='Balance Sheet'!{{c}}${bs_rows['Cash']}", NUM_FMT),
    ("Total Debt", f"='Balance Sheet'!{{c}}${bs_rows['Debt']}", NUM_FMT),
    ("Debt/Equity", None, '0.00x'),
]

r = 4
kpi_rev_row = r
for label, formula_tpl, fmt in kpis:
    ws_kpi.cell(r, 1, label).font = Font(bold=True)
    for i in range(YEAR_COUNT):
        c_letter = col(i + 2)
        if label == "Revenue Growth YoY":
            if i == 0:
                c = ws_kpi.cell(r, i + 2, "N/A")
            else:
                prev = col(i + 1)
                formula = f"=({c_letter}{kpi_rev_row}-{prev}{kpi_rev_row})/{prev}{kpi_rev_row}"
                c = ws_kpi.cell(r, i + 2, value=formula)
                style_formula(c)
        elif label == "Debt/Equity":
            formula = f"='Balance Sheet'!{c_letter}${bs_rows['Debt']}/'Balance Sheet'!{c_letter}${bs_rows['Total Equity']}"
            c = ws_kpi.cell(r, i + 2, value=formula)
            style_formula(c)
        elif formula_tpl:
            formula = formula_tpl.replace("{c}", c_letter)
            c = ws_kpi.cell(r, i + 2, value=formula)
            style_formula(c)
        c.number_format = fmt
    r += 1

# Add an intentional anomaly — one hardcoded value where formula should be
ws_rev.cell(product_rows["Product A - Enterprise SaaS"], 6, 99999999)
ws_rev.cell(product_rows["Product A - Enterprise SaaS"], 6).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# Save
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "sample_financial_model.xlsx")
wb.save(output_path)
print(f"Sample Excel saved to: {os.path.abspath(output_path)}")
print(f"Sheets: {wb.sheetnames}")
total_cells = sum(ws.max_row * ws.max_column for ws in wb.worksheets)
print(f"Estimated cells: ~{total_cells}")
